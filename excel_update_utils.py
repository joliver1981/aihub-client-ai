# excel_update_utils.py
"""
Excel Update with Change Tracking Utilities

Provides functionality to:
1. Update existing Excel rows by matching key columns
2. Detect and highlight changed values
3. Track new and deleted rows
4. Generate change logs

This extends the existing excel_utils.py functionality with UPDATE operations.
"""

import pandas as pd
import json
import os
import logging
from logging.handlers import WatchedFileHandler
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple, Set
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.comments import Comment
import re
from CommonUtils import rotate_logs_on_startup, get_log_path


# Configure logging
def setup_logging():
    """Configure logging for the workflow execution"""
    logger = logging.getLogger("ExcelUpdate")
    log_level_name = os.getenv('LOG_LEVEL', 'DEBUG')
    log_level = getattr(logging, log_level_name, logging.DEBUG)
    logger.setLevel(log_level)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler = WatchedFileHandler(filename=os.getenv('EXCEL_UPDATE_LOG', get_log_path('excel_update_log.txt')), encoding='utf-8')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    return logger

rotate_logs_on_startup(os.getenv('EXCEL_UPDATE_LOG', get_log_path('excel_update_log.txt')))

logger = setup_logging()

# Default highlight colors
DEFAULT_CHANGE_COLOR = "FFFF00"     # Yellow for changed cells
DEFAULT_NEW_ROW_COLOR = "90EE90"    # Light green for new rows
DEFAULT_DELETED_ROW_COLOR = "FFB6C1" # Light pink for deleted/missing rows
DEFAULT_UNCHANGED_COLOR = None       # No fill for unchanged


class ExcelUpdateConfig:
    """Configuration for Excel update operations"""
    
    def __init__(
        self,
        key_columns: List[str],
        highlight_changes: bool = True,
        change_highlight_color: str = DEFAULT_CHANGE_COLOR,
        new_row_color: str = DEFAULT_NEW_ROW_COLOR,
        deleted_row_color: str = DEFAULT_DELETED_ROW_COLOR,
        track_deleted_rows: bool = True,
        add_new_records: bool = True,  # NEW: Whether to insert records that don't match existing keys
        mark_deleted_as: str = "strikethrough",  # "strikethrough", "color", "comment"
        add_change_timestamp: bool = True,
        timestamp_column: str = "Last Updated",
        change_log_sheet: str = None,  # If set, create a change log sheet
        preserve_formatting: bool = True,
        case_sensitive_keys: bool = False
    ):
        self.key_columns = key_columns
        self.highlight_changes = highlight_changes
        self.change_highlight_color = change_highlight_color.replace("#", "")
        self.new_row_color = new_row_color.replace("#", "")
        self.deleted_row_color = deleted_row_color.replace("#", "")
        self.track_deleted_rows = track_deleted_rows
        self.add_new_records = add_new_records
        self.mark_deleted_as = mark_deleted_as
        self.add_change_timestamp = add_change_timestamp
        self.timestamp_column = timestamp_column
        self.change_log_sheet = change_log_sheet
        self.preserve_formatting = preserve_formatting
        self.case_sensitive_keys = case_sensitive_keys


class ChangeRecord:
    """Represents a single change to a cell"""
    
    def __init__(
        self,
        row_key: str,
        column_name: str,
        old_value: Any,
        new_value: Any,
        change_type: str,  # "modified", "added", "deleted"
        row_number: int = None
    ):
        self.row_key = row_key
        self.column_name = column_name
        self.old_value = old_value
        self.new_value = new_value
        self.change_type = change_type
        self.row_number = row_number
        self.timestamp = datetime.now()
    
    def to_dict(self) -> Dict:
        return {
            "row_key": self.row_key,
            "column": self.column_name,
            "old_value": str(self.old_value) if self.old_value is not None else "",
            "new_value": str(self.new_value) if self.new_value is not None else "",
            "change_type": self.change_type,
            "row_number": self.row_number,
            "timestamp": self.timestamp.isoformat()
        }


class ExcelUpdateResult:
    """Result of an Excel update operation"""
    
    def __init__(self):
        self.success = False
        self.file_path = ""
        self.rows_updated = 0
        self.rows_added = 0
        self.rows_deleted = 0
        self.cells_changed = 0
        self.changes: List[ChangeRecord] = []
        self.errors: List[str] = []
        self.warnings: List[str] = []
    
    def to_dict(self) -> Dict:
        return {
            "success": self.success,
            "file_path": self.file_path,
            "rows_updated": self.rows_updated,
            "rows_added": self.rows_added,
            "rows_deleted": self.rows_deleted,
            "cells_changed": self.cells_changed,
            "total_changes": len(self.changes),
            "changes_summary": [c.to_dict() for c in self.changes[:100]],  # Limit to 100
            "errors": self.errors,
            "warnings": self.warnings
        }


def build_row_key(row_data: Dict, key_columns: List[str], case_sensitive: bool = False) -> str:
    """
    Build a composite key from specified columns.
    
    Args:
        row_data: Dictionary of column_name -> value
        key_columns: List of column names to use as key
        case_sensitive: Whether to use case-sensitive key matching
        
    Returns:
        Composite key string (pipe-separated)
    """
    # Build case-insensitive column lookup map
    row_data_lower = {k.lower(): v for k, v in row_data.items()}
    
    key_parts = []
    for col in key_columns:
        # Try exact match first, then case-insensitive
        if col in row_data:
            value = row_data[col]
        elif col.lower() in row_data_lower:
            value = row_data_lower[col.lower()]
        else:
            value = ""
            
        if value is None:
            value = ""
        else:
            value = str(value).strip()
            if not case_sensitive:
                value = value.lower()
        key_parts.append(value)
    
    return "|".join(key_parts)


def read_excel_to_dict(
    file_path: str,
    sheet_name: str = None,
    key_columns: List[str] = None,
    case_sensitive: bool = False
) -> Tuple[Dict[str, Dict], Dict[str, int], List[str], int]:
    """
    Read an Excel file into a dictionary indexed by key columns.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name (uses active if None)
        key_columns: Columns to use as row key
        case_sensitive: Whether keys are case-sensitive
        
    Returns:
        Tuple of:
        - rows_dict: {key: {column: value}}
        - row_numbers: {key: row_number}
        - column_names: List of column names in order
        - data_start_row: Row number where data starts (after headers)
    """
    logger.info(f"Reading Excel file: {file_path}")
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    wb = load_workbook(file_path, data_only=True)
    
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # Get headers from row 1
    max_col = ws.max_column
    column_names = []
    col_indices = {}
    
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            header = str(header).strip()
            column_names.append(header)
            col_indices[header] = col
    
    logger.debug(f"Found columns: {column_names}")
    
    # Read all data rows
    rows_dict = {}
    row_numbers = {}
    data_start_row = 2
    
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        has_data = False
        
        for col_name, col_idx in col_indices.items():
            cell_value = ws.cell(row=row_num, column=col_idx).value
            row_data[col_name] = cell_value
            if cell_value is not None:
                has_data = True
        
        if has_data and key_columns:
            row_key = build_row_key(row_data, key_columns, case_sensitive)
            if row_key and row_key != "|".join([""] * len(key_columns)):
                rows_dict[row_key] = row_data
                row_numbers[row_key] = row_num
    
    wb.close()
    
    logger.info(f"Read {len(rows_dict)} rows from Excel")
    return rows_dict, row_numbers, column_names, data_start_row


def compare_values(old_value: Any, new_value: Any) -> bool:
    """
    Compare two values, handling type differences and None values.
    
    Returns True if values are DIFFERENT.
    """
    # Handle None/empty equivalence
    if old_value is None and new_value is None:
        return False
    if old_value is None:
        old_value = ""
    if new_value is None:
        new_value = ""
    
    # Convert to strings for comparison
    old_str = str(old_value).strip()
    new_str = str(new_value).strip()
    
    # Handle numeric comparisons
    try:
        old_num = float(old_str) if old_str else None
        new_num = float(new_str) if new_str else None
        if old_num is not None and new_num is not None:
            return abs(old_num - new_num) > 0.0001
    except (ValueError, TypeError):
        pass
    
    return old_str != new_str


def update_excel_with_changes(
    file_path: str,
    new_data: List[Dict],
    config: ExcelUpdateConfig,
    schema: Dict = None,
    sheet_name: str = None,
    output_path: str = None
) -> ExcelUpdateResult:
    """
    Update an existing Excel file with new data, tracking and highlighting changes.
    
    Args:
        file_path: Path to existing Excel file
        new_data: List of dictionaries with new data (each dict is a row)
        config: ExcelUpdateConfig with update settings
        schema: Optional schema for column mapping
        sheet_name: Target sheet name
        output_path: Output path (if different from input)
        
    Returns:
        ExcelUpdateResult with details of all changes
    """
    result = ExcelUpdateResult()
    result.file_path = output_path or file_path
    
    try:
        logger.info(f"Starting Excel update: {file_path}")
        logger.info(f"Key columns: {config.key_columns}")
        logger.info(f"New data rows: {len(new_data)}")
        
        # Validate key columns
        if not config.key_columns:
            raise ValueError("At least one key column must be specified for update operations")
        
        # Read existing data
        existing_rows, existing_row_nums, existing_columns, data_start_row = read_excel_to_dict(
            file_path,
            sheet_name=sheet_name,
            key_columns=config.key_columns,
            case_sensitive=config.case_sensitive_keys
        )
        
        logger.info(f"Existing rows: {len(existing_rows)}")
        
        # Build index of new data
        new_rows = {}
        for row in new_data:
            if isinstance(row, dict):
                row_key = build_row_key(row, config.key_columns, config.case_sensitive_keys)
                if row_key:
                    new_rows[row_key] = row
        
        logger.info(f"New rows indexed: {len(new_rows)}")
        
        # Open workbook for modification
        wb = load_workbook(file_path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Build column index (name -> column number)
        col_index = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                col_index[str(header).strip()] = col
        
        # Check for new columns in new_data that don't exist in Excel
        all_new_columns = set()
        for row in new_data:
            all_new_columns.update(row.keys())
        
        new_columns_to_add = all_new_columns - set(col_index.keys())
        
        # Add timestamp column if configured
        if config.add_change_timestamp and config.timestamp_column not in col_index:
            new_columns_to_add.add(config.timestamp_column)
        
        # Add new columns to Excel
        next_col = ws.max_column + 1
        for new_col_name in sorted(new_columns_to_add):
            if new_col_name not in col_index:
                ws.cell(row=1, column=next_col, value=new_col_name)
                col_index[new_col_name] = next_col
                next_col += 1
                logger.info(f"Added new column: {new_col_name}")
        
        # Prepare highlight fills
        change_fill = PatternFill(start_color=config.change_highlight_color, 
                                   end_color=config.change_highlight_color, 
                                   fill_type="solid") if config.highlight_changes else None
        new_row_fill = PatternFill(start_color=config.new_row_color,
                                    end_color=config.new_row_color,
                                    fill_type="solid") if config.highlight_changes else None
        deleted_fill = PatternFill(start_color=config.deleted_row_color,
                                    end_color=config.deleted_row_color,
                                    fill_type="solid") if config.highlight_changes else None
        
        # Track which existing keys we've seen
        seen_keys = set()
        timestamp_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Track duplicate incoming keys
        incoming_key_counts = {}
        for new_row in new_data:
            row_key = build_row_key(new_row, config.key_columns, config.case_sensitive_keys)
            if row_key:
                incoming_key_counts[row_key] = incoming_key_counts.get(row_key, 0) + 1
        
        # Log duplicate keys (potential data issue)
        duplicate_keys = {k: v for k, v in incoming_key_counts.items() if v > 1}
        if duplicate_keys:
            logger.warning(f"Duplicate keys detected in incoming data: {len(duplicate_keys)} keys appear multiple times")
            for key, count in list(duplicate_keys.items())[:5]:  # Log first 5
                logger.warning(f"  Key '{key}' appears {count} times")
        
        logger.info(f"Processing {len(new_data)} incoming rows, {len(incoming_key_counts)} unique keys")
        
        # Process each row in new data
        for new_row in new_data:
            row_key = build_row_key(new_row, config.key_columns, config.case_sensitive_keys)
            if not row_key:
                continue
            
            seen_keys.add(row_key)
            
            if row_key in existing_rows:
                # UPDATE existing row
                row_num = existing_row_nums[row_key]
                old_row = existing_rows[row_key]
                row_has_changes = False
                
                for col_name, new_value in new_row.items():
                    if col_name not in col_index:
                        continue
                    
                    col_num = col_index[col_name]
                    old_value = old_row.get(col_name)
                    
                    if compare_values(old_value, new_value):
                        # Value changed - update it
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = new_value
                        
                        # Apply highlight
                        if change_fill:
                            cell.fill = change_fill
                        
                        # Record change
                        change = ChangeRecord(
                            row_key=row_key,
                            column_name=col_name,
                            old_value=old_value,
                            new_value=new_value,
                            change_type="modified",
                            row_number=row_num
                        )
                        result.changes.append(change)
                        result.cells_changed += 1
                        row_has_changes = True
                
                if row_has_changes:
                    result.rows_updated += 1
                    
                    # Add timestamp
                    if config.add_change_timestamp and config.timestamp_column in col_index:
                        ts_col = col_index[config.timestamp_column]
                        ws.cell(row=row_num, column=ts_col, value=timestamp_now)
            
            else:
                # INSERT new row (only if add_new_records is enabled)
                if not config.add_new_records:
                    logger.debug(f"Skipping new record (add_new_records=False): {row_key}")
                    continue
                    
                new_row_num = ws.max_row + 1
                
                for col_name, value in new_row.items():
                    if col_name in col_index:
                        cell = ws.cell(row=new_row_num, column=col_index[col_name])
                        cell.value = value
                        
                        # Apply new row highlight
                        if new_row_fill:
                            cell.fill = new_row_fill
                
                # Add timestamp for new row
                if config.add_change_timestamp and config.timestamp_column in col_index:
                    ts_col = col_index[config.timestamp_column]
                    ws.cell(row=new_row_num, column=ts_col, value=timestamp_now)
                
                # Record as added
                change = ChangeRecord(
                    row_key=row_key,
                    column_name="[NEW ROW]",
                    old_value=None,
                    new_value=str(new_row),
                    change_type="added",
                    row_number=new_row_num
                )
                result.changes.append(change)
                result.rows_added += 1
        
        # Handle deleted rows (rows in Excel but not in new data)
        if config.track_deleted_rows:
            deleted_keys = set(existing_rows.keys()) - seen_keys
            
            for deleted_key in deleted_keys:
                row_num = existing_row_nums[deleted_key]
                
                if config.mark_deleted_as == "strikethrough":
                    # Apply strikethrough font
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_num, column=col)
                        if cell.font:
                            cell.font = Font(strike=True, color=cell.font.color)
                        else:
                            cell.font = Font(strike=True)
                
                elif config.mark_deleted_as == "color" and deleted_fill:
                    # Apply deleted color
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_num, column=col).fill = deleted_fill
                
                elif config.mark_deleted_as == "comment":
                    # Add comment to first cell
                    first_col = col_index.get(config.key_columns[0], 1)
                    cell = ws.cell(row=row_num, column=first_col)
                    cell.comment = Comment(
                        f"Row no longer present in source data as of {timestamp_now}",
                        "AI Hub"
                    )
                
                # Record deletion
                change = ChangeRecord(
                    row_key=deleted_key,
                    column_name="[DELETED ROW]",
                    old_value=str(existing_rows[deleted_key]),
                    new_value=None,
                    change_type="deleted",
                    row_number=row_num
                )
                result.changes.append(change)
                result.rows_deleted += 1
        
        # Create change log sheet if configured
        if config.change_log_sheet and result.changes:
            _create_change_log_sheet(wb, config.change_log_sheet, result.changes)
        
        # Save workbook
        save_path = output_path or file_path
        wb.save(save_path)
        wb.close()
        
        result.success = True
        result.file_path = save_path
        
        logger.info(f"Update complete: {result.rows_updated} updated, {result.rows_added} added, "
                   f"{result.rows_deleted} deleted, {result.cells_changed} cells changed")
        
    except Exception as e:
        logger.error(f"Excel update error: {str(e)}", exc_info=True)
        result.success = False
        result.errors.append(str(e))
    
    return result


def _create_change_log_sheet(wb: Workbook, sheet_name: str, changes: List[ChangeRecord]):
    """Create or update a change log sheet in the workbook."""
    
    # Create sheet if it doesn't exist
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find next empty row
        start_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(sheet_name)
        # Add headers
        headers = ["Timestamp", "Row Key", "Column", "Change Type", "Old Value", "New Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        start_row = 2
    
    # Add change records
    for i, change in enumerate(changes):
        row = start_row + i
        ws.cell(row=row, column=1, value=change.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=2, value=change.row_key)
        ws.cell(row=row, column=3, value=change.column_name)
        ws.cell(row=row, column=4, value=change.change_type)
        ws.cell(row=row, column=5, value=str(change.old_value)[:1000] if change.old_value else "")
        ws.cell(row=row, column=6, value=str(change.new_value)[:1000] if change.new_value else "")
        
        # Color code by change type
        if change.change_type == "modified":
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif change.change_type == "added":
            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif change.change_type == "deleted":
            fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        else:
            fill = None
        
        if fill:
            ws.cell(row=row, column=4).fill = fill


# Convenience function for workflow execution
def execute_excel_update(
    input_file: str,
    new_data: List[Dict],
    key_columns: List[str],
    output_file: str = None,
    sheet_name: str = None,
    highlight_changes: bool = True,
    track_deleted: bool = True,
    change_log_sheet: str = None,
    field_mapping: Dict[str, str] = None
) -> Dict:
    """
    High-level function for workflow execution of Excel update.
    
    Args:
        input_file: Path to existing Excel file
        new_data: List of dicts or single dict with new data
        key_columns: Column(s) to match rows
        output_file: Output path (defaults to input_file)
        sheet_name: Target sheet
        highlight_changes: Whether to highlight changed cells
        track_deleted: Whether to mark deleted rows
        change_log_sheet: Optional sheet name for change log
        field_mapping: Optional mapping of data fields to Excel columns
        
    Returns:
        Result dict with success status and change details
    """
    # Ensure new_data is a list
    if isinstance(new_data, dict):
        new_data = [new_data]
    
    # Apply field mapping if provided
    if field_mapping:
        mapped_data = []
        for row in new_data:
            mapped_row = {}
            for data_field, excel_col in field_mapping.items():
                if data_field in row:
                    mapped_row[excel_col] = row[data_field]
            # Also include unmapped fields
            for field, value in row.items():
                if field not in field_mapping:
                    mapped_row[field] = value
            mapped_data.append(mapped_row)
        new_data = mapped_data
    
    config = ExcelUpdateConfig(
        key_columns=key_columns,
        highlight_changes=highlight_changes,
        track_deleted_rows=track_deleted,
        change_log_sheet=change_log_sheet
    )
    
    result = update_excel_with_changes(
        file_path=input_file,
        new_data=new_data,
        config=config,
        sheet_name=sheet_name,
        output_path=output_file
    )
    
    return result.to_dict()
