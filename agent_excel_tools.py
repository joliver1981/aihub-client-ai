"""
Excel Live Data Source Tools for AI Hub Agents

Provides query-time access to Excel files uploaded as agent knowledge.
Auto-activated when an agent has Excel knowledge documents.

Tools:
    get_excel_summary       - Get file structure, columns, types, stats
    read_excel_data         - Query specific data slices with filtering
    aggregate_excel_data    - Run groupby/aggregation operations
    create_excel_chart      - Generate interactive Chart.js charts
    update_excel_data       - Modify original Excel file in-place
    analyze_excel_data      - Natural language analysis via PandasAI
"""

import logging
from logging.handlers import WatchedFileHandler
import os
import json
import time
import pandas as pd
import numpy as np
from typing import Optional, List, Dict, Any
from langchain.tools import tool
from openpyxl import load_workbook
from CommonUtils import get_db_connection, get_app_path, get_log_path
import config as cfg

# ============================================================================
# Logging Setup
# ============================================================================

_log_file = os.getenv('EXCEL_TOOLS_LOG', get_log_path('excel_tools_log.txt'))

logger = logging.getLogger("ExcelTools")
log_level_name = os.getenv('LOG_LEVEL', 'DEBUG')
log_level = getattr(logging, log_level_name, logging.DEBUG)
logger.setLevel(log_level)

if not logger.handlers:
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler = WatchedFileHandler(filename=_log_file, encoding='utf-8')
    handler.setFormatter(formatter)
    logger.addHandler(handler)


# ============================================================================
# Module-Level Helper Functions
# ============================================================================

def generate_excel_metadata(file_path: str) -> dict:
    """
    Generate a lightweight metadata profile for an Excel file.
    Called at ingest time. Stored in Documents.document_metadata as JSON.

    Args:
        file_path: Absolute path to the .xlsx/.xls file

    Returns:
        dict with structure:
        {
            "sheets": [
                {
                    "name": "Sheet1",
                    "row_count": 1500,
                    "column_count": 12,
                    "columns": [
                        {"name": "Region", "dtype": "text", "sample_values": ["North", "South"]},
                        {"name": "Revenue", "dtype": "number", "min": 100.0, "max": 99500.0, "mean": 12340.5},
                        {"name": "Date", "dtype": "date", "min": "2024-01-01", "max": "2024-12-31"}
                    ],
                    "sample_rows": [ {"Region": "North", "Revenue": 5000, ...}, ... ],
                    "has_headers": true
                }
            ],
            "total_rows": 1500,
            "file_size_bytes": 245760
        }
    """
    import openpyxl

    sample_count = int(getattr(cfg, 'EXCEL_METADATA_SAMPLE_ROWS', 5))
    max_stats_cols = int(getattr(cfg, 'EXCEL_METADATA_MAX_STATS_COLUMNS', 50))

    metadata = {
        "sheets": [],
        "total_rows": 0,
        "file_size_bytes": os.path.getsize(file_path) if os.path.exists(file_path) else 0
    }

    try:
        # Get sheet names and visibility
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_sheets = wb.sheetnames
        sheet_states = {name: wb[name].sheet_state for name in all_sheets}
        wb.close()

        for sheet_name in all_sheets:
            if sheet_states.get(sheet_name) != 'visible':
                continue

            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            except Exception as e:
                logger.warning(f"Failed to read sheet '{sheet_name}' for metadata: {e}")
                continue

            if df.empty:
                continue

            sheet_meta = {
                "name": sheet_name,
                "row_count": len(df),
                "column_count": len(df.columns),
                "columns": [],
                "sample_rows": [],
                "has_headers": True
            }

            metadata["total_rows"] += len(df)

            # Profile each column
            for i, col_name in enumerate(df.columns):
                col_info = {"name": str(col_name)}
                series = df[col_name]

                # Detect dtype
                if pd.api.types.is_numeric_dtype(series):
                    col_info["dtype"] = "number"
                    if i < max_stats_cols:
                        col_info["min"] = float(series.min()) if not pd.isna(series.min()) else None
                        col_info["max"] = float(series.max()) if not pd.isna(series.max()) else None
                        col_info["mean"] = round(float(series.mean()), 2) if not pd.isna(series.mean()) else None
                elif pd.api.types.is_datetime64_any_dtype(series):
                    col_info["dtype"] = "date"
                    non_null = series.dropna()
                    if len(non_null) > 0:
                        col_info["min"] = str(non_null.min().date())
                        col_info["max"] = str(non_null.max().date())
                else:
                    col_info["dtype"] = "text"
                    unique_vals = series.dropna().unique()[:5]
                    col_info["sample_values"] = [str(v)[:100] for v in unique_vals]

                sheet_meta["columns"].append(col_info)

            # Sample rows
            sample_df = df.head(sample_count).fillna("")
            for _, row in sample_df.iterrows():
                row_dict = {}
                for col in df.columns:
                    val = row[col]
                    if pd.isna(val):
                        row_dict[str(col)] = None
                    elif isinstance(val, (np.integer,)):
                        row_dict[str(col)] = int(val)
                    elif isinstance(val, (np.floating,)):
                        row_dict[str(col)] = round(float(val), 4)
                    elif isinstance(val, pd.Timestamp):
                        row_dict[str(col)] = str(val.date())
                    else:
                        row_dict[str(col)] = str(val)[:200]
                sheet_meta["sample_rows"].append(row_dict)

            metadata["sheets"].append(sheet_meta)

        return metadata

    except Exception as e:
        logger.error(f"Error generating Excel metadata for '{file_path}': {e}")
        return {"sheets": [], "total_rows": 0, "file_size_bytes": 0, "error": str(e)}


def get_excel_file_path(document_id: str) -> Optional[str]:
    """
    Retrieve the persistent file path for an Excel knowledge document.
    Queries Documents.original_path and verifies the file exists on disk.

    Args:
        document_id: The document_id from the Documents table

    Returns:
        Absolute file path if file exists, None otherwise
    """
    try:
        import pyodbc
        conn = pyodbc.connect(
            f"DRIVER={{SQL Server}};SERVER={cfg.DATABASE_SERVER};"
            f"DATABASE={cfg.DATABASE_NAME};UID={cfg.DATABASE_UID};PWD={cfg.DATABASE_PWD}"
        )
        cursor = conn.cursor()
        cursor.execute("EXEC tenant.sp_setTenantContext ?", os.getenv('API_KEY'))
        cursor.execute(
            "SELECT original_path FROM Documents WHERE document_id = ?",
            document_id
        )
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if row and row[0] and os.path.exists(row[0]):
            return row[0]

        logger.warning(
            f"Excel file not found for document_id={document_id}. "
            f"Path in DB: {row[0] if row else 'NULL'}"
        )
        return None
    except Exception as e:
        logger.error(f"Error getting Excel file path for {document_id}: {e}")
        return None


def get_excel_metadata(document_id: str) -> Optional[dict]:
    """
    Retrieve stored metadata profile from Documents.document_metadata.

    Args:
        document_id: The document_id from the Documents table

    Returns:
        Parsed JSON dict of metadata, or None
    """
    try:
        import pyodbc
        conn = pyodbc.connect(
            f"DRIVER={{SQL Server}};SERVER={cfg.DATABASE_SERVER};"
            f"DATABASE={cfg.DATABASE_NAME};UID={cfg.DATABASE_UID};PWD={cfg.DATABASE_PWD}"
        )
        cursor = conn.cursor()
        cursor.execute("EXEC tenant.sp_setTenantContext ?", os.getenv('API_KEY'))
        cursor.execute(
            "SELECT document_metadata FROM Documents WHERE document_id = ?",
            document_id
        )
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if row and row[0]:
            return json.loads(row[0])
        return None
    except Exception as e:
        logger.error(f"Error getting Excel metadata for {document_id}: {e}")
        return None


def _update_excel_metadata(document_id: str, metadata: dict):
    """Store/update metadata profile in Documents.document_metadata."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC tenant.sp_setTenantContext ?", os.getenv('API_KEY'))
        cursor.execute(
            "UPDATE Documents SET document_metadata = ? WHERE document_id = ?",
            json.dumps(metadata, default=str), document_id
        )
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        logger.warning(f"Failed to update Excel metadata for {document_id}: {e}")


# ============================================================================
# ExcelTool Class
# ============================================================================

class ExcelTool:
    """
    Provides Excel query tools for agents with Excel knowledge documents.

    Pattern follows KnowledgeTool in agent_knowledge_integration.py:
    - __init__ accepts agent context
    - get_tools() returns list of @tool-decorated closure functions
    - get_system_prompt_addition() returns prompt text
    """

    def __init__(self, agent_id: int, excel_docs: List[dict]):
        """
        Args:
            agent_id: The agent ID
            excel_docs: List of knowledge doc dicts that are Excel files.
                Each dict has keys: knowledge_id, agent_id, document_id,
                description, filename, document_type
        """
        self.agent_id = agent_id
        self.excel_docs = excel_docs

        # Pre-load metadata for system prompt generation
        self._metadata_cache = {}
        for doc in excel_docs:
            meta = get_excel_metadata(doc['document_id'])
            if meta:
                self._metadata_cache[doc['document_id']] = meta

        logger.info(
            f"Initialized ExcelTool for agent {agent_id} with "
            f"{len(excel_docs)} Excel document(s), "
            f"{len(self._metadata_cache)} with metadata"
        )

    def get_tools(self) -> list:
        """Return all Excel tools as a list."""
        return [
            self._get_excel_summary_tool(),
            self._get_read_excel_data_tool(),
            self._get_aggregate_excel_data_tool(),
            self._get_create_excel_chart_tool(),
            self._get_update_excel_data_tool(),
            self._get_analyze_excel_data_tool(),
        ]

    def get_system_prompt_addition(self) -> str:
        """Generate system prompt addition describing available Excel data."""
        lines = [
            "",
            "",
            "You have access to Excel data query tools for the following spreadsheet files:"
        ]

        for doc in self.excel_docs:
            meta = self._metadata_cache.get(doc['document_id'])
            if meta:
                sheet_count = len(meta.get('sheets', []))
                total_rows = meta.get('total_rows', '?')
                lines.append(
                    f"- {doc['filename']} (document_id: \"{doc['document_id']}\"): "
                    f"{sheet_count} sheet(s), {total_rows} total rows"
                )
                for sheet in meta.get('sheets', []):
                    col_names = [c['name'] for c in sheet.get('columns', [])[:15]]
                    cols_str = ', '.join(col_names)
                    if len(sheet.get('columns', [])) > 15:
                        cols_str += f', ... (+{len(sheet["columns"]) - 15} more)'
                    lines.append(
                        f"  Sheet \"{sheet['name']}\": {sheet['row_count']} rows, "
                        f"columns: [{cols_str}]"
                    )
            else:
                lines.append(
                    f"- {doc['filename']} (document_id: \"{doc['document_id']}\")"
                )

        lines.extend([
            "",
            "EXCEL TOOL INSTRUCTIONS:",
            "- PREFERRED: Use analyze_excel_data for ANY analytical question about the data",
            "  (counts, totals, averages, comparisons, filtering, rankings, complex analysis).",
            "  This tool accepts natural language questions and returns accurate analytical results.",
            "- Use get_excel_summary FIRST to understand the file structure before querying.",
            "- Use read_excel_data only when you need to display specific raw data rows to the user.",
            "- Use create_excel_chart to generate interactive charts for the user.",
            "- Use update_excel_data to modify cells or add rows in-place.",
            "- Do NOT use aggregate_excel_data for analytical questions; use analyze_excel_data instead.",
            "- These Excel tools are ONLY for spreadsheet (.xlsx/.xls) files. For other document types",
            "  (PDFs, resumes, Word documents, text files, etc.), use the get_user_specific_knowledge tool instead.",
            "- Always pass the document_id as a string parameter to Excel tools."
        ])

        return '\n'.join(lines)

    # ========================================================================
    # Tool 1: get_excel_summary
    # ========================================================================

    def _get_excel_summary_tool(self):
        """Returns @tool for getting Excel file summary/metadata."""

        @tool
        def get_excel_summary(document_id: str) -> str:
            """
            Get a summary of an Excel file's structure and content.
            Returns sheet names, column headers with data types, row counts,
            sample data, and basic statistics for numeric columns.

            ALWAYS call this FIRST before using other Excel tools so you
            understand the file structure, column names, and data types.

            Parameters:
                document_id: The document ID of the Excel file (string)

            Returns:
                Formatted text summary of file structure and content
            """
            try:
                # Try cached metadata first
                meta = get_excel_metadata(document_id)
                if not meta:
                    # Try to regenerate from file
                    file_path = get_excel_file_path(document_id)
                    if not file_path:
                        return (
                            f"Error: Excel file not found for document "
                            f"{document_id}. The file may not be available "
                            f"for live querying."
                        )
                    meta = generate_excel_metadata(file_path)
                    _update_excel_metadata(document_id, meta)

                # Format as readable text
                lines = [f"Excel File Summary (document_id: {document_id})"]
                lines.append(f"Total rows across all sheets: {meta.get('total_rows', 'unknown')}")
                lines.append(f"Number of sheets: {len(meta.get('sheets', []))}")
                lines.append(f"File size: {meta.get('file_size_bytes', 0):,} bytes")
                lines.append("")

                for sheet in meta.get('sheets', []):
                    lines.append(f"=== Sheet: \"{sheet['name']}\" ===")
                    lines.append(f"Rows: {sheet['row_count']}, Columns: {sheet['column_count']}")
                    lines.append("Columns:")

                    for col in sheet.get('columns', []):
                        col_line = f"  - {col['name']} ({col['dtype']})"
                        if col['dtype'] == 'number':
                            stats_parts = []
                            if col.get('min') is not None:
                                stats_parts.append(f"min={col['min']}")
                            if col.get('max') is not None:
                                stats_parts.append(f"max={col['max']}")
                            if col.get('mean') is not None:
                                stats_parts.append(f"mean={col['mean']:.2f}")
                            if stats_parts:
                                col_line += f" [{', '.join(stats_parts)}]"
                        elif col['dtype'] == 'date':
                            if col.get('min') and col.get('max'):
                                col_line += f" [range: {col['min']} to {col['max']}]"
                        elif col.get('sample_values'):
                            samples = col['sample_values'][:3]
                            col_line += f" [examples: {', '.join(str(s) for s in samples)}]"
                        lines.append(col_line)

                    if sheet.get('sample_rows'):
                        lines.append(f"\nSample data (first {len(sheet['sample_rows'])} rows):")
                        cols = [c['name'] for c in sheet.get('columns', [])[:10]]
                        header = " | ".join(str(c) for c in cols)
                        separator = " | ".join("---" for _ in cols)
                        lines.append(header)
                        lines.append(separator)
                        for sample_row in sheet['sample_rows'][:3]:
                            vals = [str(sample_row.get(c, ''))[:30] for c in cols]
                            lines.append(" | ".join(vals))
                    lines.append("")

                return '\n'.join(lines)

            except Exception as e:
                logger.error(f"Error in get_excel_summary: {e}")
                return f"Error getting Excel summary: {str(e)}"

        return get_excel_summary

    # ========================================================================
    # Tool 2: read_excel_data
    # ========================================================================

    def _get_read_excel_data_tool(self):
        """Returns @tool for reading Excel data slices."""

        @tool
        def read_excel_data(
            document_id: str,
            sheet_name: Optional[str] = None,
            columns: Optional[str] = None,
            start_row: Optional[int] = None,
            end_row: Optional[int] = None,
            filter_condition: Optional[str] = None
        ) -> str:
            """
            Read specific data from an Excel file. Returns a markdown table.

            Parameters:
                document_id: The document ID of the Excel file (string)
                sheet_name: Name of the sheet to read (default: first visible sheet)
                columns: Comma-separated column names to include (default: all columns)
                start_row: Starting data row, 1-based (default: 1)
                end_row: Ending data row (default: first 100 rows)
                filter_condition: Pandas query filter string, e.g.:
                    "Region == 'North'"
                    "Revenue > 1000"
                    "Status == 'Active' and Amount >= 500"

            Returns:
                Markdown table with row count summary
            """
            try:
                file_path = get_excel_file_path(document_id)
                if not file_path:
                    return (
                        f"Error: Excel file not found for document "
                        f"{document_id}. Use get_excel_summary to see "
                        f"available documents."
                    )

                max_rows = int(getattr(cfg, 'EXCEL_QUERY_MAX_ROWS', 500))
                default_rows = int(getattr(cfg, 'EXCEL_QUERY_DEFAULT_ROWS', 100))

                # Read the sheet
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name if sheet_name else 0
                )
                total_rows = len(df)

                # Apply column filter
                if columns:
                    col_list = [c.strip() for c in columns.split(',')]
                    missing = [c for c in col_list if c not in df.columns]
                    if missing:
                        return (
                            f"Error: Columns not found: {missing}. "
                            f"Available columns: {list(df.columns)}"
                        )
                    df = df[col_list]

                # Apply pandas query filter
                if filter_condition:
                    try:
                        df = df.query(filter_condition)
                    except Exception as e:
                        return (
                            f"Error applying filter '{filter_condition}': "
                            f"{str(e)}. Use pandas query syntax, e.g. "
                            f"\"Column == 'value'\" or \"Amount > 100\"."
                        )

                filtered_rows = len(df)

                # Apply row range
                actual_start = max(0, (start_row or 1) - 1)
                df = df.iloc[actual_start:]

                if end_row is not None:
                    row_limit = min(end_row - actual_start, max_rows)
                    df = df.head(row_limit)
                else:
                    df = df.head(default_rows)

                result_rows = len(df)

                # Convert to markdown table
                md_table = df.to_markdown(index=False)

                # Build summary line
                summary_parts = [f"Showing {result_rows} of {filtered_rows} rows"]
                if filter_condition:
                    summary_parts.append(f"(filtered from {total_rows} total)")
                else:
                    summary_parts.append(f"(total: {total_rows})")
                if result_rows < filtered_rows:
                    summary_parts.append(
                        "Use start_row/end_row to page through more data."
                    )

                return f"{' '.join(summary_parts)}\n\n{md_table}"

            except Exception as e:
                logger.error(f"Error in read_excel_data: {e}")
                return f"Error reading Excel data: {str(e)}"

        return read_excel_data

    # ========================================================================
    # Tool 3: aggregate_excel_data
    # ========================================================================

    def _get_aggregate_excel_data_tool(self):
        """Returns @tool for running aggregations on Excel data."""

        @tool
        def aggregate_excel_data(
            document_id: str,
            sheet_name: Optional[str] = None,
            group_by: Optional[str] = None,
            aggregations: Optional[str] = None,
            filter_condition: Optional[str] = None
        ) -> str:
            """
            Run aggregation/groupby operations on Excel data.
            Returns a markdown table of results.

            Parameters:
                document_id: The document ID of the Excel file (string)
                sheet_name: Sheet name (default: first sheet)
                group_by: Comma-separated column names to group by,
                    e.g. "Region" or "Region,Product"
                aggregations: JSON string of column:function pairs, e.g.:
                    '{"Revenue": "sum", "Units": "mean"}'
                    '{"Price": ["min", "max"], "Quantity": "count"}'
                    Supported: sum, mean, min, max, count, median, std
                filter_condition: Optional pandas query filter applied
                    before aggregation

            Returns:
                Markdown table of aggregated results
            """
            try:
                file_path = get_excel_file_path(document_id)
                if not file_path:
                    return f"Error: Excel file not found for document {document_id}"

                max_source = int(getattr(cfg, 'EXCEL_AGGREGATION_MAX_ROWS', 50000))

                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name if sheet_name else 0
                )

                if len(df) > max_source:
                    return (
                        f"Error: Sheet has {len(df)} rows, exceeding "
                        f"limit of {max_source} for aggregation."
                    )

                source_rows = len(df)

                # Apply pre-filter
                if filter_condition:
                    try:
                        df = df.query(filter_condition)
                    except Exception as e:
                        return f"Error applying filter: {str(e)}"

                filtered_rows = len(df)

                # Parse aggregations
                if aggregations:
                    try:
                        agg_dict = json.loads(aggregations)
                    except json.JSONDecodeError:
                        return (
                            "Error: 'aggregations' must be valid JSON. "
                            "Example: '{\"Revenue\": \"sum\", \"Units\": \"mean\"}'"
                        )
                    # Validate column names
                    for col_name in agg_dict:
                        if col_name not in df.columns:
                            return (
                                f"Error: Column '{col_name}' not found. "
                                f"Available: {list(df.columns)}"
                            )
                else:
                    # Default: summarize all numeric columns
                    numeric_cols = df.select_dtypes(include='number').columns.tolist()
                    if not numeric_cols:
                        return (
                            "Error: No numeric columns found for default aggregation. "
                            "Specify 'aggregations' parameter."
                        )
                    agg_dict = {col: ['sum', 'mean', 'count'] for col in numeric_cols[:5]}

                # Perform aggregation
                if group_by:
                    group_cols = [c.strip() for c in group_by.split(',')]
                    missing = [c for c in group_cols if c not in df.columns]
                    if missing:
                        return (
                            f"Error: Group-by columns not found: {missing}. "
                            f"Available: {list(df.columns)}"
                        )
                    result_df = df.groupby(group_cols).agg(agg_dict)
                    # Flatten multi-level column names
                    if isinstance(result_df.columns, pd.MultiIndex):
                        result_df.columns = [
                            f"{col[0]}_{col[1]}" if col[1] else col[0]
                            for col in result_df.columns
                        ]
                    result_df = result_df.reset_index()
                else:
                    # Global aggregation (no groupby)
                    result_series = df.agg(agg_dict)
                    if isinstance(result_series, pd.DataFrame):
                        result_df = result_series.T
                        if isinstance(result_df.columns, pd.MultiIndex):
                            result_df.columns = [
                                f"{col[0]}_{col[1]}" if col[1] else col[0]
                                for col in result_df.columns
                            ]
                    else:
                        result_df = pd.DataFrame([result_series])

                md_table = result_df.to_markdown(index=False)

                summary_parts = [f"Aggregation result: {len(result_df)} row(s)"]
                if group_by:
                    summary_parts.append(f"(grouped by: {group_by})")
                summary_parts.append(f"from {filtered_rows} source rows")
                if filter_condition:
                    summary_parts.append(f"(filtered from {source_rows})")

                return f"{' '.join(summary_parts)}\n\n{md_table}"

            except Exception as e:
                logger.error(f"Error in aggregate_excel_data: {e}")
                return f"Error aggregating Excel data: {str(e)}"

        return aggregate_excel_data

    # ========================================================================
    # Tool 4: create_excel_chart
    # ========================================================================

    def _get_create_excel_chart_tool(self):
        """Returns @tool for generating Chart.js charts from Excel data."""

        @tool
        def create_excel_chart(
            document_id: str,
            chart_type: str,
            x_column: str,
            y_columns: str,
            sheet_name: Optional[str] = None,
            title: Optional[str] = None,
            group_by: Optional[str] = None,
            filter_condition: Optional[str] = None,
            aggregation: Optional[str] = None
        ) -> str:
            """
            Generate an interactive Chart.js chart from Excel data.
            The chart is rendered interactively in the chat interface.

            Parameters:
                document_id: The document ID of the Excel file (string)
                chart_type: Chart type - one of: bar, line, pie, scatter, doughnut
                x_column: Column name for X-axis labels
                y_columns: Comma-separated column names for Y-axis data series,
                    e.g. "Revenue" or "Revenue,Costs,Profit"
                sheet_name: Sheet name (default: first sheet)
                title: Chart title (auto-generated if not provided)
                group_by: Column to group by before charting (triggers aggregation)
                filter_condition: Pandas query filter to apply before charting
                aggregation: Aggregation function when using group_by
                    (sum, mean, count, min, max). Default: sum

            Returns:
                Chart data for rendering plus a text description
            """
            try:
                file_path = get_excel_file_path(document_id)
                if not file_path:
                    return f"Error: Excel file not found for document {document_id}"

                max_points = int(getattr(cfg, 'EXCEL_CHART_MAX_DATA_POINTS', 1000))

                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name if sheet_name else 0
                )

                # Apply filter
                if filter_condition:
                    try:
                        df = df.query(filter_condition)
                    except Exception as e:
                        return f"Error applying filter: {str(e)}"

                y_col_list = [c.strip() for c in y_columns.split(',')]

                # Validate columns exist
                all_needed = [x_column] + y_col_list
                if group_by:
                    all_needed.append(group_by)
                missing = [c for c in all_needed if c not in df.columns]
                if missing:
                    return (
                        f"Error: Columns not found: {missing}. "
                        f"Available: {list(df.columns)}"
                    )

                # Validate chart type
                valid_types = ['bar', 'line', 'pie', 'scatter', 'doughnut']
                if chart_type not in valid_types:
                    return f"Error: Invalid chart_type '{chart_type}'. Use one of: {valid_types}"

                # Apply groupby + aggregation if specified
                if group_by:
                    agg_fn = aggregation or 'sum'
                    agg_dict = {col: agg_fn for col in y_col_list}
                    df = df.groupby([x_column]).agg(agg_dict).reset_index()

                # Limit data points
                if len(df) > max_points:
                    df = df.head(max_points)
                    logger.warning(f"Chart data truncated to {max_points} points")

                # Sort by x_column for clean charts
                try:
                    df = df.sort_values(x_column)
                except Exception:
                    pass

                # Build Chart.js data structure
                labels = df[x_column].astype(str).tolist()

                # Color palette
                colors = [
                    'rgba(54, 162, 235, 0.8)',
                    'rgba(255, 99, 132, 0.8)',
                    'rgba(75, 192, 192, 0.8)',
                    'rgba(255, 206, 86, 0.8)',
                    'rgba(153, 102, 255, 0.8)',
                    'rgba(255, 159, 64, 0.8)',
                    'rgba(199, 199, 199, 0.8)',
                    'rgba(83, 102, 255, 0.8)',
                    'rgba(255, 99, 255, 0.8)',
                    'rgba(99, 255, 132, 0.8)',
                ]

                datasets = []
                for i, y_col in enumerate(y_col_list):
                    color = colors[i % len(colors)]
                    border_color = color.replace('0.8', '1')

                    data_values = df[y_col].fillna(0).tolist()
                    # Convert numpy types to native Python for JSON serialization
                    data_values = [
                        float(v) if isinstance(v, (np.integer, np.floating)) else v
                        for v in data_values
                    ]

                    dataset = {
                        "label": y_col,
                        "data": data_values,
                        "borderColor": border_color,
                        "borderWidth": 1
                    }

                    # Set background colors based on chart type
                    if chart_type in ('pie', 'doughnut'):
                        dataset["backgroundColor"] = [
                            colors[j % len(colors)] for j in range(len(data_values))
                        ]
                    elif chart_type == 'bar':
                        dataset["backgroundColor"] = [color] * len(data_values)
                    else:
                        dataset["backgroundColor"] = color

                    if chart_type == 'line':
                        dataset["fill"] = False
                        dataset["tension"] = 0.1

                    datasets.append(dataset)

                chart_content = {
                    "labels": labels,
                    "datasets": datasets
                }

                chart_title = title or f"{', '.join(y_col_list)} by {x_column}"

                # Build the chart block structure
                chart_block = {
                    "type": "chart",
                    "content": chart_content,
                    "metadata": {
                        "chart_type": chart_type,
                        "title": chart_title,
                        "interactive": True,
                        "downloadable": True
                    }
                }

                # Save chart block to side channel for direct injection
                # into the response (bypasses LLM output which strips markers)
                try:
                    from GeneralAgent import _current_agent_context
                    from RichContentManager import rich_content_manager
                    user_id = str(getattr(_current_agent_context, 'user_id', None) or "0")
                    rich_content_manager.save(chart_block, user_id)
                    logger.info(f"Chart block saved to side channel for user {user_id}")
                except Exception as e:
                    logger.warning(f"Could not save chart to side channel: {e}")

                description = (
                    f"Chart generated successfully: {chart_type} chart showing "
                    f"{', '.join(y_col_list)} by {x_column} "
                    f"({len(labels)} data points). "
                    f"The chart is now displayed in the chat."
                )

                return description

            except Exception as e:
                logger.error(f"Error in create_excel_chart: {e}")
                return f"Error creating chart: {str(e)}"

        return create_excel_chart

    # ========================================================================
    # Tool 5: update_excel_data
    # ========================================================================

    def _get_update_excel_data_tool(self):
        """Returns @tool for modifying Excel files in-place."""

        @tool
        def update_excel_data(
            document_id: str,
            sheet_name: Optional[str] = None,
            updates: Optional[str] = None
        ) -> str:
            """
            Modify the original Excel file in-place.
            Changes are saved immediately to the file.

            Parameters:
                document_id: The document ID of the Excel file (string)
                sheet_name: Sheet name to modify (default: active sheet)
                updates: JSON string describing changes. Supported formats:

                  Cell updates (by cell reference):
                    {"cells": {"A1": "new value", "B2": 42, "C3": "2024-01-15"}}

                  Row updates (by 1-based row number, using column names):
                    {"rows": {"2": {"Name": "Updated Name", "Amount": 500},
                              "5": {"Status": "Closed"}}}

                  Add new rows at the end:
                    {"add_rows": [
                        {"Name": "New Entry", "Amount": 1000, "Status": "Open"},
                        {"Name": "Another", "Amount": 2000, "Status": "Open"}
                    ]}

                  Combined (all three in one call):
                    {"cells": {...}, "rows": {...}, "add_rows": [...]}

            Returns:
                Success message with count of changes made
            """
            try:
                file_path = get_excel_file_path(document_id)
                if not file_path:
                    return f"Error: Excel file not found for document {document_id}"

                if not updates:
                    return (
                        "Error: 'updates' parameter is required. "
                        "Provide a JSON string describing changes."
                    )

                try:
                    update_data = json.loads(updates)
                except json.JSONDecodeError as e:
                    return f"Error: 'updates' must be valid JSON. Parse error: {str(e)}"

                wb = load_workbook(file_path)
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active

                changes_made = 0

                # 1. Cell-level updates by reference (e.g., "A1", "B2")
                if 'cells' in update_data:
                    for cell_ref, value in update_data['cells'].items():
                        try:
                            ws[cell_ref] = value
                            changes_made += 1
                        except Exception as e:
                            logger.warning(f"Failed to set cell {cell_ref}: {e}")

                # Build header->column index map (needed for row and add_rows)
                headers = {}
                if 'rows' in update_data or 'add_rows' in update_data:
                    for col_idx in range(1, (ws.max_column or 0) + 1):
                        header_val = ws.cell(row=1, column=col_idx).value
                        if header_val:
                            headers[str(header_val)] = col_idx

                # 2. Row-level updates by row number + column name
                if 'rows' in update_data:
                    for row_num_str, row_data in update_data['rows'].items():
                        try:
                            row_num = int(row_num_str)
                        except ValueError:
                            logger.warning(f"Invalid row number: {row_num_str}")
                            continue
                        for col_name, value in row_data.items():
                            if col_name in headers:
                                ws.cell(
                                    row=row_num,
                                    column=headers[col_name],
                                    value=value
                                )
                                changes_made += 1
                            else:
                                logger.warning(f"Column '{col_name}' not found in headers")

                # 3. Add new rows at the bottom
                if 'add_rows' in update_data:
                    for new_row_data in update_data['add_rows']:
                        next_row = (ws.max_row or 1) + 1
                        for col_name, value in new_row_data.items():
                            if col_name in headers:
                                ws.cell(
                                    row=next_row,
                                    column=headers[col_name],
                                    value=value
                                )
                                changes_made += 1
                            else:
                                logger.warning(
                                    f"Column '{col_name}' not in headers for add_row"
                                )

                # Save the workbook
                wb.save(file_path)
                wb.close()

                # Regenerate metadata profile to reflect changes
                try:
                    new_meta = generate_excel_metadata(file_path)
                    _update_excel_metadata(document_id, new_meta)
                except Exception as e:
                    logger.warning(f"Failed to update metadata after changes: {e}")

                return (
                    f"Successfully applied {changes_made} change(s) to "
                    f"the Excel file (sheet: {ws.title})."
                )

            except Exception as e:
                logger.error(f"Error in update_excel_data: {e}")
                return f"Error updating Excel data: {str(e)}"

        return update_excel_data

    # ========================================================================
    # Tool 6: analyze_excel_data (PandasAI-powered natural language analysis)
    # ========================================================================

    def _get_analyze_excel_data_tool(self):
        """Returns @tool for natural language analysis of Excel data via PandasAI."""

        # Capture metadata cache reference for the closure
        metadata_cache = self._metadata_cache

        @tool
        def analyze_excel_data(
            document_id: str,
            question: str,
            sheet_name: Optional[str] = None
        ) -> str:
            """
            Ask a natural language question about an Excel file and get an
            accurate analytical answer. Uses AI-powered code generation to
            run the correct pandas operations on the data.

            PREFERRED tool for ANY analytical question about Excel data
            including:
            - Counting (how many orders, unique customers, etc.)
            - Aggregations (total revenue, average price, sum of quantities)
            - Filtering (orders over $1000, sales in Q4, by region)
            - Comparisons (which region has highest sales, top N products)
            - Complex analysis (year-over-year growth, rankings, trends)

            Parameters:
                document_id: The document ID of the Excel file (string)
                question: Natural language question about the data
                sheet_name: Optional sheet name (default: first sheet)

            Returns:
                The analytical result as text or a formatted table
            """
            try:
                file_path = get_excel_file_path(document_id)
                if not file_path:
                    return (
                        f"Error: Excel file not found for document "
                        f"{document_id}. Use get_excel_summary to see "
                        f"available documents."
                    )

                # Load the DataFrame as-is (no fragile pre-cleaning)
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name if sheet_name else 0
                )

                if df.empty:
                    return "The spreadsheet sheet is empty — no data to analyze."

                logger.info(
                    f"analyze_excel_data: loaded {len(df)} rows, "
                    f"{len(df.columns)} columns from {file_path}"
                )

                # Build description from cached metadata for better code generation
                description = None
                meta = metadata_cache.get(document_id)
                if meta:
                    desc_parts = []
                    for sheet in meta.get('sheets', []):
                        if sheet_name and sheet['name'] != sheet_name:
                            continue
                        col_descs = []
                        for col in sheet.get('columns', []):
                            col_desc = f"{col['name']} ({col['dtype']})"
                            if col['dtype'] == 'number' and col.get('mean') is not None:
                                col_desc += (
                                    f" [range: {col.get('min')}-{col.get('max')}]"
                                )
                            elif col.get('sample_values'):
                                samples = ', '.join(
                                    str(s) for s in col['sample_values'][:3]
                                )
                                col_desc += f" [examples: {samples}]"
                            col_descs.append(col_desc)
                        desc_parts.append(
                            f"Sheet '{sheet['name']}': "
                            + ', '.join(col_descs)
                        )
                        if not sheet_name:
                            break  # Only describe the first sheet if none specified
                    description = '; '.join(desc_parts) if desc_parts else None

                # Preprocess question with context (follows LLMAnalyticalEngine
                # _preprocess_pandas_input_question pattern)
                processed_question = (
                    question
                    + '\n\nIMPORTANT: Base all time references such as '
                    + '"this year", "last month", etc. on the current date: '
                    + str(time.strftime("%Y-%m-%d"))
                    + '\n\nNote: The spreadsheet data may contain non-data '
                    + 'rows such as title rows, subtotal rows, or grand '
                    + 'total rows. Identify and exclude these from your '
                    + 'analysis. Focus only on actual data records.'
                )

                # Create PandasAI Agent (same config as LLMAnalyticalEngine)
                from pandasai import Agent as PandasAIAgent
                from api_keys_config import create_pandasai_llm

                llm = create_pandasai_llm(use_alternate_api=True)

                agent_kwargs = {
                    "config": {
                        "llm": llm,
                        "enable_cache": False,
                        "open_charts": False,
                    },
                    "memory_size": 5,
                }
                if description:
                    agent_kwargs["description"] = description

                pandas_agent = PandasAIAgent([df], **agent_kwargs)

                logger.info(
                    f"analyze_excel_data: querying PandasAI with: "
                    f"{question[:200]}"
                )

                # Execute the query
                result = pandas_agent.chat(processed_question)

                # Check for PandasAI error response and retry once
                from pandasai.core.response.error import ErrorResponse
                if isinstance(result, ErrorResponse):
                    logger.warning(
                        "PandasAI returned error on first attempt, retrying..."
                    )
                    result = pandas_agent.chat(processed_question)
                    if isinstance(result, ErrorResponse):
                        return (
                            "Unable to analyze the data for this question. "
                            "Try rephrasing your question or use the "
                            "structured Excel tools (read_excel_data, "
                            "aggregate_excel_data) as a fallback."
                        )

                # Format the result
                if isinstance(result, pd.DataFrame):
                    if result.empty:
                        return (
                            "The analysis returned no matching results. "
                            "Try a different question or broader criteria."
                        )
                    total = len(result)
                    if total > 100:
                        truncated = result.head(100)
                        return (
                            f"Result ({total} rows, showing first 100):"
                            f"\n\n{truncated.to_markdown(index=False)}"
                        )
                    return (
                        f"Result ({total} rows):"
                        f"\n\n{result.to_markdown(index=False)}"
                    )
                elif result is None:
                    return (
                        "The analysis completed but returned no result. "
                        "Try rephrasing the question."
                    )
                else:
                    return str(result)

            except Exception as e:
                logger.error(f"Error in analyze_excel_data: {e}")
                return f"Error analyzing Excel data: {str(e)}"

        return analyze_excel_data
