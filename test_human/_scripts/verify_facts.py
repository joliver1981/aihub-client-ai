"""Compute fingerprint facts from generated fixtures so the test plans match exactly."""
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent

print("=" * 60)
print("F1 — Monthly sales by region")
print("=" * 60)
wb = openpyxl.load_workbook(ROOT / "01_Finance" / "fixtures" / "F1_monthly_sales_by_region.xlsx")
ws = wb.active
# Headers row 4, data starts row 5, ends row 64 (60 rows)
data = []
for r in range(5, 65):
    region = ws.cell(r, 1).value
    channel = ws.cell(r, 2).value
    fam = ws.cell(r, 3).value
    units = ws.cell(r, 4).value
    rev = ws.cell(r, 5).value
    if region:
        data.append((region, channel, fam, units, rev))

print(f"Rows: {len(data)}")
print(f"Grand total: ${sum(r[4] for r in data):,.2f}")
# By region
by_region = {}
for r in data:
    by_region.setdefault(r[0], 0)
    by_region[r[0]] += r[4]
print("By region:")
for k, v in sorted(by_region.items(), key=lambda x: -x[1]):
    print(f"  {k}: ${v:,.2f}")
# By channel
by_channel = {}
for r in data:
    by_channel.setdefault(r[1], 0)
    by_channel[r[1]] += r[4]
print("By channel:")
for k, v in sorted(by_channel.items(), key=lambda x: -x[1]):
    print(f"  {k}: ${v:,.2f}")
# By family - units
by_fam_units = {}
for r in data:
    by_fam_units.setdefault(r[2], 0)
    by_fam_units[r[2]] += r[3]
print("By family (units):")
for k, v in sorted(by_fam_units.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v:,}")
# Highest single row
top = max(data, key=lambda r: r[4])
print(f"Highest single row: {top}")

print()
print("=" * 60)
print("I1 — IT Asset inventory")
print("=" * 60)
wb = openpyxl.load_workbook(ROOT / "03_IT" / "fixtures" / "I1_asset_inventory.xlsx")
ws = wb.active
data = []
for r in range(5, 75):
    aid = ws.cell(r, 1).value
    atype = ws.cell(r, 2).value
    site = ws.cell(r, 4).value
    pd = ws.cell(r, 5).value
    cost = ws.cell(r, 7).value
    if aid and not str(aid).startswith("TOTAL"):
        data.append((aid, atype, site, pd, cost))
print(f"Rows: {len(data)}")
print(f"Total cost: ${sum(r[4] for r in data):,}")
by_site = {}
for r in data:
    by_site.setdefault(r[2], 0)
    by_site[r[2]] += 1
print("By site:")
for k, v in sorted(by_site.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")
by_type = {}
for r in data:
    by_type.setdefault(r[1], 0)
    by_type[r[1]] += 1
print("By type:")
for k, v in sorted(by_type.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")
# Oldest
oldest = min(data, key=lambda r: r[3])
print(f"Oldest: {oldest}")

print()
print("=" * 60)
print("P1 — Demand forecast 12-month")
print("=" * 60)
wb = openpyxl.load_workbook(ROOT / "04_Planning" / "fixtures" / "P1_demand_forecast.xlsx")
ws = wb.active
months = ["Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct"]
data = []
for r in range(5, 25):
    fam = ws.cell(r, 1).value
    reg = ws.cell(r, 2).value
    monthly = [ws.cell(r, 3 + i).value for i in range(12)]
    if fam and reg:
        data.append((fam, reg, monthly, sum(monthly)))
print(f"Rows (family × region combos): {len(data)}")
print(f"Grand total 12-mo: {sum(r[3] for r in data):,}")
# Peak month for Tents
tents_rows = [r for r in data if r[0] == "Tents"]
month_totals = [sum(r[2][i] for r in tents_rows) for i in range(12)]
peak_idx = month_totals.index(max(month_totals))
print(f"Tents peak month: {months[peak_idx]} with {month_totals[peak_idx]:,} units")
# Lowest month overall
all_month = [sum(r[2][i] for r in data) for i in range(12)]
low_idx = all_month.index(min(all_month))
print(f"Lowest month overall: {months[low_idx]} with {all_month[low_idx]:,} units")
# Highest region
by_reg = {}
for r in data:
    by_reg.setdefault(r[1], 0)
    by_reg[r[1]] += r[3]
print("By region (12-mo):")
for k, v in sorted(by_reg.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v:,}")

print()
print("=" * 60)
print("O1 — Inventory turnover (Q3 sheet)")
print("=" * 60)
wb = openpyxl.load_workbook(ROOT / "02_Operations" / "fixtures" / "O1_inventory_turnover.xlsx")
ws = wb["Q3_FY25"]
# Three blocks per quarter, each block has Warehouse header then SKU rows
# Just compute total stock value across all rows that have a $ in col 4
total_stock = 0
rows_seen = 0
hot = []
slow = []
for r in range(1, ws.max_row + 1):
    sku = ws.cell(r, 1).value
    desc = ws.cell(r, 2).value
    turn = ws.cell(r, 3).value
    val = ws.cell(r, 4).value
    if sku and str(sku).startswith("NW") is False and turn is not None and val is not None and isinstance(turn, (int, float)) and isinstance(val, (int, float)):
        # Skip subtotal rows (their col 1 is "Subtotal — ...")
        if isinstance(sku, str) and sku.startswith("Subtotal"):
            continue
        if isinstance(sku, str) and sku.startswith("Warehouse"):
            continue
        rows_seen += 1
        total_stock += val
        if turn >= 7.0:
            hot.append((sku, desc, turn, val))
        if turn < 2.0:
            slow.append((sku, desc, turn, val))
print(f"Q3 SKU rows: {rows_seen}")
print(f"Q3 total stock value: ${total_stock:,}")
print(f"Hot SKUs (turn ≥ 7): {len(hot)}")
print(f"Slow movers (turn < 2): {len(slow)}")
# Highest turn in Q3
all_q3 = []
for r in range(1, ws.max_row + 1):
    sku = ws.cell(r, 1).value
    turn = ws.cell(r, 3).value
    val = ws.cell(r, 4).value
    if isinstance(sku, str) and sku.startswith("NW") is False and isinstance(turn, (int, float)) and isinstance(val, (int, float)):
        if not sku.startswith(("Subtotal", "Warehouse")):
            all_q3.append((sku, turn, val))
all_q3.sort(key=lambda x: -x[1])
print(f"Highest turn Q3: {all_q3[0]}")
all_q3.sort(key=lambda x: x[1])
print(f"Lowest turn Q3: {all_q3[0]}")
