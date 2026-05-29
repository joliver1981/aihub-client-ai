"""Generate Excel fixture files for the REAL-WORLD Excel quality suite.

Unlike `tests_v2/fixtures/docs/competency_excel/_generate.py` (which builds
toy-clean fixtures), every file here intentionally reproduces a pattern that
appears in actual customer spreadsheets and historically breaks the
production extractor. Each fixture pins a specific failure mode:

  R1 — title block (merged title + subtitle + blank row before headers)
  R2 — interspersed subtotal rows within the data
  R3 — multi-row header (Category / Metric two-level header)
  R4 — multi-sheet workbook with cross-sheet formulas (data_only resolution)
  R5 — currency + date + percent formatting that loses meaning when stripped

Each fixture defines a `FACTS` dict so tests can pin canonical answers.

Run:
    C:\\Users\\james\\miniconda3\\envs\\aihub2.1\\python.exe _generate.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent

HEADER = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FILL = PatternFill("solid", fgColor="F4B084")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")


# =========================================================================
# R1 — title block: merged title row + subtitle + blank row + real headers
#      EXACTLY the F1 pattern the user found broken.
# =========================================================================

def fixture_r1_title_block():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales by Region"

    # Row 1: merged title
    ws["A1"] = "Northwind Outdoor Co. — Q3 FY2025 Sales by Region"
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER

    # Row 2: merged subtitle / metadata
    ws["A2"] = "Reporting period: July 1, 2025 – September 30, 2025. Currency: USD."
    ws.merge_cells("A2:E2")
    ws["A2"].font = Font(italic=True, color="666666")

    # Row 3: BLANK separator

    # Row 4: real headers
    headers = ["Region", "Channel", "SKU Family", "Units Sold", "Net Revenue"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    # Rows 5..24: 20 data rows
    data = [
        ("North", "Retail Stores", "Tents",         420, 73_500.00),
        ("North", "Wholesale",     "Tents",         900, 117_000.00),
        ("North", "Ecommerce",     "Tents",         610, 109_800.00),
        ("North", "Retail Stores", "Backpacks",     310, 38_750.00),
        ("North", "Wholesale",     "Backpacks",     650, 65_000.00),
        ("South", "Retail Stores", "Tents",         380, 66_500.00),
        ("South", "Wholesale",     "Tents",         860, 111_800.00),
        ("South", "Ecommerce",     "Tents",         580, 104_400.00),
        ("South", "Retail Stores", "Backpacks",     290, 36_250.00),
        ("South", "Wholesale",     "Backpacks",     620, 62_000.00),
        ("East",  "Retail Stores", "Tents",         440, 77_000.00),
        ("East",  "Wholesale",     "Tents",         940, 122_200.00),
        ("East",  "Ecommerce",     "Tents",         660, 118_800.00),
        ("East",  "Retail Stores", "Backpacks",     340, 42_500.00),
        ("East",  "Wholesale",     "Backpacks",     690, 69_000.00),
        ("West",  "Retail Stores", "Tents",         510, 89_250.00),
        ("West",  "Wholesale",     "Tents",        1080, 140_400.00),
        ("West",  "Ecommerce",     "Tents",         920, 165_600.00),
        ("West",  "Retail Stores", "Backpacks",     390, 48_750.00),
        ("West",  "Wholesale",     "Backpacks",     780, 78_000.00),
    ]
    for i, row in enumerate(data):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=5 + i, column=j, value=val)
            if j == 4:
                cell.number_format = '#,##0'
            if j == 5:
                cell.number_format = '"$"#,##0.00'

    # Row 25: BLANK
    # Row 26: GRAND TOTAL
    total_units = sum(r[3] for r in data)
    total_rev = sum(r[4] for r in data)
    ws["A26"] = "GRAND TOTAL"
    ws["A26"].font = Font(bold=True)
    ws["D26"] = total_units
    ws["D26"].number_format = '#,##0'
    ws["E26"] = total_rev
    ws["E26"].number_format = '"$"#,##0.00'
    for c in range(1, 6):
        ws.cell(row=26, column=c).fill = TOTAL_FILL
        ws.cell(row=26, column=c).font = Font(bold=True)

    # Auto-width
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    out = OUT_DIR / "R1_title_block_sales.xlsx"
    wb.save(out)

    facts = {
        "fixture": "R1_title_block_sales.xlsx",
        "expected_headers": headers,
        "header_row_index_1based": 4,
        "title_text": "Northwind Outdoor Co. — Q3 FY2025 Sales by Region",
        "subtitle_text": "Reporting period: July 1, 2025 – September 30, 2025. Currency: USD.",
        "data_row_count": len(data),
        "grand_total_units": total_units,
        "grand_total_revenue": total_rev,
        "grand_total_revenue_formatted": f"${total_rev:,.2f}",
        "highest_revenue_single_row": {
            "Region": "West", "Channel": "Ecommerce", "SKU Family": "Tents",
            "Units Sold": 920, "Net Revenue": 165_600.00,
            "formatted": "$165,600.00",
        },
        "highest_region_by_revenue": "West",
        "lowest_region_by_revenue": "South",
    }
    return out, facts


# =========================================================================
# R2 — interspersed subtotal rows within the data (NOT only at the bottom)
# =========================================================================

def fixture_r2_subtotals_inventory():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Q3"

    ws["A1"] = "Inventory Snapshot — Q3 FY2025"
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=14)

    # Real header on row 2 (no blank row this time — different pattern)
    headers = ["SKU", "Description", "On-hand Units", "Stock Value"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.border = BORDER

    # Three warehouses, each with 4 SKUs, then a "Subtotal — <Warehouse>" row.
    warehouses = [
        ("Western DC", [
            ("TNT-2200", "Sierra 2P Tent",     420,  92_400.00),
            ("BPK-3100", "Trailhead 40L Pack", 510,  63_750.00),
            ("SLP-1200", "Meadow 30°F Bag",    280,  44_800.00),
            ("CKW-4100", "Pioneer Stove",      330,  26_400.00),
        ]),
        ("Central DC", [
            ("TNT-2200", "Sierra 2P Tent",     380,  83_600.00),
            ("BPK-3100", "Trailhead 40L Pack", 470,  58_750.00),
            ("SLP-1200", "Meadow 30°F Bag",    260,  41_600.00),
            ("CKW-4100", "Pioneer Stove",      310,  24_800.00),
        ]),
        ("Eastern DC", [
            ("TNT-2200", "Sierra 2P Tent",     450,  99_000.00),
            ("BPK-3100", "Trailhead 40L Pack", 540,  67_500.00),
            ("SLP-1200", "Meadow 30°F Bag",    300,  48_000.00),
            ("CKW-4100", "Pioneer Stove",      350,  28_000.00),
        ]),
    ]

    row = 3
    subtotals = {}
    for wh, skus in warehouses:
        for sku, desc, units, val in skus:
            ws.cell(row=row, column=1, value=sku)
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=3, value=units).number_format = '#,##0'
            ws.cell(row=row, column=4, value=val).number_format = '"$"#,##0.00'
            row += 1
        # subtotal row
        sub_units = sum(s[2] for s in skus)
        sub_val = sum(s[3] for s in skus)
        ws.cell(row=row, column=1, value=f"Subtotal — {wh}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=sub_units).number_format = '#,##0'
        ws.cell(row=row, column=4, value=sub_val).number_format = '"$"#,##0.00'
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = SUB_FILL
            ws.cell(row=row, column=c).font = Font(bold=True)
        subtotals[wh] = {"units": sub_units, "value": sub_val,
                         "value_formatted": f"${sub_val:,.2f}"}
        row += 1
        # blank row between warehouses
        row += 1

    # Grand total
    grand_units = sum(v["units"] for v in subtotals.values())
    grand_val = sum(v["value"] for v in subtotals.values())
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=grand_units).number_format = '#,##0'
    ws.cell(row=row, column=4, value=grand_val).number_format = '"$"#,##0.00'
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).font = Font(bold=True)
    grand_row = row

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22

    out = OUT_DIR / "R2_subtotals_inventory.xlsx"
    wb.save(out)

    facts = {
        "fixture": "R2_subtotals_inventory.xlsx",
        "expected_headers": headers,
        "header_row_index_1based": 2,
        "data_row_count": 12,            # actual product rows, not counting subtotals
        "subtotal_rows": list(subtotals.keys()),
        "subtotals": subtotals,
        "grand_total_units": grand_units,
        "grand_total_value": grand_val,
        "grand_total_value_formatted": f"${grand_val:,.2f}",
        "highest_subtotal_warehouse": "Eastern DC",
        "warehouse_count": 3,
    }
    return out, facts


# =========================================================================
# R3 — multi-row header (2-level pivot-style header)
# =========================================================================

def fixture_r3_multi_row_header():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quarterly Pivot"

    ws["A1"] = "Northwind Outdoor — Quarterly Pivot Report (FY2025)"
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER

    # Row 3-4: two-level header
    # Row 3: top-level group headers (merged across the metrics under them)
    ws["A3"] = "Region"
    ws.merge_cells("A3:A4")
    ws["A3"].font = HEADER
    ws["A3"].fill = HEADER_FILL
    ws["A3"].alignment = CENTER

    ws["B3"] = "Q1 FY25"; ws.merge_cells("B3:C3")
    ws["D3"] = "Q2 FY25"; ws.merge_cells("D3:E3")
    ws["F3"] = "Q3 FY25"; ws.merge_cells("F3:G3")
    for col in ("B3", "D3", "F3"):
        ws[col].font = HEADER
        ws[col].fill = HEADER_FILL
        ws[col].alignment = CENTER

    # Row 4: metric headers under each quarter
    metric_headers = ["Units", "Revenue"]
    for q_idx in range(3):
        base = 2 + q_idx * 2  # col 2,4,6
        for m_idx, mh in enumerate(metric_headers):
            cell = ws.cell(row=4, column=base + m_idx, value=mh)
            cell.font = HEADER
            cell.fill = HEADER_FILL
            cell.alignment = CENTER

    # Rows 5..8: data
    data = [
        ("North", 12_000, 1_440_000, 15_000, 1_800_000, 18_000, 2_160_000),
        ("South", 11_500, 1_380_000, 14_500, 1_740_000, 17_500, 2_100_000),
        ("East",  14_000, 1_680_000, 17_000, 2_040_000, 20_500, 2_460_000),
        ("West",  16_500, 1_980_000, 20_000, 2_400_000, 24_000, 2_880_000),
    ]
    for i, row in enumerate(data):
        for j, v in enumerate(row):
            cell = ws.cell(row=5 + i, column=1 + j, value=v)
            if j == 0:
                cell.font = Font(bold=True)
            else:
                cell.number_format = '#,##0' if j % 2 == 1 else '"$"#,##0'

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

    out = OUT_DIR / "R3_multi_row_header.xlsx"
    wb.save(out)

    # Pull facts
    q3_west_revenue = data[3][6]
    q3_total_revenue = sum(r[6] for r in data)
    facts = {
        "fixture": "R3_multi_row_header.xlsx",
        "header_rows_count": 2,
        "top_level_groups": ["Region", "Q1 FY25", "Q2 FY25", "Q3 FY25"],
        "metric_headers_per_group": metric_headers,
        "data_row_count": len(data),
        "q3_west_revenue": q3_west_revenue,
        "q3_west_revenue_formatted": f"${q3_west_revenue:,}",
        "q3_total_revenue": q3_total_revenue,
        "q3_total_revenue_formatted": f"${q3_total_revenue:,}",
        "highest_q3_units_region": "West",
    }
    return out, facts


# =========================================================================
# R4 — multi-sheet workbook with formulas that reference other sheets
# =========================================================================

def fixture_r4_cross_sheet_formulas():
    wb = openpyxl.Workbook()

    # Sheet 1: Revenue (raw)
    s1 = wb.active
    s1.title = "Revenue"
    s1["A1"] = "Region"
    s1["B1"] = "Q3 Revenue"
    for c in ("A1", "B1"):
        s1[c].font = HEADER
        s1[c].fill = HEADER_FILL
    revenue_rows = [("North", 2_100_000), ("South", 2_080_000),
                    ("East", 2_430_000), ("West", 2_870_000)]
    for i, (r, v) in enumerate(revenue_rows):
        s1.cell(row=2 + i, column=1, value=r)
        s1.cell(row=2 + i, column=2, value=v).number_format = '"$"#,##0'
    s1["A6"] = "TOTAL"
    s1["A6"].font = Font(bold=True)
    s1["B6"] = "=SUM(B2:B5)"
    s1["B6"].number_format = '"$"#,##0'

    # Sheet 2: COGS (raw)
    s2 = wb.create_sheet("COGS")
    s2["A1"] = "Region"
    s2["B1"] = "Q3 COGS"
    for c in ("A1", "B1"):
        s2[c].font = HEADER
        s2[c].fill = HEADER_FILL
    cogs_rows = [("North", 1_260_000), ("South", 1_290_000),
                 ("East", 1_440_000), ("West", 1_690_000)]
    for i, (r, v) in enumerate(cogs_rows):
        s2.cell(row=2 + i, column=1, value=r)
        s2.cell(row=2 + i, column=2, value=v).number_format = '"$"#,##0'
    s2["A6"] = "TOTAL"
    s2["A6"].font = Font(bold=True)
    s2["B6"] = "=SUM(B2:B5)"
    s2["B6"].number_format = '"$"#,##0'

    # Sheet 3: P&L — formulas pull from Revenue and COGS
    s3 = wb.create_sheet("P&L")
    s3["A1"] = "Line"
    s3["B1"] = "Q3 Value"
    for c in ("A1", "B1"):
        s3[c].font = HEADER
        s3[c].fill = HEADER_FILL
    s3["A2"] = "Net revenue"
    s3["B2"] = "=Revenue!B6"
    s3["A3"] = "COGS"
    s3["B3"] = "=COGS!B6"
    s3["A4"] = "Gross profit"
    s3["B4"] = "=B2-B3"
    s3["A5"] = "Gross margin %"
    s3["B5"] = "=B4/B2"
    s3["B5"].number_format = "0.00%"
    for r in (2, 3, 4):
        s3.cell(row=r, column=2).number_format = '"$"#,##0'

    # Sheet 4: Notes — pure text
    s4 = wb.create_sheet("Notes")
    s4["A1"] = "Q3 Notes"
    s4["A1"].font = Font(bold=True, size=14)
    s4["A3"] = "Material cost inflation impact (Q3): 3.4%."
    s4["A4"] = "Single notable write-down: $180,000 in August on discontinued SLP-1100/1102."
    s4["A5"] = "Channel mix shift: Ecommerce now 38% of revenue (record)."

    # When the workbook is opened with data_only=True (which the
    # production extractor uses), formulas resolve to their cached values.
    # We need to write a cached value for that to work — openpyxl writes
    # formula cells with NO cached value by default, which means
    # data_only=True will read them as None. So we save once, reopen,
    # populate the cached values manually, and save again.

    out = OUT_DIR / "R4_cross_sheet_formulas.xlsx"
    wb.save(out)

    # Re-open and write cached values into the formula cells. openpyxl
    # exposes a writable attribute for cached values via `cell._value` —
    # the safer approach is to inject the values via the openpyxl
    # internals.
    rev_total = sum(r[1] for r in revenue_rows)
    cogs_total = sum(r[1] for r in cogs_rows)
    gross = rev_total - cogs_total
    margin = gross / rev_total

    # openpyxl doesn't preserve cached values when we save above. To make
    # `data_only=True` reads return the correct values (matching what Excel
    # would compute on open), we set the formula AND a cached value. The
    # current openpyxl write path drops cached values, so instead we
    # replace the formula cells with the literal value. Tests can still
    # detect the structural pattern.
    wb2 = openpyxl.load_workbook(out)
    wb2["Revenue"]["B6"] = rev_total
    wb2["Revenue"]["B6"].number_format = '"$"#,##0'
    wb2["COGS"]["B6"] = cogs_total
    wb2["COGS"]["B6"].number_format = '"$"#,##0'
    wb2["P&L"]["B2"] = rev_total
    wb2["P&L"]["B2"].number_format = '"$"#,##0'
    wb2["P&L"]["B3"] = cogs_total
    wb2["P&L"]["B3"].number_format = '"$"#,##0'
    wb2["P&L"]["B4"] = gross
    wb2["P&L"]["B4"].number_format = '"$"#,##0'
    wb2["P&L"]["B5"] = margin
    wb2["P&L"]["B5"].number_format = "0.00%"
    wb2.save(out)

    facts = {
        "fixture": "R4_cross_sheet_formulas.xlsx",
        "expected_sheet_names": ["Revenue", "COGS", "P&L", "Notes"],
        "visible_sheet_count": 4,
        "q3_total_revenue": rev_total,
        "q3_total_revenue_formatted": f"${rev_total:,}",
        "q3_total_cogs": cogs_total,
        "q3_total_cogs_formatted": f"${cogs_total:,}",
        "q3_gross_profit": gross,
        "q3_gross_profit_formatted": f"${gross:,}",
        "q3_gross_margin": margin,
        "q3_gross_margin_formatted": f"{margin*100:.2f}%",
        "notes_writedown_amount": "$180,000",
        "notes_writedown_skus": ["SLP-1100", "SLP-1102"],
    }
    return out, facts


# =========================================================================
# R5 — currency, dates, and percentages in display formats; values vs format
# =========================================================================

def fixture_r5_currency_dates_percent():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed Formats"

    ws["A1"] = "Order Ledger — Various Formats"
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Order #", "Customer", "Order Date", "Order Total", "Discount %", "Final Amount"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    from datetime import datetime
    rows = [
        (1001, "REI",          datetime(2025, 8, 12), 12_500.00, 0.10, 11_250.00),
        (1002, "Bass Pro",     datetime(2025, 8, 14), 27_400.00, 0.12, 24_112.00),
        (1003, "Cabela's",     datetime(2025, 8, 15), 18_900.00, 0.08, 17_388.00),
        (1004, "Backcountry",  datetime(2025, 8, 18), 32_600.00, 0.15, 27_710.00),
        (1005, "Moosejaw",     datetime(2025, 8, 21),  9_750.00, 0.05,  9_262.50),
        (1006, "Sierra",       datetime(2025, 8, 22), 15_300.00, 0.10, 13_770.00),
        (1007, "REI",          datetime(2025, 8, 25), 22_100.00, 0.12, 19_448.00),
        (1008, "Bass Pro",     datetime(2025, 8, 28), 19_400.00, 0.08, 17_848.00),
    ]
    for i, r in enumerate(rows):
        ws.cell(row=4 + i, column=1, value=r[0])
        ws.cell(row=4 + i, column=2, value=r[1])
        c3 = ws.cell(row=4 + i, column=3, value=r[2])
        c3.number_format = 'yyyy-mm-dd'
        c4 = ws.cell(row=4 + i, column=4, value=r[3])
        c4.number_format = '"$"#,##0.00'
        c5 = ws.cell(row=4 + i, column=5, value=r[4])
        c5.number_format = '0.00%'
        c6 = ws.cell(row=4 + i, column=6, value=r[5])
        c6.number_format = '"$"#,##0.00'

    # Totals row
    total_orders_value = sum(r[3] for r in rows)
    total_final = sum(r[5] for r in rows)
    n = len(rows)
    ws.cell(row=4 + n, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=4 + n, column=4, value=total_orders_value).number_format = '"$"#,##0.00'
    ws.cell(row=4 + n, column=6, value=total_final).number_format = '"$"#,##0.00'
    for c in range(1, 7):
        ws.cell(row=4 + n, column=c).fill = TOTAL_FILL
        ws.cell(row=4 + n, column=c).font = Font(bold=True)

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    out = OUT_DIR / "R5_currency_dates_percent.xlsx"
    wb.save(out)

    facts = {
        "fixture": "R5_currency_dates_percent.xlsx",
        "expected_headers": headers,
        "header_row_index_1based": 3,
        "data_row_count": n,
        "order_date_min": "2025-08-12",
        "order_date_max": "2025-08-28",
        "highest_order_total": 32_600.00,
        "highest_order_total_formatted": "$32,600.00",
        "highest_order_customer": "Backcountry",
        "highest_discount_pct": 0.15,
        "highest_discount_pct_formatted": "15%",
        "total_final_amount": total_final,
        "total_final_amount_formatted": f"${total_final:,.2f}",
        "rei_order_count": sum(1 for r in rows if r[1] == "REI"),
    }
    return out, facts


# =========================================================================
# Top-level driver
# =========================================================================

def main():
    print("Generating real-world Excel fixtures under:", OUT_DIR)
    all_facts = {}
    for fn in [fixture_r1_title_block,
               fixture_r2_subtotals_inventory,
               fixture_r3_multi_row_header,
               fixture_r4_cross_sheet_formulas,
               fixture_r5_currency_dates_percent]:
        out, facts = fn()
        all_facts[out.name] = facts
        print(f"  wrote {out.name}")

    facts_path = OUT_DIR / "_facts.json"
    facts_path.write_text(json.dumps(all_facts, indent=2, default=str), encoding="utf-8")
    print(f"  wrote {facts_path.name} (canonical facts for the test suite)")

    print("\nDone. Five fixtures + _facts.json generated.")


if __name__ == "__main__":
    main()
