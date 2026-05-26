"""Generate Excel fixture files for the agent-knowledge Excel COMPETENCY suite.

Each file is a different "shape" of spreadsheet to probe a different
extraction / retrieval / reasoning weakness. Every fixture has at least
one **fingerprint** value — a string that no other fixture has — so the
test can reliably ask "what is X?" and score whether the agent surfaced
the right number / name.

Run:
    C:\\Users\\james\\miniconda3\\envs\\aihub2.1\\python.exe _generate.py
"""
from __future__ import annotations

import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUT_DIR = Path(__file__).resolve().parent
random.seed(42)  # deterministic fixtures

# Reusable styles ----------------------------------------------------------

HEADER = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUBHDR = Font(bold=True, size=10)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
MONEY_FMT = '"$"#,##0.00'


# =========================================================================
# 1. CLEAN EMPLOYEES — baseline single-sheet HR table
# =========================================================================

def fixture_clean_employees():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = ["EmployeeId", "FirstName", "LastName", "Department",
               "Title", "AnnualSalary", "HireDate", "City"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.border = BORDER

    # 30 employees, deterministic
    rows = [
        (1001, "Lena",  "Ashford",   "Engineering", "Principal Engineer",   192000, "2017-03-14", "Boulder"),
        (1002, "Marcus","Okonkwo",   "Engineering", "Senior Engineer",      148000, "2019-08-22", "Boulder"),
        (1003, "Priya", "Mehta",     "Engineering", "Engineering Manager",  178500, "2016-11-02", "Boulder"),
        (1004, "Ivan",  "Strelnikov","Engineering", "Staff Engineer",       163000, "2018-04-19", "Boulder"),
        (1005, "Greta", "Knudsen",   "Engineering", "Senior Engineer",      151000, "2020-01-13", "Boulder"),
        (1006, "Aki",   "Tanaka",    "Engineering", "Engineer II",          122500, "2021-06-30", "Boulder"),
        (1007, "Wes",   "Halloran",  "Engineering", "Engineer I",            98000, "2023-02-11", "Boulder"),
        (1008, "Soraya","Vahid",     "Engineering", "Engineering Director", 215000, "2014-07-08", "Boulder"),
        (1009, "Bram",  "Veldhuizen","Engineering", "Engineer II",          120000, "2022-09-26", "Boulder"),
        (1010, "Inez",  "Calderon",  "Engineering", "Senior Engineer",      149500, "2019-12-04", "Boulder"),
        (1011, "Theo",  "Brandt",    "Finance",     "CFO",                  225000, "2015-05-20", "Munich"),
        (1012, "Cleo",  "Park",      "Finance",     "Controller",           142000, "2018-08-12", "Munich"),
        (1013, "Rafa",  "Sandoval",  "Finance",     "Senior Accountant",    102500, "2020-04-03", "Munich"),
        (1014, "Mira",  "Chen",      "Finance",     "FP&A Manager",         128000, "2019-02-18", "Munich"),
        (1015, "Heinrich","Vogt",    "Operations",  "Plant Manager",        148500, "2016-10-09", "Munich"),
        (1016, "Janna", "Ostrowski", "Operations",  "Logistics Coordinator", 78500, "2021-03-22", "Munich"),
        (1017, "Pavel", "Dvorak",    "Operations",  "Operations Analyst",    82000, "2022-11-15", "Munich"),
        (1018, "Sebastian","Vogel",  "Platform",    "Director of Platform", 198000, "2017-09-01", "Boulder"),
        (1019, "Devon", "Cole",      "Platform",    "Identity Lead",        165000, "2019-05-27", "Boulder"),
        (1020, "Akiko", "Tamura",    "Platform",    "Order Pipeline Lead",  167500, "2018-07-14", "Boulder"),
        (1021, "Niamh", "Doyle",     "Sales",       "VP of Sales",          205000, "2016-04-25", "Dublin"),
        (1022, "Quinn", "Beaumont",  "Sales",       "Account Executive",    112000, "2020-06-18", "Dublin"),
        (1023, "Ola",   "Adebayo",   "Sales",       "Sales Engineer",       128000, "2021-08-30", "Dublin"),
        (1024, "Yuki",  "Sato",      "Sales",       "Account Executive",    109500, "2022-02-07", "Dublin"),
        (1025, "Hugo",  "Lefevre",   "Marketing",   "Head of Marketing",    158000, "2017-11-19", "Paris"),
        (1026, "Esme",  "Whitcombe", "Marketing",   "Content Lead",          96500, "2021-01-25", "Paris"),
        (1027, "Karim", "Rashid",    "Marketing",   "Demand Gen Manager",   108000, "2020-09-10", "Paris"),
        (1028, "Saoirse","O'Riordan","HR",          "Head of People",       142000, "2018-12-03", "Dublin"),
        (1029, "Tilda", "Klein",     "HR",          "Recruiter",             82500, "2022-05-14", "Dublin"),
        (1030, "Bao",   "Tran",      "Legal",       "General Counsel",      195000, "2016-08-29", "Boulder"),
    ]
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if c_idx == 6:
                cell.number_format = MONEY_FMT

    # Set column widths
    widths = [12, 11, 13, 14, 25, 14, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Anchor fact for "not present" question
    ws.cell(row=len(rows) + 4, column=1,
            value="Note: Joe Smith is NOT employed at this company.").font = Font(italic=True)

    out = OUT_DIR / "01_clean_employees.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


# =========================================================================
# 2. MERGED HEADERS — multi-row headers with merged cells
# =========================================================================

def fixture_merged_headers_sales():
    """Sales table where headers span TWO rows and merge across regions."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Regional Sales"

    ws["A1"] = "Quarterly Sales Report — FY 2025 (USD thousands)"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:I1")

    # Row 3-4: merged headers
    # | Quarter | North America |  Europe       |  Asia-Pacific |  Total |
    # |         | West | East   | UK | DE | FR  | JP | AU       |        |
    ws["A3"] = "Quarter"
    ws.merge_cells("A3:A4")
    ws["B3"] = "North America"
    ws.merge_cells("B3:C3")
    ws["B4"] = "West"
    ws["C4"] = "East"
    ws["D3"] = "Europe"
    ws.merge_cells("D3:F3")
    ws["D4"] = "UK"
    ws["E4"] = "Germany"
    ws["F4"] = "France"
    ws["G3"] = "Asia-Pacific"
    ws.merge_cells("G3:H3")
    ws["G4"] = "Japan"
    ws["H4"] = "Australia"
    ws["I3"] = "Total"
    ws.merge_cells("I3:I4")

    for row in (3, 4):
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.font = HEADER
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

    # Data — Q1 to Q4
    data = [
        ("Q1 2025", 1240, 1815, 880,  990,  710, 425, 380, None),
        ("Q2 2025", 1395, 1922, 915, 1050,  745, 462, 410, None),
        ("Q3 2025", 1521, 2104, 947, 1108,  792, 491, 437, None),
        ("Q4 2025", 1612, 2231, 982, 1175,  831, 519, 458, None),
    ]
    for r_idx, row in enumerate(data, 5):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if c_idx >= 2 and c_idx <= 8:
                cell.number_format = "#,##0"

    # Total column — anchored values (precomputed)
    totals = [sum(d[1:8]) for d in data]
    for r_idx, t in enumerate(totals, 5):
        c = ws.cell(row=r_idx, column=9, value=t)
        c.font = Font(bold=True)
        c.fill = SUBHDR_FILL
        c.number_format = "#,##0"
        c.border = BORDER

    # Anchor fact for the test
    ws.cell(row=11, column=1,
            value="Anchor — Q3 2025 Europe (UK+Germany+France) totaled $2,847K").font = Font(italic=True, bold=True)
    # That's 947+1108+792 = 2,847 — verified

    widths = [10, 9, 9, 8, 11, 8, 8, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = OUT_DIR / "02_merged_headers_sales.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


# =========================================================================
# 3. MULTI-SHEET COMPANY — 4 sheets with cross-sheet relationships
# =========================================================================

def fixture_multi_sheet_company():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Products ──
    p = wb.active
    p.title = "Products"
    p.append(["ProductId", "ProductName", "Category", "UnitPrice", "InStock"])
    products = [
        ("PRD-101", "OrbiCore Sensor",        "Sensors",     189.00, 1240),
        ("PRD-102", "OrbiCore Sensor Pro",    "Sensors",     289.00,  410),
        ("PRD-103", "Helios Gateway",         "Networking",  549.00,   88),
        ("PRD-104", "Helios Gateway Mini",    "Networking",  319.00,  175),
        ("PRD-105", "Nimbus Repeater",        "Networking",   95.00,  502),
        ("PRD-106", "Thalia Display Panel",   "Displays",    412.00,  220),
        ("PRD-107", "Thalia Display Panel XL","Displays",    689.00,   45),
        ("PRD-108", "Atlas Battery Pack",     "Power",       129.00,  860),
        ("PRD-109", "Atlas Battery Pack Plus","Power",       219.00,  340),
        ("PRD-110", "Lyra Antenna Kit",       "Accessories",  39.00, 2105),
    ]
    for row in products:
        p.append(row)
    for r in p.iter_rows(min_row=1, max_row=1):
        for c in r:
            c.font = HEADER
            c.fill = HEADER_FILL

    # ── Sheet 2: Customers ──
    c = wb.create_sheet("Customers")
    c.append(["CustomerId", "CustomerName", "Tier", "PrimaryContact", "Country"])
    customers = [
        ("CUST-501", "Vellichor Industries",         "Platinum",  "M. Thalberg", "United States"),
        ("CUST-502", "Bramble & Forest Co.",         "Gold",      "F. Hawthorne","United Kingdom"),
        ("CUST-503", "PolarKraft Packaging GmbH",    "Gold",      "L. Schroeder","Germany"),
        ("CUST-504", "Sundial Foods Cooperative",    "Silver",    "R. Calderon", "Spain"),
        ("CUST-505", "Tessuto Holdings",             "Platinum",  "G. Lombardi", "Italy"),
        ("CUST-506", "Greenline Distributors",       "Silver",    "M. O'Connor", "Ireland"),
        ("CUST-507", "Cypress Container Corp",       "Gold",      "D. Whitman",  "United States"),
        ("CUST-508", "Mira Pharmaceuticals AG",      "Platinum",  "J. Engelhart","Switzerland"),
    ]
    for row in customers:
        c.append(row)
    for r in c.iter_rows(min_row=1, max_row=1):
        for cell in r:
            cell.font = HEADER
            cell.fill = HEADER_FILL

    # ── Sheet 3: Orders ──
    o = wb.create_sheet("Orders")
    o.append(["OrderId", "CustomerId", "ProductId", "Quantity",
              "OrderDate", "Status"])
    orders = [
        ("ORD-9001", "CUST-501", "PRD-103",  12, "2026-01-08", "Shipped"),
        ("ORD-9002", "CUST-501", "PRD-108",  40, "2026-01-08", "Shipped"),
        ("ORD-9003", "CUST-502", "PRD-101",  60, "2026-01-12", "Shipped"),
        ("ORD-9004", "CUST-503", "PRD-106",   8, "2026-01-15", "Pending"),
        ("ORD-9005", "CUST-504", "PRD-110", 200, "2026-01-19", "Shipped"),
        ("ORD-9006", "CUST-505", "PRD-107",   3, "2026-01-22", "Shipped"),
        ("ORD-9007", "CUST-505", "PRD-104",  15, "2026-01-22", "Shipped"),
        ("ORD-9008", "CUST-506", "PRD-105",  35, "2026-02-02", "Cancelled"),
        ("ORD-9009", "CUST-507", "PRD-102",  18, "2026-02-04", "Shipped"),
        ("ORD-9010", "CUST-508", "PRD-109",  22, "2026-02-09", "Shipped"),
        ("ORD-9011", "CUST-501", "PRD-101", 120, "2026-02-12", "Shipped"),
        ("ORD-9012", "CUST-503", "PRD-103",   4, "2026-02-15", "Pending"),
    ]
    for row in orders:
        o.append(row)
    for r in o.iter_rows(min_row=1, max_row=1):
        for cell in r:
            cell.font = HEADER
            cell.fill = HEADER_FILL

    # ── Sheet 4: Order Detail (cross-sheet formulas) ──
    d = wb.create_sheet("OrderDetail")
    d.append(["OrderId", "CustomerName", "ProductName", "Quantity",
              "UnitPrice", "LineTotal", "Country"])
    # Hardcoded computed cross-sheet values so we don't depend on
    # uncached formulas resolving.
    customer_map = {row[0]: row[1:] for row in customers}
    product_map  = {row[0]: row[1:] for row in products}
    for o_row in orders:
        oid, cid, pid, qty, _date, _stat = o_row
        cust_name, _tier, _contact, country = customer_map[cid]
        prod_name, _cat, unit_price, _stock = product_map[pid]
        d.append([oid, cust_name, prod_name, qty, unit_price,
                  round(qty * unit_price, 2), country])
    for r in d.iter_rows(min_row=1, max_row=1):
        for cell in r:
            cell.font = HEADER
            cell.fill = HEADER_FILL

    # Anchor facts:
    #   Largest order by line total: ORD-9011 (CUST-501 / Vellichor) × 120 PRD-101 @ $189 = $22,680
    #   PRD-107 (Thalia Display Panel XL) ordered only by Tessuto Holdings
    anchor = wb.create_sheet("Anchors")
    anchor.append(["Fact", "Value"])
    anchor.append(["Largest order by line total", "ORD-9011 / $22,680.00"])
    anchor.append(["Largest order customer", "Vellichor Industries"])
    anchor.append(["PRD-107 Thalia Display Panel XL ordered by",
                   "Tessuto Holdings (only)"])
    anchor.append(["Cancelled orders", "1 (ORD-9008)"])
    anchor.append(["Customers in Germany", "PolarKraft Packaging GmbH"])

    out = OUT_DIR / "03_multi_sheet_company.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


# =========================================================================
# 4. HIDDEN SHEET — visible operational data + hidden exec compensation
# =========================================================================

def fixture_hidden_sheet_exec_comp():
    """A workbook where one sheet is hidden. The competency question is
    twofold: does the agent surface visible-sheet facts correctly, AND
    does it inadvertently leak the hidden-sheet fingerprint string?

    We use a UNIQUE marker string in the hidden sheet ("ZX-HIDDEN-7Q-MARKER")
    so the test can detect leakage with a simple substring scan over the
    agent's answer."""
    wb = openpyxl.Workbook()

    # Visible sheet
    v = wb.active
    v.title = "Operations Metrics"
    v.append(["Metric", "Q1 2026", "Q2 2026 (proj)"])
    v.append(["Manufacturing yield (%)", 94.2, 95.1])
    v.append(["Average defect rate (ppm)", 312, 280])
    v.append(["Order fulfillment SLA met (%)", 97.4, 98.0])
    v.append(["Avg cycle time (days)", 4.2, 3.9])
    v.append(["Carbon intensity (kg CO2/unit)", 0.85, 0.82])

    # Anchor in the visible sheet
    v.append([])
    v.append(["Anchor: Q1 2026 manufacturing yield was 94.2%"])

    # Hidden sheet — executive comp with a unique marker
    h = wb.create_sheet("ExecComp")
    h.sheet_state = "hidden"
    h.append(["Executive", "BaseSalary", "Bonus", "Equity", "Marker"])
    h.append(["CEO",  650000, 850000, 1500000, "ZX-HIDDEN-7Q-MARKER"])
    h.append(["CFO",  425000, 380000,  720000, "ZX-HIDDEN-7Q-MARKER"])
    h.append(["COO",  410000, 360000,  680000, "ZX-HIDDEN-7Q-MARKER"])
    # The marker is a unique sentinel so the test can grep for leakage.
    # ZX-HIDDEN-7Q-MARKER appears in no other fixture.

    out = OUT_DIR / "04_hidden_sheet_exec_comp.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


# =========================================================================
# 5. LARGE INVENTORY — 500 SKUs to stress retrieval at scale
# =========================================================================

def fixture_large_inventory():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["SKU", "ProductFamily", "Variant", "WarehouseCity",
               "UnitsOnHand", "UnitPrice", "ReorderPoint"])
    for r in ws.iter_rows(min_row=1, max_row=1):
        for c in r:
            c.font = HEADER
            c.fill = HEADER_FILL

    families = ["OrbiCore", "Helios", "Nimbus", "Thalia", "Atlas",
                "Lyra", "Hyperion", "Cygnus", "Vega", "Polaris"]
    variants = ["Standard", "Pro", "Mini", "XL", "Plus", "Lite"]
    cities = ["Boulder", "Munich", "Dublin", "Paris", "Singapore",
              "Sao Paulo", "Toronto", "Sydney"]

    # Deterministic data
    for i in range(1, 501):
        fam = families[i % len(families)]
        var = variants[i % len(variants)]
        city = cities[(i * 3) % len(cities)]
        units = ((i * 17) % 850) + 5
        price = round(50 + ((i * 13) % 950) + (i % 9) * 0.5, 2)
        reorder = max(20, units // 4)
        ws.append([f"SKU-{i:04d}", fam, var, city, units, price, reorder])
        cell = ws.cell(row=i + 1, column=6)
        cell.number_format = MONEY_FMT

    # Specific fingerprint:
    # SKU-0173 — Hyperion Mini in Singapore — set unique price + units
    target_row = 174  # 173 + 1 header
    ws.cell(row=target_row, column=1, value="SKU-0173")
    ws.cell(row=target_row, column=2, value="Hyperion")
    ws.cell(row=target_row, column=3, value="Mini")
    ws.cell(row=target_row, column=4, value="Singapore")
    ws.cell(row=target_row, column=5, value=42)  # unique low stock
    ws.cell(row=target_row, column=6, value=1247.99)  # unique high price
    ws.cell(row=target_row, column=6).number_format = MONEY_FMT
    ws.cell(row=target_row, column=7, value=15)

    # Anchor on a dedicated sheet so the LLM can find it
    a = wb.create_sheet("Anchors")
    a.append(["Fact", "Value"])
    a.append(["SKU-0173 details", "Hyperion Mini, Singapore, 42 units, $1247.99"])
    a.append(["Total SKUs in this inventory", "500"])
    a.append(["Warehouse cities covered", "8"])
    a.append(["Highest-priced SKU", "SKU-0173 at $1247.99"])

    out = OUT_DIR / "05_large_inventory.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


# =========================================================================
# 6. MULTI-TABLE DASHBOARD — three distinct tables on one sheet
# =========================================================================

def fixture_multi_table_dashboard():
    """One sheet, three tables separated by blank rows. Tests whether the
    extractor segments them or smushes them together. Each table has a
    fingerprint value."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exec Dashboard"

    # Header
    ws["A1"] = "Cobalt Industries — Executive Dashboard, March 2026"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:E1")

    # ── Table 1: Top-line KPIs ──
    ws["A3"] = "TABLE 1 — Top-Line KPIs"
    ws["A3"].font = SUBHDR
    ws["A3"].fill = SUBHDR_FILL
    ws.merge_cells("A3:E3")
    headers1 = ["KPI", "Target", "Actual", "Variance", "Status"]
    for i, h in enumerate(headers1, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
    kpis = [
        ("Quarterly revenue ($M)", 14.5, 15.82, "+1.32",  "Above target"),
        ("Gross margin (%)",       58.0, 56.4,  "-1.6",   "Below target"),
        ("Active customers",      1850,  1923, "+73",     "Above target"),
        ("Employee headcount",     410,  398, "-12",     "Below target"),
        ("Net Promoter Score",     45,   48,   "+3",     "Above target"),
    ]
    for r_idx, k in enumerate(kpis, 5):
        for c_idx, v in enumerate(k, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    # ── Table 2: Regional breakdown ──
    # Leave 2 blank rows
    start = 5 + len(kpis) + 2
    ws.cell(row=start, column=1, value="TABLE 2 — Regional Revenue Breakdown").font = SUBHDR
    ws.cell(row=start, column=1).fill = SUBHDR_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=5)
    headers2 = ["Region", "Revenue ($M)", "% of total", "YoY growth", "Top customer"]
    for i, h in enumerate(headers2, 1):
        cell = ws.cell(row=start + 1, column=i, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
    regions = [
        ("North America",   7.42,  "46.9%", "+12.3%", "Vellichor Industries"),
        ("Europe",          5.18,  "32.7%",  "+8.7%", "PolarKraft Packaging GmbH"),
        ("Asia-Pacific",    2.41,  "15.2%", "+18.4%", "Mira Pharmaceuticals AG"),
        ("Latin America",   0.81,   "5.1%",  "+4.2%", "Sundial Foods Cooperative"),
    ]
    for r_idx, region in enumerate(regions, start + 2):
        for c_idx, v in enumerate(region, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    # ── Table 3: Top customers ──
    start2 = start + 2 + len(regions) + 2
    ws.cell(row=start2, column=1, value="TABLE 3 — Top 5 Customers by Revenue").font = SUBHDR
    ws.cell(row=start2, column=1).fill = SUBHDR_FILL
    ws.merge_cells(start_row=start2, start_column=1, end_row=start2, end_column=5)
    headers3 = ["Rank", "Customer", "Revenue ($K)", "Industry", "Tenure (yrs)"]
    for i, h in enumerate(headers3, 1):
        cell = ws.cell(row=start2 + 1, column=i, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
    top_customers = [
        (1, "Vellichor Industries",          3260, "Manufacturing", 7),
        (2, "PolarKraft Packaging GmbH",     1823, "Packaging",     4),
        (3, "Mira Pharmaceuticals AG",       1409, "Pharma",        5),
        (4, "Tessuto Holdings",              1172, "Textiles",      3),
        (5, "Sundial Foods Cooperative",      926, "Food & Bev",    6),
    ]
    for r_idx, t in enumerate(top_customers, start2 + 2):
        for c_idx, v in enumerate(t, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    # Anchor fact on a separate sheet
    a = wb.create_sheet("Anchors")
    a.append(["Fact", "Value"])
    a.append(["Q1 2026 quarterly revenue actual ($M)", "15.82"])
    a.append(["Asia-Pacific YoY growth", "+18.4%"])
    a.append(["Rank 3 customer by revenue", "Mira Pharmaceuticals AG ($1,409K)"])
    a.append(["Below-target KPIs", "Gross margin, Employee headcount"])

    for i, w in enumerate([28, 16, 16, 14, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = OUT_DIR / "06_multi_table_dashboard.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


# =========================================================================

if __name__ == "__main__":
    fixture_clean_employees()
    fixture_merged_headers_sales()
    fixture_multi_sheet_company()
    fixture_hidden_sheet_exec_comp()
    fixture_large_inventory()
    fixture_multi_table_dashboard()
    print("All competency fixtures generated.")
