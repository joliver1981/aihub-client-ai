"""
Fixture generator — Email Report Reconciliation competency test.

Builds a "vendor statement" the way a real counterparty would email it: an
Excel workbook (and a PDF twin) listing invoices and the amounts THEY think are
outstanding. It is built from REAL ERPDB invoices, then seeded with a handful
of deliberate DISCREPANCIES vs the database:

  - amount mismatches (their figure != ERPDB amount_due)
  - an invoice on their statement that does NOT exist in ERPDB
  - an ERPDB invoice they OMITTED

An ANSWER_KEY.md records every planted discrepancy so a human can grade the
agent's reconciliation. Deterministic given the DB snapshot.

Run:  python make_fixtures.py            (pull ERPDB, build xlsx+pdf+key)
      python make_fixtures.py --email    (also drop into the email outbox note)
"""
import os
import sys
from datetime import date

import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", "_fixtures"))
OUT = os.path.join(ROOT, "reconciliation")
CS = "DRIVER={SQL Server};SERVER=10.0.0.6;DATABASE=ERPDB;UID=ai_user;PWD=Bradynov11"

STATEMENT_VENDOR = "Global Parts Distributors"


def pull_invoices():
    c = pyodbc.connect(CS, timeout=10); cur = c.cursor()
    # amount_due is the DB's outstanding figure per invoice
    cur.execute("""
        SELECT TOP 14 invoice_id, customer_name, customer_po,
               CAST(amount_due AS FLOAT) AS amount_due, status
        FROM dbo.Invoices
        WHERE amount_due IS NOT NULL AND amount_due > 0
        ORDER BY invoice_id
    """)
    rows = [{"invoice_id": r[0], "customer": r[1] or "", "po": r[2] or "",
             "db_amount": round(float(r[3]), 2), "status": r[4] or ""}
            for r in cur.fetchall()]
    c.close()
    return rows


def build():
    os.makedirs(OUT, exist_ok=True)
    db = pull_invoices()
    if len(db) < 8:
        print("ERROR: not enough ERPDB invoices to build the statement", file=sys.stderr)
        sys.exit(1)

    # Their statement = the DB figures, with planted discrepancies.
    statement, key = [], []
    for i, r in enumerate(db):
        their = r["db_amount"]
        note = "match"
        if i == 2:
            their = round(r["db_amount"] + 500.00, 2); note = "amount HIGH by $500.00"
        elif i == 5:
            their = round(r["db_amount"] - 250.00, 2); note = "amount LOW by $250.00"
        elif i == 8:
            their = round(r["db_amount"] * 1.10, 2)
            note = f"amount HIGH by ${round(their - r['db_amount'],2):,.2f} (10%)"
        statement.append({"invoice_id": r["invoice_id"], "po": r["po"],
                          "their_amount": their})
        if note != "match":
            key.append({"invoice_id": r["invoice_id"], "type": "amount_mismatch",
                        "their": their, "db": r["db_amount"], "detail": note})
    # a phantom invoice only on their statement
    phantom = {"invoice_id": "GP-STMT-9001", "po": "PO-00000", "their_amount": 3199.99}
    statement.append(phantom)
    key.append({"invoice_id": phantom["invoice_id"], "type": "not_in_erpdb",
                "their": phantom["their_amount"], "db": None,
                "detail": "on statement, NOT in ERPDB"})
    # an omitted invoice: on DB, dropped from their statement
    omitted = db[10]
    key.append({"invoice_id": omitted["invoice_id"], "type": "omitted_from_statement",
                "their": None, "db": omitted["db_amount"],
                "detail": "in ERPDB, MISSING from their statement"})
    statement = [s for s in statement if s["invoice_id"] != omitted["invoice_id"]]

    _xlsx(statement)
    _pdf(statement)
    _key(statement, key, db)
    print(f"statement rows: {len(statement)} | planted discrepancies: {len(key)}")
    print(f"-> {OUT}\\vendor_statement.xlsx / .pdf")
    print(f"-> {ROOT}\\02_ANSWER_KEY.md")


def _xlsx(statement):
    wb = Workbook(); ws = wb.active; ws.title = "Statement"
    ws["A1"] = f"{STATEMENT_VENDOR} — Statement of Account"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"As of {date.today():%Y-%m-%d}   |   Amounts they show as outstanding"
    hdr = ["Invoice #", "Your PO", "Amount Outstanding"]
    ws.append([]); ws.append(hdr)
    for c in ws[4]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDE6F5")
    total = 0.0
    for s in statement:
        ws.append([s["invoice_id"], s["po"], s["their_amount"]])
        total += s["their_amount"]
    ws.append([]); ws.append(["", "TOTAL", round(total, 2)])
    ws[f"C{ws.max_row}"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    for r in ws.iter_rows(min_row=5):
        if r[2].value is not None and isinstance(r[2].value, (int, float)):
            r[2].number_format = '$#,##0.00'
    wb.save(os.path.join(OUT, "vendor_statement.xlsx"))


def _pdf(statement):
    path = os.path.join(OUT, "vendor_statement.pdf")
    c = canvas.Canvas(path, pagesize=LETTER); w, h = LETTER
    y = h - 1 * inch
    c.setFont("Helvetica-Bold", 16)
    c.drawString(1 * inch, y, f"{STATEMENT_VENDOR} — Statement of Account")
    c.setFont("Helvetica", 9); y -= 18
    c.drawString(1 * inch, y, f"As of {date.today():%Y-%m-%d}  |  amounts shown as outstanding")
    y -= 30; c.setFont("Helvetica-Bold", 10)
    c.drawString(1 * inch, y, "Invoice #"); c.drawString(3 * inch, y, "Your PO")
    c.drawRightString(w - 1 * inch, y, "Amount Outstanding")
    c.line(1 * inch, y - 4, w - 1 * inch, y - 4); y -= 20
    c.setFont("Helvetica", 10); total = 0.0
    for s in statement:
        c.drawString(1 * inch, y, s["invoice_id"])
        c.drawString(3 * inch, y, s["po"])
        c.drawRightString(w - 1 * inch, y, f"${s['their_amount']:,.2f}")
        total += s["their_amount"]; y -= 16
    y -= 6; c.line(4.6 * inch, y, w - 1 * inch, y); y -= 18
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(w - 2.1 * inch, y, "TOTAL:")
    c.drawRightString(w - 1 * inch, y, f"${total:,.2f}")
    c.showPage(); c.save()


def _key(statement, key, db):
    stmt_total = round(sum(s["their_amount"] for s in statement), 2)
    db_total = round(sum(r["db_amount"] for r in db), 2)
    lines = ["# Answer key — email report reconciliation", "",
             f"Vendor statement: **{STATEMENT_VENDOR}**",
             f"- Statement total (their figures): **${stmt_total:,.2f}**",
             f"- ERPDB total (all {len(db)} pulled invoices): **${db_total:,.2f}**",
             "",
             "## Planted discrepancies the agent MUST find",
             "| Invoice | Type | Their amount | ERPDB amount | Note |",
             "|---|---|---|---|---|"]
    for k in key:
        their = f"${k['their']:,.2f}" if k["their"] is not None else "—"
        dbv = f"${k['db']:,.2f}" if k["db"] is not None else "—"
        lines.append(f"| {k['invoice_id']} | {k['type']} | {their} | {dbv} | {k['detail']} |")
    lines += ["",
              "A correct reconciliation email should call out every row above: "
              "the amount mismatches (with the delta), the invoice that isn't in "
              "ERPDB, and the ERPDB invoice missing from their statement — and "
              "should NOT flag the matching invoices."]
    with open(os.path.join(ROOT, "02_ANSWER_KEY.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


if __name__ == "__main__":
    build()
