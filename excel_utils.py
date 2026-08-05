# excel_utils.py

"""
Excel Template Population Utilities

Provides intelligent Excel template detection, AI-powered data mapping,
and population using Pandas + openpyxl.

Three schema modes:
1. Existing Template - Read schema from Excel file
2. User-Defined - Schema provided in config  
3. AI-Generated - AI determines schema from input data

Two template types (auto-detected):
- Table: Headers in row 1, data appended as rows
- Form: Labels with adjacent value cells
"""

import pandas as pd
import json
import os
import logging
from logging.handlers import WatchedFileHandler
from typing import Dict, List, Any, Optional, Tuple, Union
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
import re
import threading
from collections import OrderedDict
from datetime import datetime, date
from CommonUtils import rotate_logs_on_startup, get_log_path
import system_prompts as sysprompts


# Configure logging
def setup_logging():
    """Configure logging for the excel utils"""
    logger = logging.getLogger("ExcelUtils")
    log_level_name = os.getenv('LOG_LEVEL', 'DEBUG')
    log_level = getattr(logging, log_level_name, logging.DEBUG)
    logger.setLevel(log_level)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler = WatchedFileHandler(filename=os.getenv('EXCEL_UTILS_LOG', get_log_path('excel_utils_log.txt')), encoding='utf-8')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    return logger

rotate_logs_on_startup(os.getenv('EXCEL_UTILS_LOG', get_log_path('excel_utils_log.txt')))

logger = setup_logging()

try:
    from AppUtils import azureMiniQuickPrompt
    from AppUtils import quickPrompt
except ImportError:
    logger.warning("Could not import quickPrompt from AppUtils. AI features might be unavailable.")
    from AppUtils import azureQuickPrompt as quickPrompt


def detect_template_type(ws) -> str:
    """
    Auto-detect whether a worksheet is table-style or form-style using AI.
    
    Table-style: Headers in row 1, data in subsequent rows (columnar layout)
    Form-style: Labels with adjacent value cells (scattered layout)
    
    Args:
        ws: openpyxl worksheet object
        
    Returns:
        "table" or "form"
    """
    logger.debug("Detecting template type...")
    
    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    
    logger.debug(f"Worksheet dimensions: {max_row} rows x {max_col} columns")
    
    # Quick heuristic for obvious cases
    if max_row <= 1:
        logger.debug("Only 1 row detected, assuming table type with headers only")
        return "table"
    
    # Build a sample of the worksheet structure for AI analysis
    sample_data = []
    
    # Get first few rows (up to 5 rows, up to 20 columns)
    rows_to_check = min(max_row, 5)
    cols_to_check = min(max_col, 20)
    
    for row in range(1, rows_to_check + 1):
        row_data = []
        for col in range(1, cols_to_check + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                # Truncate long values
                cell_str = str(cell_value)[:50]
                row_data.append(f"[{row},{col}]: {cell_str}")
            else:
                row_data.append(f"[{row},{col}]: (empty)")
        sample_data.append(" | ".join(row_data))
    
    sample_text = "\n".join(sample_data)
    
    # Use AI to detect template type
    if quickPrompt:
        try:
            result = _ai_detect_template_type(sample_text, max_row, max_col)
            logger.debug(f"AI detected template type: {result}")
            return result
        except Exception as e:
            logger.warning(f"AI template detection failed: {str(e)}, falling back to heuristic")
            return _heuristic_detect_template_type(ws, max_row, max_col)
    else:
        logger.debug("quickPrompt not available, using heuristic detection")
        return _heuristic_detect_template_type(ws, max_row, max_col)


def _ai_detect_template_type(sample_text: str, max_row: int, max_col: int) -> str:
    """
    Use AI to detect whether the template is table-style or form-style.
    
    Args:
        sample_text: Sample of worksheet cell values
        max_row: Total rows in worksheet
        max_col: Total columns in worksheet
        
    Returns:
        "table" or "form"
    """
    logger.debug("Using AI to detect template type...")
    
    system_prompt = """You are an expert at analyzing Excel spreadsheet structures.
Your task is to determine if a spreadsheet is TABLE-style or FORM-style.

TABLE-style characteristics:
- Row 1 contains column headers (field names)
- Data rows are below the headers
- Each row represents a record
- Columns represent fields/attributes
- Example: Customer | Date | Amount | Status (with data rows below)

FORM-style characteristics:
- Labels and values are scattered across the sheet
- Often has label-value pairs side by side or stacked
- Single record with fields in various positions
- Example: "Name:" in A1 with value in B1, "Date:" in A2 with value in B2

You must respond with ONLY a single word: either "table" or "form"
Do not include any explanation or additional text."""

    user_prompt = f"""Analyze this Excel spreadsheet sample and determine if it's TABLE-style or FORM-style.

Worksheet dimensions: {max_row} rows x {max_col} columns

Sample data (first few rows and columns):
{sample_text}

Based on this structure, is this a TABLE or FORM layout? Respond with only "table" or "form"."""

    # FIXED 2026-08-03: this passed `temperature=0`, which azureMiniQuickPrompt has
    # never accepted (its parameter is `temp`), so the call raised TypeError every
    # single time and silently fell back to the heuristic below — AI template
    # detection has never actually run. Correcting the name alone would have been a
    # REGRESSION while this ran once per exported row (measured 1.40 -> 2.98 s/row);
    # it is safe now only because detect_template_schema is memoised per
    # (file, sheet, header signature), so it runs ONCE per export instead of per row.
    _count_llm_call("template_type_detection")
    response = azureMiniQuickPrompt(
        user_prompt,
        system_prompt,
        temp=0
    )
    
    if response:
        result = response.strip().lower()
        if result in ("table", "form"):
            return result
        else:
            logger.warning(f"AI returned unexpected value: {result}, defaulting to table")
            return "table"
    else:
        logger.warning("AI returned empty response, defaulting to table")
        return "table"


def _heuristic_detect_template_type(ws, max_row: int, max_col: int) -> str:
    """
    Fallback heuristic detection for template type.
    
    Args:
        ws: openpyxl worksheet object
        max_row: Total rows in worksheet
        max_col: Total columns in worksheet
        
    Returns:
        "table" or "form"
    """
    logger.debug("Using heuristic template type detection...")
    
    # Check row 1 for potential headers
    row1_values = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    row1_filled = sum(1 for v in row1_values if v is not None and str(v).strip())
    
    logger.debug(f"Row 1 has {row1_filled} filled cells out of {max_col}")
    
    # Check for label-value pattern
    label_value_pairs = 0
    for row in range(1, min(max_row + 1, 20)):
        for col in range(1, max_col):
            cell_val = ws.cell(row=row, column=col).value
            next_cell_val = ws.cell(row=row, column=col + 1).value
            
            if cell_val and isinstance(cell_val, str):
                cell_str = str(cell_val).strip()
                if cell_str.endswith(':') or (len(cell_str) > 2 and next_cell_val is None):
                    label_value_pairs += 1
    
    logger.debug(f"Detected {label_value_pairs} potential label-value pairs")
    
    # Decision logic - more lenient for table detection
    row1_percentage = row1_filled / max_col if max_col > 0 else 0
    
    # Consider it a table if row 1 has multiple headers
    # Use lower threshold (30%) or absolute count (10+)
    if row1_filled >= 3 and (row1_percentage >= 0.3 or row1_filled >= 10):
        logger.debug("Detected as TABLE type (row 1 appears to be headers)")
        return "table"
    elif label_value_pairs >= 3 and row1_filled < 5:
        logger.debug("Detected as FORM type (multiple label-value pairs found)")
        return "form"
    else:
        logger.debug("Defaulting to TABLE type")
        return "table"


_SCHEMA_CACHE = {}
_SCHEMA_CACHE_MAX = 64


def _header_signature(template_path: str, sheet_name: str = None):
    """Cheap fingerprint of a sheet's SHAPE: its header row plus column count.

    Deliberately not mtime/size — an append changes both on every row, which would
    make the cache miss every time, which is the bug we are fixing. What matters
    is whether the LAYOUT changed, and appending rows never changes the header.
    Returns None if the file cannot be read, which disables caching for that call.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        header = tuple(str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1)))
        sig = (ws.title, header, ws.max_column)
        wb.close()
        return sig
    except Exception:
        return None


def detect_template_schema(template_path: str, sheet_name: str = None) -> Dict:
    """Memoised wrapper — see _detect_template_schema_uncached for the real work.

    WHY (2026-08-03): this was called once PER EXPORTED ROW from the Excel Export
    append loop, and it performs an AI call to classify the sheet as table-vs-form.
    The answer cannot change while you are appending rows to your own output file,
    so it is computed once per (file, sheet, header-shape) and reused. Cache is
    keyed on the header signature rather than the file's mtime precisely because
    appending mutates mtime on every row.
    """
    key = (os.path.abspath(template_path or ""), sheet_name,
           _header_signature(template_path, sheet_name))
    if key[2] is not None and key in _SCHEMA_CACHE:
        logger.debug("Template schema cache HIT for %s", template_path)
        return _SCHEMA_CACHE[key]
    schema = _detect_template_schema_uncached(template_path, sheet_name)
    if key[2] is not None:
        if len(_SCHEMA_CACHE) >= _SCHEMA_CACHE_MAX:
            _SCHEMA_CACHE.clear()
        _SCHEMA_CACHE[key] = schema
    return schema


def _detect_template_schema_uncached(template_path: str, sheet_name: str = None) -> Dict:
    """
    Analyze an Excel template and extract its schema.
    Auto-detects whether template is table-style or form-style.
    
    Args:
        template_path: Path to the Excel template file
        sheet_name: Specific sheet to analyze (optional, defaults to active)
    
    Returns:
        {
            "type": "table" | "form",
            "sheet_name": "Sheet1",
            "columns": ["Col A", "Col B", ...],      # for table type
            "data_start_row": 2,                      # for table type
            "cells": {                                # for form type
                "Field Label": {"cell": "B2", "label_cell": "A2"},
                ...
            },
            "raw_structure": {...}                    # for AI context
        }
    """
    logger.info(f"Detecting template schema from: {template_path}")
    
    if not os.path.exists(template_path):
        error_msg = f"Template file not found: {template_path}"
        logger.error(error_msg)
        raise FileNotFoundError(error_msg)
    
    try:
        wb = load_workbook(template_path, data_only=True)
        logger.debug(f"Loaded workbook with sheets: {wb.sheetnames}")
        
        # Get target sheet
        requested_sheet_name = sheet_name  # Preserve the originally requested name
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.debug(f"Using specified sheet: {sheet_name}")
        else:
            ws = wb.active
            if not requested_sheet_name:
                sheet_name = ws.title
            else:
                # Sheet doesn't exist yet — use active sheet for schema detection
                # but preserve the requested sheet_name so it gets created later
                logger.debug(f"Sheet '{sheet_name}' not found, using active sheet for schema detection but preserving target name")
            logger.debug(f"Using active sheet: {ws.title}, target sheet_name: {sheet_name}")

        # Detect template type
        template_type = detect_template_type(ws)

        schema = {
            "type": template_type,
            "sheet_name": sheet_name,
            "template_path": template_path
        }
        
        if template_type == "table":
            schema.update(_extract_table_schema(ws))
        else:
            schema.update(_extract_form_schema(ws))
        
        wb.close()
        
        logger.info(f"Schema detection complete. Type: {template_type}")
        logger.debug(f"Full schema: {json.dumps(schema, indent=2, default=str)}")
        
        return schema
        
    except Exception as e:
        logger.error(f"Error detecting template schema: {str(e)}", exc_info=True)
        raise


def _extract_table_schema(ws) -> Dict:
    """
    Extract schema from a table-style worksheet.
    
    Args:
        ws: openpyxl worksheet
        
    Returns:
        Dict with columns, data_start_row, and raw_structure
    """
    logger.debug("Extracting table schema...")
    
    max_col = ws.max_column
    max_row = ws.max_row
    
    # Get headers from row 1
    columns = []
    for col in range(1, max_col + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value is not None and str(cell_value).strip():
            columns.append({
                "name": str(cell_value).strip(),
                "column_letter": get_column_letter(col),
                "column_index": col
            })
    
    logger.debug(f"Found {len(columns)} columns: {[c['name'] for c in columns]}")
    
    # Determine where data starts (usually row 2)
    data_start_row = 2
    
    # Find the last row with data
    last_data_row = 1
    for row in range(2, max_row + 1):
        row_has_data = any(
            ws.cell(row=row, column=c['column_index']).value is not None 
            for c in columns
        )
        if row_has_data:
            last_data_row = row
    
    logger.debug(f"Data starts at row {data_start_row}, last data row: {last_data_row}")
    
    # Build raw structure for AI context
    raw_structure = {
        "headers": [c['name'] for c in columns],
        "sample_data": []
    }
    
    # Get sample data (up to 3 rows)
    for row in range(data_start_row, min(last_data_row + 1, data_start_row + 3)):
        row_data = {}
        for c in columns:
            cell_val = ws.cell(row=row, column=c['column_index']).value
            row_data[c['name']] = cell_val
        if any(v is not None for v in row_data.values()):
            raw_structure["sample_data"].append(row_data)
    
    return {
        "columns": columns,
        "data_start_row": data_start_row,
        "last_data_row": last_data_row,
        "raw_structure": raw_structure
    }


def _extract_form_schema(ws) -> Dict:
    """
    Extract schema from a form-style worksheet using AI to identify fields.
    
    Args:
        ws: openpyxl worksheet
        
    Returns:
        Dict with cells mapping and raw_structure
    """
    logger.debug("Extracting form schema...")
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Build a representation of the worksheet for AI analysis
    cell_data = []
    for row in range(1, min(max_row + 1, 50)):  # Limit to first 50 rows
        for col in range(1, min(max_col + 1, 20)):  # Limit to first 20 columns
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                cell_data.append({
                    "cell": f"{get_column_letter(col)}{row}",
                    "value": str(cell_value)[:100],  # Truncate long values
                    "row": row,
                    "col": col
                })
    
    logger.debug(f"Found {len(cell_data)} non-empty cells for form analysis")
    
    raw_structure = {
        "cells": cell_data,
        "dimensions": {"rows": max_row, "cols": max_col}
    }
    
    # Use AI to identify form fields
    cells_mapping = {}
    
    if quickPrompt and cell_data:
        try:
            cells_mapping = _ai_identify_form_fields(cell_data)
        except Exception as e:
            logger.warning(f"AI form field identification failed: {str(e)}")
            # Fallback: simple heuristic detection
            cells_mapping = _heuristic_form_fields(cell_data)
    else:
        cells_mapping = _heuristic_form_fields(cell_data)
    
    return {
        "cells": cells_mapping,
        "raw_structure": raw_structure
    }


def _ai_identify_form_fields(cell_data: List[Dict]) -> Dict:
    """
    Use AI to identify fillable form fields from cell data.
    
    Args:
        cell_data: List of cell information dicts
        
    Returns:
        Dict mapping field names to cell locations
    """
    logger.debug("Using AI to identify form fields...")
    
    system_prompt = """You are an expert at analyzing Excel form templates. 
Your task is to identify fillable fields in a form-style Excel template.
You must return ONLY valid JSON with no additional text or explanation."""

    user_prompt = f"""Analyze this Excel template structure and identify the fillable fields.
    
Template cells (cell address and current value):
{json.dumps(cell_data, indent=2)}

Identify fields where:
1. A cell contains a label (like "Customer Name:", "Date:", "Account #", etc.)
2. The adjacent cell (usually to the right or below) is where data should be entered

Return ONLY a JSON object in this exact format:
{{
    "fields": [
        {{
            "label": "Customer Name",
            "value_cell": "B2",
            "label_cell": "A2",
            "expected_type": "string"
        }}
    ]
}}

Rules:
- Include only actual fillable fields, not headers or titles
- expected_type should be: string, number, date, currency, or email
- If a cell appears to already have a value that looks like example data, still include it
- Return valid JSON only, no markdown or explanation"""

    try:
        response = quickPrompt(user_prompt, system=system_prompt, temp=0.0)
        logger.debug(f"AI response for form fields: {response[:500]}...")
        
        # Clean up response (remove markdown if present)
        response = response.strip()
        if response.startswith("```"):
            response = re.sub(r'^```json?\s*', '', response)
            response = re.sub(r'\s*```$', '', response)
        
        result = json.loads(response)
        
        # Convert to our expected format
        cells_mapping = {}
        for field in result.get("fields", []):
            label = field.get("label", "")
            if label:
                cells_mapping[label] = {
                    "cell": field.get("value_cell", ""),
                    "label_cell": field.get("label_cell", ""),
                    "expected_type": field.get("expected_type", "string")
                }
        
        logger.debug(f"AI identified {len(cells_mapping)} form fields")
        return cells_mapping
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse AI response as JSON: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"AI form field identification error: {str(e)}")
        raise


def _heuristic_form_fields(cell_data: List[Dict]) -> Dict:
    """
    Fallback heuristic to identify form fields without AI.
    
    Args:
        cell_data: List of cell information dicts
        
    Returns:
        Dict mapping field names to cell locations
    """
    logger.debug("Using heuristic form field detection...")
    
    cells_mapping = {}
    
    # Build a lookup by position
    cell_lookup = {(c['row'], c['col']): c for c in cell_data}
    
    for cell in cell_data:
        value = str(cell['value']).strip()
        row, col = cell['row'], cell['col']
        
        # Check if this looks like a label
        is_label = (
            value.endswith(':') or
            value.endswith('?') or
            any(keyword in value.lower() for keyword in 
                ['name', 'date', 'number', 'address', 'email', 'phone', 'id', 'amount', 'total'])
        )
        
        if is_label:
            # Check for value cell to the right
            right_cell = cell_lookup.get((row, col + 1))
            # Check for value cell below
            below_cell = cell_lookup.get((row + 1, col))
            
            label_name = value.rstrip(':').rstrip('?').strip()
            
            if right_cell or not below_cell:
                # Assume value is to the right
                value_cell = f"{get_column_letter(col + 1)}{row}"
            else:
                # Assume value is below
                value_cell = f"{get_column_letter(col)}{row + 1}"
            
            cells_mapping[label_name] = {
                "cell": value_cell,
                "label_cell": cell['cell'],
                "expected_type": "string"
            }
    
    logger.debug(f"Heuristic identified {len(cells_mapping)} form fields")
    return cells_mapping


def generate_schema_from_data(raw_data: Any, ai_instructions: str = None) -> Dict:
    """
    AI-generated schema mode: Analyze input data and create optimal schema.
    
    Args:
        raw_data: Input data to analyze
        ai_instructions: Hints about desired output structure
    
    Returns:
        Schema dict compatible with map_data_to_schema
    """
    logger.info("Generating schema from data using AI...")
    logger.debug(f"Input data type: {type(raw_data)}")
    logger.debug(f"Input data preview: {str(raw_data)[:500]}...")
    
    if not quickPrompt:
        raise RuntimeError("AI features unavailable - quickPrompt not imported")
    
    # Convert data to string if needed
    if isinstance(raw_data, (dict, list)):
        data_str = json.dumps(raw_data, indent=2, default=str)
    else:
        data_str = str(raw_data)
    
    system_prompt = """You are an expert at analyzing data and creating optimal Excel schemas.
Your task is to analyze input data and determine the best column structure for an Excel spreadsheet.
You must return ONLY valid JSON with no additional text or explanation."""

    user_prompt = f"""Analyze this data and create an optimal Excel schema for storing it.

INPUT DATA:
{data_str[:3000]}

{f"ADDITIONAL INSTRUCTIONS: {ai_instructions}" if ai_instructions else ""}

Create a schema that:
1. Captures all important data fields
2. Uses clear, professional column names
3. Identifies appropriate data types

Return ONLY a JSON object in this exact format:
{{
    "type": "table",
    "columns": [
        {{"name": "Column Name", "expected_type": "string"}},
        {{"name": "Amount", "expected_type": "currency"}},
        {{"name": "Date", "expected_type": "date"}}
    ],
    "data_start_row": 2
}}

Rules:
- Column names should be clear and professional
- expected_type should be: string, number, date, currency, email, or boolean
- Return valid JSON only, no markdown or explanation"""

    try:
        _count_llm_call("schema_generation")
        response = quickPrompt(user_prompt, system=system_prompt, temp=0.0)
        logger.debug(f"AI schema generation response: {response[:500]}...")
        
        # Clean up response
        response = response.strip()
        if response.startswith("```"):
            response = re.sub(r'^```json?\s*', '', response)
            response = re.sub(r'\s*```$', '', response)
        
        schema = json.loads(response)
        
        # Ensure required fields
        schema.setdefault("type", "table")
        schema.setdefault("data_start_row", 2)
        
        # Add column indices
        for i, col in enumerate(schema.get("columns", []), 1):
            col["column_index"] = i
            col["column_letter"] = get_column_letter(i)
        
        logger.info(f"Generated schema with {len(schema.get('columns', []))} columns")
        logger.debug(f"Generated schema: {json.dumps(schema, indent=2)}")
        
        return schema
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse AI schema response: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Schema generation error: {str(e)}", exc_info=True)
        raise


def _direct_table_mapping(raw_data: Any, schema: Dict):
    """Map records to table columns WITHOUT the AI, or return None to fall through.

    Returns {"rows": [...]} only when the mapping is forced — i.e. every record's
    field names are exactly the set of target column names (case/whitespace
    insensitive). In that situation the AI mapper can only echo the input back,
    so calling it is a pure round-trip cost.

    Returns None — deliberately conservative — for anything else:
      * form-style schemas (cells, not columns)
      * no columns, or duplicate column names after normalising (ambiguous)
      * records that are not dicts
      * a record missing a column, or carrying a field the sheet does not have
    Every one of those still goes to the unchanged AI path, so semantic mapping
    (e.g. "company" -> "Customer Name") is completely unaffected.
    """
    if (schema or {}).get("type", "table") != "table":
        return None
    cols = (schema or {}).get("columns") or []
    col_names = [c.get("name", c) if isinstance(c, dict) else c for c in cols]
    col_names = [str(c) for c in col_names if c is not None and str(c).strip() != ""]
    if not col_names:
        return None

    norm_to_col = {}
    for c in col_names:
        key = c.strip().lower()
        if key in norm_to_col:
            return None                      # duplicate columns -> ambiguous
        norm_to_col[key] = c

    records = raw_data if isinstance(raw_data, list) else [raw_data]
    if not records or not all(isinstance(r, dict) for r in records):
        return None

    rows = []
    for rec in records:
        keyed = {}
        for k, v in rec.items():
            keyed[str(k).strip().lower()] = v
        if len(keyed) != len(rec) or set(keyed) != set(norm_to_col):
            return None                      # not an exact match -> let the AI do it
        rows.append({norm_to_col[k]: v for k, v in keyed.items()})
    return {"rows": rows}


# ---------------------------------------------------------------------------
# Mapping-spec path (2026-08-05, approved by james).
#
# The semantic mapping between a record's field names and a sheet's columns is
# a function of the SCHEMA, not the row — yet the legacy path pays one
# primary-model call per row re-deriving it (a 1,000-row export on the
# semantic path ≈ 1M+ tokens). Here the mini model is asked ONCE per distinct
# (field set, column set) for a reusable mapping spec — which source field
# feeds which column, with one transform from a closed vocabulary — and the
# spec is applied to every row in pure Python. The model chooses the mapping;
# deterministic code executes it, failing closed to the raw value per cell.
#
# Trust boundary: the model's output is never executed as-is. The response is
# schema-constrained (strict json_schema with enums over the ACTUAL field
# names, column names, and transform ids), then re-validated deterministically
# (unknown source/target dropped, unknown transform -> "none"), so an
# off-vocabulary answer can cost polish, never correctness.
# ---------------------------------------------------------------------------

_LLM_CALL_STATS_LOCK = threading.Lock()
_LLM_CALL_STATS: Dict[str, int] = {}


def _count_llm_call(kind: str) -> int:
    """Count every model call this module makes, by kind. Returns the new
    module-lifetime total so call sites can log it."""
    with _LLM_CALL_STATS_LOCK:
        _LLM_CALL_STATS[kind] = _LLM_CALL_STATS.get(kind, 0) + 1
        total = sum(_LLM_CALL_STATS.values())
    logger.info(f"excel_utils LLM call #{total} (kind={kind})")
    return total


def get_llm_call_stats() -> Dict[str, int]:
    """Snapshot of {kind: count} for this module. Used by ops logging and the
    per-export call-count tripwire tests."""
    with _LLM_CALL_STATS_LOCK:
        return dict(_LLM_CALL_STATS)


def reset_llm_call_stats() -> None:
    with _LLM_CALL_STATS_LOCK:
        _LLM_CALL_STATS.clear()


def _transform_none(value):
    return value


def _transform_text(value):
    if value is None:
        return None
    return str(value)


def _transform_number(value):
    """Numeric coercion: strips currency symbols/thousands separators, handles
    accounting negatives '(1,234.56)'. Raises on anything it cannot parse —
    the applier writes the raw value instead."""
    if isinstance(value, bool):
        raise ValueError("boolean is not a number")
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        raise ValueError(f"cannot coerce {type(value).__name__} to number")
    s = value.strip()
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()
    s = re.sub(r"[$€£¥,\s]", "", s)
    if s == "":
        raise ValueError("empty numeric string")
    number = int(s) if re.fullmatch(r"[+-]?\d+", s) else float(s)
    return -number if negative else number


# Ordered, fixed date formats tried before the generic parser. Ambiguous
# numeric dates resolve month-first (US convention) — documented policy, not
# a guess made per row.
_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%d-%b-%Y",
    "%d %b %Y",
    "%b %d, %Y",
    "%B %d, %Y",
    "%Y%m%d",
)


def _transform_date_iso(value):
    """ISO date rendering. datetime/date objects convert directly; strings go
    through the fixed format list then dateutil (month-first). Raises when the
    value cannot be parsed — the applier writes the raw value instead."""
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if not isinstance(value, str):
        raise ValueError(f"cannot parse {type(value).__name__} as date")
    s = value.strip()
    if not s:
        raise ValueError("empty date string")
    for fmt in _DATE_FORMATS:
        try:
            parsed = datetime.strptime(s, fmt)
            return _transform_date_iso(parsed)
        except ValueError:
            continue
    from dateutil import parser as _dateutil_parser
    parsed = _dateutil_parser.parse(s, dayfirst=False)
    return _transform_date_iso(parsed)


# The single source of truth for the transform vocabulary. The json_schema
# enum, the prompt's ALLOWED TRANSFORMS list, and the validator all derive
# from these keys — adding a transform means adding a tested function here.
_MAPPING_TRANSFORMS = {
    "none": _transform_none,
    "text": _transform_text,
    "number": _transform_number,
    "date_iso": _transform_date_iso,
}

_TRANSFORM_DESCRIPTIONS = {
    "none": "write the value unchanged (use when unsure)",
    "text": "convert the value to plain text",
    "number": "numeric value; strips currency symbols and thousands separators",
    "date_iso": "date rendered as ISO YYYY-MM-DD",
}


def _normalize_transform_name(name) -> Optional[str]:
    """Fold formatting variants ('date-iso', 'Date ISO') onto the canonical
    id. Returns None when the name is not in the vocabulary even after
    folding — the caller degrades it to 'none'. Format normalisation only;
    no semantic guessing."""
    if not isinstance(name, str):
        return None
    key = re.sub(r"[\s\-]+", "_", name.strip().lower())
    return key if key in _MAPPING_TRANSFORMS else None


# Spec cache: one generation per distinct (field set, column set, instructions,
# context) per process. Failed generations are cached too, so a persistently
# failing schema degrades to the legacy path without re-paying a model call
# per row.
_SPEC_FAILED = object()
_MAPPING_SPEC_CACHE: "OrderedDict[tuple, Any]" = OrderedDict()
_MAPPING_SPEC_CACHE_LOCK = threading.Lock()
_MAPPING_SPEC_CACHE_MAX = 128


def _mapping_spec_enabled() -> bool:
    return os.getenv("EXCEL_AI_MAPPING_SPEC", "true").strip().lower() in (
        "true", "1", "t", "y", "yes")


def _spec_cache_get(key):
    with _MAPPING_SPEC_CACHE_LOCK:
        return _MAPPING_SPEC_CACHE.get(key)


def _spec_cache_put(key, value):
    with _MAPPING_SPEC_CACHE_LOCK:
        _MAPPING_SPEC_CACHE[key] = value
        while len(_MAPPING_SPEC_CACHE) > _MAPPING_SPEC_CACHE_MAX:
            _MAPPING_SPEC_CACHE.popitem(last=False)


def clear_mapping_spec_cache() -> None:
    with _MAPPING_SPEC_CACHE_LOCK:
        _MAPPING_SPEC_CACHE.clear()


def _build_mapping_spec_response_format(field_names: List[str], column_names: List[str]) -> Dict:
    """Strict json_schema whose enums are the ACTUAL field/column names and
    transform ids — constrained decoding cannot emit a variant spelling."""
    return {
        "type": "json_schema",
        "json_schema": {
            "name": "excel_mapping_spec",
            "strict": True,
            "schema": {
                "type": "object",
                "properties": {
                    "mappings": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "source": {"type": "string", "enum": list(field_names)},
                                "target": {"type": "string", "enum": list(column_names)},
                                "transform": {"type": "string",
                                              "enum": list(_MAPPING_TRANSFORMS.keys())},
                            },
                            "required": ["source", "target", "transform"],
                            "additionalProperties": False,
                        },
                    },
                },
                "required": ["mappings"],
                "additionalProperties": False,
            },
        },
    }


def _validate_mapping_spec(raw: Any, field_names: List[str], column_names: List[str]):
    """Deterministic re-validation of the model's spec (defense in depth
    behind the constrained decoding). Unknown sources/targets are dropped,
    unknown transforms degrade to 'none', duplicate targets keep the first —
    every repair is recorded as a warning."""
    warnings: List[str] = []
    fields_lower = {str(f).lower(): str(f) for f in field_names}
    cols_lower = {str(c).lower(): str(c) for c in column_names}

    entries = raw.get("mappings") if isinstance(raw, dict) else None
    if not isinstance(entries, list):
        return [], ["mapping spec response had no 'mappings' list"]

    mappings: List[Dict[str, str]] = []
    seen_targets = set()
    for entry in entries:
        if not isinstance(entry, dict):
            warnings.append(f"dropped non-object mapping entry: {entry!r}")
            continue
        source = fields_lower.get(str(entry.get("source", "")).lower())
        target = cols_lower.get(str(entry.get("target", "")).lower())
        if source is None or target is None:
            warnings.append(
                f"dropped mapping with unknown source/target: {entry.get('source')!r} -> {entry.get('target')!r}")
            continue
        if target in seen_targets:
            warnings.append(f"duplicate target column {target!r} - kept the first mapping")
            continue
        transform = _normalize_transform_name(entry.get("transform"))
        if transform is None:
            warnings.append(
                f"unknown transform {entry.get('transform')!r} for {target!r} - value will be written unchanged")
            transform = "none"
        seen_targets.add(target)
        mappings.append({"source": source, "target": target, "transform": transform})

    if not mappings:
        warnings.append("mapping spec validation left no usable mappings")
    return mappings, warnings


def _generate_mapping_spec(sample_records: List[Dict], field_names: List[str],
                           column_names: List[str], ai_instructions: str,
                           source_context: str):
    """ONE mini-model call producing the reusable mapping spec. Tries strict
    json_schema first (constrained decoding); falls back to json_object once.
    Returns (mappings, warnings) or (None, warnings) on failure."""
    transforms_doc = "\n".join(
        f"- {name}: {_TRANSFORM_DESCRIPTIONS[name]}" for name in _MAPPING_TRANSFORMS)
    samples = sample_records[:3]

    user_prompt = f"""Create a mapping specification from source record fields to Excel columns.

SOURCE FIELD NAMES:
{json.dumps(field_names)}

TARGET COLUMNS:
{json.dumps(column_names)}

SAMPLE RECORDS (for context only - map FIELD NAMES, not these values):
{json.dumps(samples, indent=2, default=str)[:3000]}

ALLOWED TRANSFORMS:
{transforms_doc}

{f"SOURCE CONTEXT: {source_context}" if source_context else ""}
{f"ADDITIONAL INSTRUCTIONS: {ai_instructions}" if ai_instructions else ""}

Rules:
1. Map semantically equivalent fields (e.g. "company" -> "Customer Name", "amt" -> "Amount")
2. Use each target column at most once; omit source fields that have no suitable column
3. Choose the transform from ALLOWED TRANSFORMS that fits the sample values; use "none" when unsure

Return ONLY a JSON object in this exact format:
{{"mappings": [{{"source": "field_name", "target": "Column Name", "transform": "none"}}]}}"""

    system_prompt = sysprompts.WORKFLOW_EXCEL_MAPPING_SPEC_SYSTEM
    mini = globals().get("azureMiniQuickPrompt")
    if mini is None:
        return None, ["azureMiniQuickPrompt unavailable - mapping spec disabled"]

    response = None
    attempts = [_build_mapping_spec_response_format(field_names, column_names),
                {"type": "json_object"}]
    for response_format in attempts:
        try:
            _count_llm_call("mapping_spec")
            response = mini(user_prompt, system_prompt, temp=0,
                            response_format=response_format)
            break
        except Exception as e:
            logger.warning(
                f"Mapping spec call failed with response_format="
                f"{response_format.get('type')}: {e}")
    if response is None:
        return None, ["mapping spec generation failed on both response formats"]

    try:
        cleaned = response.strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```json?\s*", "", cleaned)
            cleaned = re.sub(r"\s*```$", "", cleaned)
        parsed = json.loads(cleaned)
    except (json.JSONDecodeError, AttributeError) as e:
        logger.warning(f"Mapping spec response was not valid JSON: {e}")
        return None, [f"mapping spec response was not valid JSON: {e}"]

    mappings, warnings = _validate_mapping_spec(parsed, field_names, column_names)
    if not mappings:
        return None, warnings
    logger.info(
        f"Mapping spec generated: {len(mappings)} mapping(s) for "
        f"{len(field_names)} field(s) -> {len(column_names)} column(s)")
    return mappings, warnings


def _apply_mapping_spec(records: List[Dict], mappings: List[Dict[str, str]]) -> Dict:
    """Apply a validated spec to every record in pure Python. A transform that
    raises writes the RAW value for that cell (fail closed) and is counted in
    the warnings."""
    rows: List[Dict[str, Any]] = []
    failed_cells = 0
    for rec in records:
        rec_lower = {str(k).lower(): v for k, v in rec.items()}
        row: Dict[str, Any] = {}
        for m in mappings:
            value = rec[m["source"]] if m["source"] in rec else rec_lower.get(m["source"].lower())
            if isinstance(value, dict) and "value" in value:
                # write_extraction_to_excel's field-definition context shape:
                # {field: {"value": ..., "description": ...}} - the cell gets
                # the value, same as the legacy AI mapper produced.
                value = value["value"]
            if value is None:
                row[m["target"]] = None
                continue
            try:
                row[m["target"]] = _MAPPING_TRANSFORMS[m["transform"]](value)
            except Exception:
                row[m["target"]] = value
                failed_cells += 1
        rows.append(row)
    warnings = []
    if failed_cells:
        warnings.append(
            f"{failed_cells} value(s) failed their transform and were written unchanged")
    return {"rows": rows, "warnings": warnings}


def _map_via_spec(raw_data: Any, schema: Dict, ai_instructions: str,
                  source_context: str) -> Optional[Dict]:
    """Table mapping via a cached spec. Returns the mapped rows, or None to
    fall through to the legacy per-call AI mapper (unstructured input, spec
    disabled, or generation failure)."""
    if not _mapping_spec_enabled():
        return None
    records = raw_data if isinstance(raw_data, list) else [raw_data]
    if not records or not all(isinstance(r, dict) and r for r in records):
        return None                          # unstructured input -> legacy path
    cols = (schema or {}).get("columns") or []
    column_names = [c.get("name", c) if isinstance(c, dict) else c for c in cols]
    column_names = [str(c) for c in column_names if c is not None and str(c).strip() != ""]
    if not column_names:
        return None
    field_names = sorted({str(k) for rec in records for k in rec.keys()})
    if not field_names:
        return None

    cache_key = (frozenset(field_names), tuple(column_names),
                 ai_instructions or "", source_context or "")
    cached = _spec_cache_get(cache_key)
    if cached is _SPEC_FAILED:
        return None
    if cached is not None:
        result = _apply_mapping_spec(records, cached)
        logger.debug(
            f"Mapping spec cache hit: applied {len(cached)} mapping(s) to "
            f"{len(records)} record(s) without a model call")
        return result

    mappings, gen_warnings = _generate_mapping_spec(
        records, field_names, column_names, ai_instructions, source_context)
    if not mappings:
        _spec_cache_put(cache_key, _SPEC_FAILED)
        logger.warning(
            "Mapping spec generation failed - falling back to legacy AI "
            f"mapping for this schema: {gen_warnings}")
        return None
    _spec_cache_put(cache_key, mappings)
    result = _apply_mapping_spec(records, mappings)
    result["warnings"] = gen_warnings + result.get("warnings", [])
    for warning in result["warnings"]:
        logger.warning(f"Mapping spec warning: {warning}")
    return result


def map_data_to_schema(
    raw_data: Any,
    schema: Dict,
    ai_instructions: str = None,
    source_context: str = None
) -> Dict:
    """
    Use AI to map raw input data to the target schema.
    
    Args:
        raw_data: Input data (string, dict, list, or any extracted content)
        schema: Target schema from detect_template_schema or user-defined
        ai_instructions: Additional instructions for the AI
        source_context: Description of where the data came from
    
    Returns:
        For table type:
        {
            "rows": [
                {"Column A": "value1", "Column B": "value2"},
                ...
            ]
        }
        
        For form type:
        {
            "cells": {
                "B2": "Customer Name Value",
                "B3": "Account Number Value",
                ...
            }
        }
    """
    logger.info("Mapping data to schema using AI...")
    logger.debug(f"Schema type: {schema.get('type', 'unknown')}")
    logger.debug(f"Input data type: {type(raw_data)}")

    # FAST PATH (2026-08-03, approved by james). Skip the AI when it has nothing
    # to decide. Profiling showed 97% of a Database -> Excel export's wall clock
    # was ONE LLM call PER ROW asking how to map fields onto columns - and on an
    # append the columns were themselves created from those field names on row 1,
    # so the model was asked 1,000 times to map a->a, b->b. ~1.4s/row of nothing.
    # Fires only on an EXACT field-name/column-name set match, which is precisely
    # the case where the answer is forced; anything else (document extraction ->
    # a customer's template, missing or extra fields, duplicate column names)
    # returns None and falls through to the unchanged AI mapper.
    direct = _direct_table_mapping(raw_data, schema)
    if direct is not None:
        logger.info(
            "Field names match the target columns exactly (%d cols, %d row(s)) - "
            "mapped directly, skipping the AI call",
            len(direct.get("rows", [{}])[0]) if direct.get("rows") else 0,
            len(direct.get("rows", [])))
        return direct

    # MAPPING-SPEC PATH (2026-08-05, approved by james). For structured (dict)
    # records whose names DON'T match the columns, ask the mini model ONCE per
    # distinct schema for a reusable mapping spec and apply it in Python -
    # replacing the legacy one-primary-model-call-PER-ROW below. Returns None
    # (unstructured input, disabled via EXCEL_AI_MAPPING_SPEC=false, or spec
    # generation failure) to fall through to the unchanged legacy mapper.
    if (schema or {}).get("type", "table") == "table":
        spec_result = _map_via_spec(raw_data, schema, ai_instructions, source_context)
        if spec_result is not None:
            return spec_result

    if not quickPrompt:
        raise RuntimeError("AI features unavailable - quickPrompt not imported")
    
    # Convert data to string if needed
    if isinstance(raw_data, (dict, list)):
        data_str = json.dumps(raw_data, indent=2, default=str)
    else:
        data_str = str(raw_data)
    
    template_type = schema.get("type", "table")
    
    if template_type == "table":
        return _map_data_to_table(data_str, schema, ai_instructions, source_context)
    else:
        return _map_data_to_form(data_str, schema, ai_instructions, source_context)


def _map_data_to_table(data_str: str, schema: Dict, ai_instructions: str, source_context: str) -> Dict:
    """Map data to table-style schema."""
    logger.debug("Mapping data to table schema...")
    
    columns = schema.get("columns", [])
    column_names = [c.get("name", c) if isinstance(c, dict) else c for c in columns]
    
    logger.debug(f"Target columns: {column_names}")
    
    system_prompt = sysprompts.WORKFLOW_EXCEL_TABLE_MAPPING_SYSTEM

    user_prompt = f"""Extract data from the source and map it to these Excel columns.

SOURCE DATA:
{data_str}

TARGET COLUMNS:
{json.dumps(column_names)}

MAPPING GUIDANCE:
- Use the field descriptions to understand what each source field contains
- Match source fields to target columns based on semantic meaning

{f"SOURCE CONTEXT: {source_context}" if source_context else ""}
{f"ADDITIONAL INSTRUCTIONS: {ai_instructions}" if ai_instructions else ""}

Rules:
1. Map semantically equivalent fields (e.g., "company" -> "Customer Name", "amt" -> "Amount")
2. Convert data types appropriately:
   - Dates: Use ISO format (YYYY-MM-DD) or readable format (Month DD, YYYY)
   - Currency: Include numeric value only (no $ symbol), or as formatted string
   - Numbers: Use numeric values where appropriate
3. If a column's data cannot be found in the source, use null
4. If the source contains multiple records, return multiple rows
5. If the source is unstructured text, extract all relevant information

Return ONLY a JSON object in this exact format:
{{
    "rows": [
        {{"Column Name 1": "value1", "Column Name 2": "value2"}},
        {{"Column Name 1": "value3", "Column Name 2": "value4"}}
    ],
    "warnings": ["Optional: any fields that couldn't be mapped"]
}}

Return valid JSON only, no markdown or explanation."""

    try:
        _count_llm_call("legacy_table_mapping")
        response = quickPrompt(user_prompt, system=system_prompt, temp=0.0)
        logger.debug(f"AI mapping response: {response[:500]}...")
        
        # Clean up response
        response = response.strip()
        if response.startswith("```"):
            response = re.sub(r'^```json?\s*', '', response)
            response = re.sub(r'\s*```$', '', response)
        
        result = json.loads(response)
        
        # Log any warnings from AI
        warnings = result.get("warnings", [])
        for warning in warnings:
            logger.warning(f"AI mapping warning: {warning}")
        
        rows = result.get("rows", [])
        logger.info(f"Mapped data to {len(rows)} row(s)")
        
        return {"rows": rows, "warnings": warnings}
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse AI mapping response: {str(e)}")
        logger.error(f"Raw response: {response}")
        raise
    except Exception as e:
        logger.error(f"Data mapping error: {str(e)}", exc_info=True)
        raise


def _map_data_to_form(data_str: str, schema: Dict, ai_instructions: str, source_context: str) -> Dict:
    """Map data to form-style schema."""
    logger.debug("Mapping data to form schema...")
    
    cells_schema = schema.get("cells", {})
    
    # Build field descriptions for AI
    field_descriptions = []
    for label, info in cells_schema.items():
        cell = info.get("cell", "") if isinstance(info, dict) else info
        expected_type = info.get("expected_type", "string") if isinstance(info, dict) else "string"
        field_descriptions.append({
            "label": label,
            "cell": cell,
            "expected_type": expected_type
        })
    
    logger.debug(f"Target fields: {[f['label'] for f in field_descriptions]}")
    
    system_prompt = sysprompts.WORKFLOW_EXCEL_FORM_MAPPING_SYSTEM

    user_prompt = f"""Extract data from the source and map it to these Excel form fields.

SOURCE DATA:
{data_str[:4000]}

TARGET FORM FIELDS:
{json.dumps(field_descriptions, indent=2)}

{f"SOURCE CONTEXT: {source_context}" if source_context else ""}
{f"ADDITIONAL INSTRUCTIONS: {ai_instructions}" if ai_instructions else ""}

Rules:
1. Map semantically equivalent fields from the source to target labels
2. Convert data types appropriately for each field's expected_type
3. If a field's data cannot be found, use null
4. Extract all relevant information from unstructured text

Return ONLY a JSON object in this exact format:
{{
    "cells": {{
        "B2": "value for cell B2",
        "B3": "value for cell B3"
    }},
    "field_mapping": {{
        "Field Label": "extracted value"
    }},
    "warnings": ["Optional: any fields that couldn't be mapped"]
}}

Return valid JSON only, no markdown or explanation."""

    try:
        _count_llm_call("legacy_form_mapping")
        response = quickPrompt(user_prompt, system=system_prompt, temp=0.0)
        logger.debug(f"AI form mapping response: {response[:500]}...")
        
        # Clean up response
        response = response.strip()
        if response.startswith("```"):
            response = re.sub(r'^```json?\s*', '', response)
            response = re.sub(r'\s*```$', '', response)
        
        result = json.loads(response)
        
        # Log any warnings from AI
        warnings = result.get("warnings", [])
        for warning in warnings:
            logger.warning(f"AI mapping warning: {warning}")
        
        cells = result.get("cells", {})
        
        # If cells not directly provided, build from field_mapping
        if not cells and result.get("field_mapping"):
            for label, value in result["field_mapping"].items():
                if label in cells_schema:
                    cell_info = cells_schema[label]
                    cell_addr = cell_info.get("cell", "") if isinstance(cell_info, dict) else cell_info
                    if cell_addr:
                        cells[cell_addr] = value
        
        logger.info(f"Mapped data to {len(cells)} cell(s)")
        
        return {"cells": cells, "warnings": warnings}
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse AI form mapping response: {str(e)}")
        logger.error(f"Raw response: {response}")
        raise
    except Exception as e:
        logger.error(f"Form mapping error: {str(e)}", exc_info=True)
        raise


def populate_excel(
    output_path: str,
    mapped_data: Dict,
    schema: Dict,
    template_path: str = None,
    operation: str = "append",  # "append", "overwrite", "new_from_template"
    cell_formatting: Dict[str, Dict] = None
) -> Dict:
    """
    Write mapped data to Excel file.
    
    Args:
        output_path: Where to save the result
        mapped_data: Data from map_data_to_schema
        schema: Schema used for mapping
        template_path: Source template (for new_from_template operation)
        operation: How to write the data
            - "append": Add rows to existing file (or create if not exists)
            - "overwrite": Replace all data rows (keep headers for table type)
            - "new_from_template": Create new file from template, then populate
        cell_formatting: Optional dict of AI-suggested cell formatting from extraction
    
    Returns:
        {
            "success": True,
            "file_path": "/path/to/output.xlsx",
            "rows_written": 5,           # for table type
            "cells_populated": 12,       # for form type
            "sheet_name": "Sheet1",
            "operation": "append"
        }
    """
    logger.info(f"Populating Excel file: {output_path}")
    logger.debug(f"Operation: {operation}")
    logger.debug(f"Template path: {template_path}")
    
    template_type = schema.get("type", "table")
    sheet_name = schema.get("sheet_name", "Sheet1")
    
    try:
        # Determine source workbook
        if operation == "new_from_template" and template_path:
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            logger.debug(f"Creating new file from template: {template_path}")
            wb = load_workbook(template_path)
            
        elif os.path.exists(output_path):
            logger.debug(f"Opening existing file: {output_path}")
            wb = load_workbook(output_path)
            
        else:
            logger.debug("Creating new workbook")
            wb = Workbook()
            
            # If table type, add headers
            if template_type == "table":
                ws = wb.active
                ws.title = sheet_name
                columns = schema.get("columns", [])
                for col in columns:
                    col_name = col.get("name", col) if isinstance(col, dict) else col
                    col_idx = col.get("column_index", columns.index(col) + 1) if isinstance(col, dict) else columns.index(col) + 1
                    ws.cell(row=1, column=col_idx, value=col_name)
                logger.debug(f"Added headers: {[c.get('name', c) if isinstance(c, dict) else c for c in columns]}")
        
        # Get target worksheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif sheet_name:
            # Sheet name specified but doesn't exist yet — create it
            ws = wb.create_sheet(title=sheet_name)
            logger.debug(f"Created new sheet: {sheet_name}")
            # For table type, add headers to the new sheet
            if template_type == "table":
                columns = schema.get("columns", [])
                for col in columns:
                    col_name = col.get("name", col) if isinstance(col, dict) else col
                    col_idx = col.get("column_index", columns.index(col) + 1) if isinstance(col, dict) else columns.index(col) + 1
                    ws.cell(row=1, column=col_idx, value=col_name)
                logger.debug(f"Added headers to new sheet '{sheet_name}': {[c.get('name', c) if isinstance(c, dict) else c for c in columns]}")
        else:
            ws = wb.active
            if ws.title == "Sheet" and operation == "new_from_template":
                ws.title = sheet_name
        
        logger.debug(f"Working with sheet: {ws.title}")
        
        # Populate based on type
        start_row = None
        if template_type == "table":
            result, start_row = _populate_table(ws, mapped_data, schema, operation)
        else:
            result = _populate_form(ws, mapped_data, schema)

        # Apply AI-suggested cell formatting if provided
        formatting_result = None
        if cell_formatting and template_type == "table":
            try:
                data_start_row = start_row or 2
                rows_written = result.get("rows_written", 0)
                
                formatting_result = apply_cell_formatting(
                    ws=ws,
                    cell_formatting=cell_formatting,
                    schema=schema,
                    data_start_row=data_start_row,
                    rows_written=rows_written
                )
                logger.info(f"Cell formatting applied: {formatting_result}")
            except Exception as e:
                logger.error(f"Error applying cell formatting (data preserved): {str(e)}", exc_info=True)
                formatting_result = {"error": str(e), "cells_formatted": 0}
        
        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.debug(f"Created output directory: {output_dir}")
        
        # Save workbook
        wb.save(output_path)
        wb.close()
        
        logger.info(f"Successfully saved Excel file: {output_path}")
        
        result.update({
            "success": True,
            "file_path": output_path,
            "sheet_name": ws.title,
            "operation": operation
        })

        # Include formatting results if formatting was applied
        if formatting_result:
            result["formatting"] = formatting_result
        
        return result
        
    except Exception as e:
        logger.error(f"Error populating Excel file: {str(e)}", exc_info=True)
        raise


def _populate_table(ws, mapped_data: Dict, schema: Dict, operation: str) -> Dict:
    """Populate table-style worksheet."""
    logger.debug("Populating table...")
    
    rows = mapped_data.get("rows", [])
    columns = schema.get("columns", [])
    
    if not rows:
        logger.warning("No rows to write")
        return {"rows_written": 0}
    
    # Build column name to index mapping
    col_mapping = {}
    for col in columns:
        col_name = col.get("name", col) if isinstance(col, dict) else col
        col_idx = col.get("column_index", columns.index(col) + 1) if isinstance(col, dict) else columns.index(col) + 1
        col_mapping[col_name] = col_idx
    
    logger.debug(f"Column mapping: {col_mapping}")
    
    # For append operation, check if any columns in our schema need headers written
    # This handles the case where new columns (like assumptions/sources) were added to schema
    if operation == "append":
        for col in columns:
            col_name = col.get("name", col) if isinstance(col, dict) else col
            col_idx = col.get("column_index", columns.index(col) + 1) if isinstance(col, dict) else columns.index(col) + 1
            
            # Check if header row has this column
            existing_header = ws.cell(row=1, column=col_idx).value
            if existing_header is None or str(existing_header).strip() == "":
                # Write the header for this new column
                ws.cell(row=1, column=col_idx, value=col_name)
                logger.info(f"Added new column header '{col_name}' at column {col_idx}")
    
    # Determine starting row
    if operation == "overwrite":
        start_row = schema.get("data_start_row", 2)
        # Clear existing data
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col, value=None)
        logger.debug(f"Cleared existing data from row {start_row}")
    else:  # append (also the path taken by new_from_template)
        # Find the first free row by scanning UP for the last row that actually
        # contains data, instead of trusting ws.max_row + 1.
        #
        # openpyxl's ws.max_row counts trailing/phantom empty rows: a
        # hand-authored template often carries an empty row directly under the
        # header (or a <dimension> that spans past the last real row), so
        # max_row came back as 2 for a header-only sheet. start_row = max_row+1
        # then wrote the first record to row 3 and left a permanent blank row 2
        # under the header — which breaks sorting in Excel. Scanning for the
        # last non-empty row is robust to that.
        last_data_row = 0
        for r in range(ws.max_row, 0, -1):
            if any(ws.cell(row=r, column=c).value not in (None, "")
                   for c in range(1, ws.max_column + 1)):
                last_data_row = r
                break

        # last_data_row == 0 -> completely empty sheet
        # last_data_row == 1 -> only the header row is populated
        # otherwise          -> append immediately after the last real row
        # Row 1 is reserved for headers, so data never starts above row 2.
        start_row = max(last_data_row + 1, 2)
        logger.debug(
            f"APPEND: ws.max_row={ws.max_row}, last_data_row={last_data_row}, "
            f"start_row={start_row}")
    
    # Write rows
    rows_written = 0
    for row_data in rows:
        if not row_data:
            continue
            
        for col_name, value in row_data.items():
            if col_name in col_mapping:
                col_idx = col_mapping[col_name]
                ws.cell(row=start_row + rows_written, column=col_idx, value=value)
            else:
                logger.warning(f"Column '{col_name}' not found in schema, skipping")
        
        rows_written += 1
    
    logger.info(f"Wrote {rows_written} rows")
    
    return {"rows_written": rows_written}, start_row


def _populate_form(ws, mapped_data: Dict, schema: Dict) -> Dict:
    """Populate form-style worksheet."""
    logger.debug("Populating form...")
    
    cells = mapped_data.get("cells", {})
    
    if not cells:
        logger.warning("No cells to populate")
        return {"cells_populated": 0}
    
    cells_populated = 0
    
    for cell_addr, value in cells.items():
        # Skip None values AND empty strings to preserve existing data
        if value is None or (isinstance(value, str) and value.strip() == ''):
            logger.debug(f"Skipping cell {cell_addr} with empty/null value")
            continue
            
        try:
            ws[cell_addr] = value
            cells_populated += 1
            logger.debug(f"Set cell {cell_addr} = {str(value)[:50]}")
        except Exception as e:
            logger.warning(f"Failed to set cell {cell_addr}: {str(e)}")
    
    logger.info(f"Populated {cells_populated} cells")
    
    return {"cells_populated": cells_populated}


# Convenience function for common use case
def process_data_to_excel(
    raw_data: Any,
    output_path: str,
    template_path: str = None,
    schema_mode: str = "existing_template",  # or "ai_generated"
    ai_instructions: str = None,
    operation: str = "append",
    sheet_name: str = None
) -> Dict:
    """
    High-level convenience function to process data and write to Excel.
    
    This combines schema detection/generation, data mapping, and Excel population
    into a single call.
    
    Args:
        raw_data: Input data (any format)
        output_path: Where to save the Excel file
        template_path: Path to template (for existing_template mode)
        schema_mode: "existing_template" or "ai_generated"
        ai_instructions: Additional instructions for AI
        operation: "append", "overwrite", or "new_from_template"
        sheet_name: Target sheet name (optional)
    
    Returns:
        Result dict with success status and details
    """
    logger.info("=" * 60)
    logger.info("Starting Excel data processing")
    logger.info(f"Schema mode: {schema_mode}")
    logger.info(f"Operation: {operation}")
    logger.info(f"Output: {output_path}")
    logger.info("=" * 60)
    
    try:
        # Step 1: Get or generate schema
        if schema_mode == "existing_template":
            if not template_path:
                raise ValueError("template_path required for existing_template mode")
            schema = detect_template_schema(template_path, sheet_name)
        else:  # ai_generated
            schema = generate_schema_from_data(raw_data, ai_instructions)
        
        logger.info(f"Schema ready. Type: {schema.get('type')}")
        
        # Step 2: Map data to schema
        mapped_data = map_data_to_schema(
            raw_data, 
            schema, 
            ai_instructions=ai_instructions
        )
        
        logger.info("Data mapping complete")
        
        # Step 3: Populate Excel
        result = populate_excel(
            output_path=output_path,
            mapped_data=mapped_data,
            schema=schema,
            template_path=template_path,
            operation=operation
        )
        
        logger.info("Excel population complete")
        logger.info(f"Result: {json.dumps(result, indent=2)}")
        
        return result
        
    except Exception as e:
        logger.error(f"Error in process_data_to_excel: {str(e)}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "file_path": output_path
        }


def extraction_result_to_dataframe(
    extraction_result: Dict,
    include_assumptions: bool = False,
    include_sources: bool = False,
    include_confidence: bool = False,
    transpose: bool = True
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Convert populate_schema_with_claude output to a pandas DataFrame.
    
    Args:
        extraction_result: Output from populate_schema_with_claude
        include_assumptions: Whether to include assumptions column(s)
        include_sources: Whether to include source pages column(s)
        include_confidence: Whether to include confidence level column(s) (LOW/MED/HIGH)
        transpose: If True, fields become columns (one row). 
                   If False, fields become rows (Field, Value, ... columns)
    
    Returns:
        Tuple of (DataFrame, list of warnings)
    """
    logger.info("Converting extraction result to DataFrame")
    logger.debug(f"Options - include_assumptions: {include_assumptions}, "
                 f"include_sources: {include_sources}, include_confidence: {include_confidence}, transpose: {transpose}")
    
    fields = extraction_result.get("fields", {})
    global_assumptions = extraction_result.get("global_assumptions", [])
    warnings = []
    
    if not fields:
        logger.warning("No fields found in extraction result")
        warnings.append("No fields found in extraction result")
        return pd.DataFrame(), warnings
    
    logger.debug(f"Processing {len(fields)} fields")
    
    if transpose:
        # One row with field names as columns
        row_data = {}
        assumptions_data = {}
        sources_data = {}
        confidence_data = {}
        
        for field_name, field_info in fields.items():
            value = field_info.get("value")
            row_data[field_name] = value
            
            if include_assumptions:
                assumptions = field_info.get("assumptions", [])
                assumptions_data[f"{field_name}_assumptions"] = "; ".join(assumptions) if assumptions else ""
            
            if include_sources:
                sources = field_info.get("sources", [])
                source_pages = []
                for src in sources:
                    pages = src.get("pages", [])
                    if pages:
                        source_pages.extend([str(p) for p in pages])
                sources_data[f"{field_name}_sources"] = ", ".join(source_pages) if source_pages else ""
            
            if include_confidence:
                confidence = field_info.get("confidence", "")
                confidence_data[f"{field_name}_confidence"] = confidence if confidence else ""
        
        # Combine all data
        combined_data = {**row_data}
        if include_assumptions:
            combined_data.update(assumptions_data)
        if include_sources:
            combined_data.update(sources_data)
        if include_confidence:
            combined_data.update(confidence_data)
        
        df = pd.DataFrame([combined_data])
        
    else:
        # Multiple rows with Field, Value, Assumptions, Sources, Confidence columns
        rows = []
        for field_name, field_info in fields.items():
            row = {
                "Field": field_name,
                "Value": field_info.get("value")
            }
            
            if include_assumptions:
                assumptions = field_info.get("assumptions", [])
                row["Assumptions"] = "; ".join(assumptions) if assumptions else ""
            
            if include_sources:
                sources = field_info.get("sources", [])
                source_pages = []
                source_notes = []
                for src in sources:
                    pages = src.get("pages", [])
                    if pages:
                        source_pages.extend([str(p) for p in pages])
                    note = src.get("notes")
                    if note:
                        source_notes.append(note)
                
                row["Source Pages"] = ", ".join(source_pages) if source_pages else ""
                row["Source Notes"] = "; ".join(source_notes) if source_notes else ""
            
            if include_confidence:
                confidence = field_info.get("confidence", "")
                row["Confidence"] = confidence if confidence else ""
        
            rows.append(row)
        
        df = pd.DataFrame(rows)
    
    # Add global assumptions as a warning/note
    if global_assumptions:
        warnings.append(f"Global assumptions: {'; '.join(global_assumptions)}")
    
    logger.info(f"Created DataFrame with shape {df.shape}")
    logger.debug(f"Columns: {list(df.columns)}")
    
    return df, warnings


def apply_cell_formatting(
    ws,
    cell_formatting: Dict[str, Dict],
    schema: Dict,
    data_start_row: int,
    rows_written: int
) -> Dict:
    """
    Apply AI-suggested cell-level formatting to a worksheet.
    
    This applies formatting based on the cell_formatting dict returned by
    the AI extraction, which contains per-field formatting suggestions.
    
    Args:
        ws: openpyxl worksheet object
        cell_formatting: Dict from AI extraction result, e.g.:
            {
                "amount": {"fill": "#FFCDD2", "font_color": "#B71C1C", "bold": true, "reason": "..."},
                "status": {"fill": "#C8E6C9", "reason": "..."}
            }
        schema: The schema used for writing (contains column info)
        data_start_row: Row where data starts (typically 2)
        rows_written: Number of data rows written
        
    Returns:
        Dict with results: {"cells_formatted": int, "errors": []}
    """
    if not cell_formatting:
        logger.debug("No cell formatting to apply")
        return {"cells_formatted": 0, "errors": []}
    
    logger.info(f"Applying AI-suggested cell formatting for {len(cell_formatting)} field(s)")
    
    results = {
        "cells_formatted": 0,
        "errors": []
    }
    
    # Build column name to index mapping from schema
    columns = schema.get("columns", [])
    col_mapping = {}
    for col in columns:
        col_name = col.get("name", col) if isinstance(col, dict) else col
        col_idx = col.get("column_index", columns.index(col) + 1) if isinstance(col, dict) else columns.index(col) + 1
        col_mapping[col_name] = col_idx
    
    logger.debug(f"Column mapping for formatting: {col_mapping}")
    
    for field_name, format_spec in cell_formatting.items():
        try:
            if field_name not in col_mapping:
                logger.warning(f"Field '{field_name}' not found in columns, skipping formatting")
                continue
            
            col_idx = col_mapping[field_name]
            
            # Parse formatting spec
            fill = None
            font_kwargs = {}
            
            # Handle fill/background color
            fill_color = format_spec.get("fill") or format_spec.get("background")
            if fill_color:
                # Strip # if present
                fill_color = fill_color.lstrip("#")
                try:
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                except Exception as e:
                    logger.warning(f"Invalid fill color '{fill_color}' for field '{field_name}': {e}")
            
            # Handle font color
            font_color = format_spec.get("font_color") or format_spec.get("color")
            if font_color:
                font_color = font_color.lstrip("#")
                font_kwargs["color"] = font_color
            
            # Handle bold
            if format_spec.get("bold"):
                font_kwargs["bold"] = True
            
            # Handle italic
            if format_spec.get("italic"):
                font_kwargs["italic"] = True
            
            font = Font(**font_kwargs) if font_kwargs else None
            
            # Apply formatting to all data rows in this column
            for row in range(data_start_row, data_start_row + rows_written):
                cell = ws.cell(row=row, column=col_idx)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                results["cells_formatted"] += 1
            
            reason = format_spec.get("reason", "")
            logger.debug(f"Formatted field '{field_name}': fill={fill_color}, font_color={font_color}, "
                        f"bold={format_spec.get('bold')}, reason={reason}")
            
        except Exception as e:
            error_msg = f"Failed to format field '{field_name}': {str(e)}"
            logger.error(error_msg, exc_info=True)
            results["errors"].append(error_msg)
    
    logger.info(f"Cell formatting complete: {results['cells_formatted']} cells formatted")
    return results

def write_extraction_to_excel(
    extraction_result: Dict,
    output_path: str,
    template_path: str = None,
    operation: str = "new",  # "new", "append", "new_from_template"
    include_assumptions: bool = False,
    include_sources: bool = False,
    include_confidence: bool = False,
    sheet_name: str = None,
    field_mapping: Dict[str, str] = None,
    ai_mapping_instructions: str = None,
    field_definitions: List[Dict] = None 
) -> Dict:
    """
    Write extraction results to an Excel file.
    
    Args:
        extraction_result: Output from populate_schema_with_claude
        output_path: Where to save the Excel file
        template_path: Path to template (for new_from_template or append operations)
        operation: 
            - "new": Create new file with auto-generated columns
            - "append": Append to existing file
            - "new_from_template": Create new file from template, then populate
        include_assumptions: Include assumptions columns
        include_sources: Include source pages columns
        include_confidence: Include confidence level columns (LOW/MED/HIGH)
        sheet_name: Target sheet name (optional)
        field_mapping: Manual field->column mapping, e.g. {"customer_name": "Retailer Name"}
        ai_mapping_instructions: Instructions for AI to map fields to template columns
    
    Returns:
        Result dict with success status and details
    """
    logger.info("=" * 60)
    logger.info("Writing extraction result to Excel")
    logger.info(f"Output path: {output_path}")
    logger.info(f"Operation: {operation}")
    logger.info(f"Template: {template_path}")
    logger.info("=" * 60)
    
    try:
        # Convert extraction result to simple values dict for mapping
        fields = extraction_result.get("fields", {})
        values_dict = {k: v.get("value") for k, v in fields.items()}
        
        logger.debug(f"Extracted values: {values_dict}")

        # Extract AI-suggested cell formatting if present
        cell_formatting = extraction_result.get("cell_formatting", {})
        if cell_formatting:
            logger.info(f"AI provided cell formatting for {len(cell_formatting)} field(s)")
            logger.debug(f"Cell formatting details: {cell_formatting}")

        # Transform cell_formatting keys using field_mapping if provided
        # The AI returns formatting with extraction field names (e.g., "customer")
        # but Excel columns may have different names (e.g., "CUSTOMER")
        if cell_formatting and field_mapping:
            transformed_formatting = {}
            for field_name, format_spec in cell_formatting.items():
                # Look up the mapped column name
                mapped_name = field_mapping.get(field_name)
                if mapped_name:
                    transformed_formatting[mapped_name] = format_spec
                    logger.debug(f"Transformed formatting key '{field_name}' -> '{mapped_name}'")
                else:
                    # Keep original name if no mapping exists
                    transformed_formatting[field_name] = format_spec
            cell_formatting = transformed_formatting
            logger.debug(f"Transformed cell_formatting keys using field_mapping")
        
        # Determine if we need to map to an existing schema
        use_template_schema = operation in ["append", "new_from_template"] and template_path

        # For append without template, read schema from existing output file
        if operation == "append" and not template_path:
            if os.path.exists(output_path):
                use_template_schema = True
                template_path = output_path  # Use existing file as the schema source
            else:
                return {
                    "success": False,
                    "error": f"Cannot append - file does not exist: {output_path}",
                    "file_path": output_path
                }
        
        if use_template_schema:
            logger.info("Using template schema for mapping")
            
            # Detect template schema
            template_schema = detect_template_schema(template_path, sheet_name)
            logger.debug(f"Template schema type: {template_schema.get('type')}")
            
            if template_schema.get("type") == "table":
                # Get template columns
                template_columns = [
                    c.get("name", c) if isinstance(c, dict) else c 
                    for c in template_schema.get("columns", [])
                ]
                logger.debug(f"Template columns: {template_columns}")
                
                # Build case-insensitive lookup: lowercase -> actual template column name
                template_columns_lower = {col.lower(): col for col in template_columns}
                logger.debug(f"Template columns (case-insensitive lookup): {template_columns_lower}")
                
                # Track columns that need to be created (not in template)
                columns_to_create = []
                
                # Find the maximum existing column index for adding new columns
                max_col_idx = max(
                    (c.get("column_index", i+1) if isinstance(c, dict) else i+1)
                    for i, c in enumerate(template_schema.get("columns", []))
                ) if template_schema.get("columns") else 0
                
                # Build set of existing column names for quick lookup (will be updated as we add columns)
                existing_col_names = set(template_columns)
                existing_col_names_lower = set(template_columns_lower.keys())
                
                # Normalize field_mapping values to match actual template column names
                # This handles case mismatches between user's mapping and template columns
                normalized_field_mapping = None
                if field_mapping:
                    normalized_field_mapping = {}
                    for extract_field, mapped_col in field_mapping.items():
                        mapped_col_lower = mapped_col.lower()
                        if mapped_col_lower in template_columns_lower:
                            # Use the actual template column name (preserves original case)
                            actual_col_name = template_columns_lower[mapped_col_lower]
                            normalized_field_mapping[extract_field] = actual_col_name
                            if actual_col_name != mapped_col:
                                logger.info(f"Normalized field mapping: '{extract_field}' -> '{mapped_col}' -> '{actual_col_name}'")
                        else:
                            # Column doesn't exist - will create it with the user's specified name
                            normalized_field_mapping[extract_field] = mapped_col
                            
                            # Check if we've already queued this column for creation (case-insensitive)
                            if mapped_col_lower not in existing_col_names_lower:
                                max_col_idx += 1
                                columns_to_create.append({
                                    "name": mapped_col,
                                    "column_index": max_col_idx,
                                    "column_letter": get_column_letter(max_col_idx)
                                })
                                existing_col_names.add(mapped_col)
                                existing_col_names_lower.add(mapped_col_lower)
                                template_columns_lower[mapped_col_lower] = mapped_col
                                logger.info(f"Will create new column '{mapped_col}' at index {max_col_idx} for field '{extract_field}'")
                    
                    logger.debug(f"Normalized field mapping: {normalized_field_mapping}")
                
                # Map extraction fields to template columns
                if normalized_field_mapping:
                    # Use normalized manual mapping
                    logger.info("Using manual field mapping (normalized)")
                    mapped_row = {}
                    for field_name, value in values_dict.items():
                        if field_name in normalized_field_mapping:
                            mapped_row[normalized_field_mapping[field_name]] = value
                        elif field_name in template_columns:
                            mapped_row[field_name] = value
                        elif field_name.lower() in template_columns_lower:
                            # Case-insensitive fallback for direct field names
                            mapped_row[template_columns_lower[field_name.lower()]] = value
                        else:
                            # Field not in mapping and not in template - create column with field name
                            field_name_lower = field_name.lower()
                            if field_name_lower not in existing_col_names_lower:
                                max_col_idx += 1
                                columns_to_create.append({
                                    "name": field_name,
                                    "column_index": max_col_idx,
                                    "column_letter": get_column_letter(max_col_idx)
                                })
                                existing_col_names.add(field_name)
                                existing_col_names_lower.add(field_name_lower)
                                template_columns_lower[field_name_lower] = field_name
                                logger.info(f"Will create new column '{field_name}' at index {max_col_idx} (unmapped field)")
                            mapped_row[field_name] = value
                    
                    # Update field_mapping reference to use normalized version for metadata columns
                    field_mapping = normalized_field_mapping
                    
                else:
                    # Use AI mapping
                    logger.info("Using AI to map fields to template columns")
                    mapping_instructions = ai_mapping_instructions or \
                        "Map the extracted field names to the template column names semantically."
                    
                    # Build enhanced context with field descriptions
                    if field_definitions:
                        field_context = {
                            fd.get('name'): {
                                'value': values_dict.get(fd.get('name')),
                                'description': fd.get('description', '')
                            }
                            for fd in field_definitions
                            if fd.get('name') in values_dict
                        }
                    else:
                        field_context = values_dict
                    
                    mapped_data = map_data_to_schema(
                        raw_data=field_context,
                        schema=template_schema,
                        ai_instructions=mapping_instructions
                    )
                    
                    if mapped_data.get("rows"):
                        mapped_row = mapped_data["rows"][0]
                        
                        # Infer field_mapping from AI result by matching values
                        # This allows metadata columns to use consistent naming with mapped columns
                        inferred_mapping = {}
                        for extract_field, extract_value in values_dict.items():
                            for col_name, col_value in mapped_row.items():
                                # Match by value (handle type differences)
                                if str(extract_value) == str(col_value) or extract_value == col_value:
                                    inferred_mapping[extract_field] = col_name
                                    break
                        
                        if inferred_mapping:
                            field_mapping = inferred_mapping
                            logger.info(f"Inferred field mapping from AI result: {field_mapping}")
                        else:
                            logger.debug("Could not infer field mapping from AI result - using original field names for metadata")
                    else:
                        mapped_row = {}
                        logger.warning("AI mapping returned no rows")
                
                # Add any new columns to the schema (before metadata columns)
                if columns_to_create:
                    template_schema["columns"] = template_schema.get("columns", []) + columns_to_create
                    logger.info(f"Extended schema with {len(columns_to_create)} new column(s): {[c['name'] for c in columns_to_create]}")
                
                # Add assumptions, sources, and confidence if requested
                logger.info(f"Include assumptions: {include_assumptions}, Include sources: {include_sources}, Include confidence: {include_confidence}")
                
                if include_assumptions or include_sources or include_confidence:
                    # Use the existing_col_names and max_col_idx we've been maintaining
                    # (they include any new columns we created above)
                    logger.debug(f"Current column names for metadata check: {existing_col_names}")
                    
                    # Track new metadata columns to add to schema
                    new_columns = []
                    
                    for field_name, field_info in fields.items():
                        # Find the mapped column name
                        col_name = field_mapping.get(field_name, field_name) if field_mapping else field_name
                        
                        if include_assumptions:
                            assumptions = field_info.get("assumptions", [])
                            assumptions_col_name = f"{col_name}_assumptions"
                            assumptions_col_name_lower = assumptions_col_name.lower()
                            if assumptions:
                                mapped_row[assumptions_col_name] = "; ".join(assumptions)
                            
                            # Add column to schema if it doesn't exist (case-insensitive check)
                            if assumptions_col_name_lower not in existing_col_names_lower:
                                max_col_idx += 1
                                new_columns.append({
                                    "name": assumptions_col_name,
                                    "column_index": max_col_idx,
                                    "column_letter": get_column_letter(max_col_idx)
                                })
                                existing_col_names.add(assumptions_col_name)
                                existing_col_names_lower.add(assumptions_col_name_lower)
                                logger.debug(f"Added new schema column: {assumptions_col_name} at index {max_col_idx}")
                        
                        if include_sources:
                            sources = field_info.get("sources", [])
                            source_pages = []
                            for src in sources:
                                pages = src.get("pages", [])
                                source_pages.extend([str(p) for p in pages])
                            
                            sources_col_name = f"{col_name}_sources"
                            sources_col_name_lower = sources_col_name.lower()
                            if source_pages:
                                mapped_row[sources_col_name] = ", ".join(source_pages)
                            
                            # Add column to schema if it doesn't exist (case-insensitive check)
                            if sources_col_name_lower not in existing_col_names_lower:
                                max_col_idx += 1
                                new_columns.append({
                                    "name": sources_col_name,
                                    "column_index": max_col_idx,
                                    "column_letter": get_column_letter(max_col_idx)
                                })
                                existing_col_names.add(sources_col_name)
                                existing_col_names_lower.add(sources_col_name_lower)
                                logger.debug(f"Added new schema column: {sources_col_name} at index {max_col_idx}")
                        
                        if include_confidence:
                            confidence = field_info.get("confidence", "")
                            confidence_col_name = f"{col_name}_confidence"
                            confidence_col_name_lower = confidence_col_name.lower()
                            if confidence:
                                mapped_row[confidence_col_name] = confidence
                            
                            # Add column to schema if it doesn't exist (case-insensitive check)
                            if confidence_col_name_lower not in existing_col_names_lower:
                                max_col_idx += 1
                                new_columns.append({
                                    "name": confidence_col_name,
                                    "column_index": max_col_idx,
                                    "column_letter": get_column_letter(max_col_idx)
                                })
                                existing_col_names.add(confidence_col_name)
                                existing_col_names_lower.add(confidence_col_name_lower)
                                logger.debug(f"Added new schema column: {confidence_col_name} at index {max_col_idx}")
                    
                    # Extend template schema with new columns
                    if new_columns:
                        template_schema["columns"] = template_schema.get("columns", []) + new_columns
                        logger.info(f"Extended schema with {len(new_columns)} new assumption/source/confidence column(s)")
                
                # Prepare data for populate_excel
                mapped_data_final = {"rows": [mapped_row]}
                
            else:
                # Form-style template - use AI mapping for cells
                logger.info("Template is form-style, using AI cell mapping")
                mapped_data_final = map_data_to_schema(
                    raw_data=values_dict,
                    schema=template_schema,
                    ai_instructions=ai_mapping_instructions
                )
            
            # Write to Excel
            result = populate_excel(
                output_path=output_path,
                mapped_data=mapped_data_final,
                schema=template_schema,
                template_path=template_path,
                operation="append" if operation == "append" else "new_from_template",
                cell_formatting=cell_formatting
            )
            
        else:
            # New file without template - use extraction fields as columns
            logger.info("Creating new file from extraction fields")
            
            # Convert to DataFrame
            df, warnings = extraction_result_to_dataframe(
                extraction_result,
                include_assumptions=include_assumptions,
                include_sources=include_sources,
                include_confidence=include_confidence,
                transpose=True  # Fields as columns
            )
            
            # Build schema from DataFrame columns
            schema = {
                "type": "table",
                "columns": [{"name": col, "column_index": i+1, "column_letter": get_column_letter(i+1)}
                           for i, col in enumerate(df.columns)],
                "data_start_row": 2
            }
            # Preserve sheet_name so populate_excel uses the correct sheet
            if sheet_name:
                schema["sheet_name"] = sheet_name

            # Convert DataFrame row to dict
            if len(df) > 0:
                row_data = df.iloc[0].to_dict()
                mapped_data_final = {"rows": [row_data]}
            else:
                mapped_data_final = {"rows": []}
            
            # Write to Excel
            result = populate_excel(
                output_path=output_path,
                mapped_data=mapped_data_final,
                schema=schema,
                template_path=None,
                operation="append" if operation == "append" else "overwrite",
                cell_formatting=cell_formatting
            )
            
            # Add warnings
            if warnings:
                result["warnings"] = result.get("warnings", []) + warnings
        
        logger.info(f"Excel write complete: {result}")
        return result
        
    except Exception as e:
        logger.error(f"Error writing extraction to Excel: {str(e)}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "file_path": output_path
        }
