import json
import logging
import os
import threading
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, g
from flask_login import current_user
from werkzeug.utils import secure_filename

from role_decorators import api_key_or_session_required
from compliance_engine import ComplianceEngine
from compliance_comparison import ComplianceComparison

logger = logging.getLogger("ComplianceRoutes")

compliance_bp = Blueprint("compliance", __name__)

_engine = ComplianceEngine()
_comparison = ComplianceComparison()


def _get_user_id():
    if getattr(g, "auth_method", None) == "api_key":
        return 1
    if hasattr(current_user, "id"):
        return current_user.id
    return 1


# ============================================================================
# UI Page
# ============================================================================

@compliance_bp.route("/compliance")
@api_key_or_session_required(min_role=2)
def compliance_page():
    return render_template("compliance_management.html")


@compliance_bp.route("/compliance/schemas")
@api_key_or_session_required(min_role=2)
def compliance_schemas_page():
    return render_template("compliance_schemas.html")


# ============================================================================
# Retailer CRUD
# ============================================================================

@compliance_bp.route("/api/compliance/retailers", methods=["GET"])
@api_key_or_session_required(min_role=2)
def list_retailers():
    retailers = _engine.get_retailers()
    for r in retailers:
        for k, v in r.items():
            if isinstance(v, datetime):
                r[k] = v.isoformat()
    return jsonify({"retailers": retailers})


@compliance_bp.route("/api/compliance/retailers", methods=["POST"])
@api_key_or_session_required(min_role=2)
def create_retailer():
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Retailer name is required"}), 400
    retailer_id = _engine.create_retailer(
        name=name,
        notes=data.get("notes", ""),
        created_by=_get_user_id(),
    )
    return jsonify({"retailer_id": retailer_id, "name": name}), 201


@compliance_bp.route("/api/compliance/retailers/<int:retailer_id>", methods=["GET"])
@api_key_or_session_required(min_role=2)
def get_retailer(retailer_id):
    retailer = _engine.get_retailer(retailer_id)
    if not retailer:
        return jsonify({"error": "Retailer not found"}), 404
    for k, v in retailer.items():
        if isinstance(v, datetime):
            retailer[k] = v.isoformat()
    return jsonify(retailer)


@compliance_bp.route("/api/compliance/retailers/<int:retailer_id>", methods=["PUT"])
@api_key_or_session_required(min_role=2)
def update_retailer(retailer_id):
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Retailer name is required"}), 400
    ok = _engine.update_retailer(retailer_id, name, data.get("notes", ""))
    if not ok:
        return jsonify({"error": "Retailer not found"}), 404
    return jsonify({"success": True})


@compliance_bp.route("/api/compliance/retailers/<int:retailer_id>", methods=["DELETE"])
@api_key_or_session_required(min_role=2)
def delete_retailer(retailer_id):
    ok = _engine.delete_retailer(retailer_id)
    if not ok:
        return jsonify({"error": "Retailer not found"}), 404
    return jsonify({"success": True})


# ============================================================================
# Document Sets
# ============================================================================

@compliance_bp.route(
    "/api/compliance/retailers/<int:retailer_id>/sets", methods=["GET"]
)
@api_key_or_session_required(min_role=2)
def list_document_sets(retailer_id):
    sets = _engine.get_document_sets(retailer_id)
    for s in sets:
        for k, v in s.items():
            if isinstance(v, datetime):
                s[k] = v.isoformat()
    return jsonify({"sets": sets})


@compliance_bp.route(
    "/api/compliance/retailers/<int:retailer_id>/sets", methods=["POST"]
)
@api_key_or_session_required(min_role=2)
def create_document_set(retailer_id):
    data = request.get_json() or {}
    category = data.get("category", "").strip()
    if not category:
        return jsonify({"error": "Category is required"}), 400
    agent_id = data.get("agent_id")
    if agent_id in ("", None):
        agent_id = None
    else:
        try:
            agent_id = int(agent_id)
        except (ValueError, TypeError):
            return jsonify({"error": "agent_id must be numeric"}), 400
    set_id = _engine.create_document_set(
        retailer_id=retailer_id,
        category=category,
        description=data.get("description", ""),
        agent_id=agent_id,
    )
    return jsonify({"set_id": set_id, "category": category}), 201


@compliance_bp.route("/api/compliance/sets/<int:set_id>", methods=["GET"])
@api_key_or_session_required(min_role=2)
def get_document_set(set_id):
    s = _engine.get_document_set(set_id)
    if not s:
        return jsonify({"error": "Document set not found"}), 404
    return jsonify(s)


@compliance_bp.route("/api/compliance/sets/<int:set_id>", methods=["PUT"])
@api_key_or_session_required(min_role=2)
def update_document_set(set_id):
    data = request.get_json() or {}
    description = data.get("description")
    agent_id = data.get("agent_id")
    clear_agent = False
    if agent_id in ("", None) and "agent_id" in data:
        clear_agent = True
        agent_id = None
    elif agent_id is not None:
        try:
            agent_id = int(agent_id)
        except (ValueError, TypeError):
            return jsonify({"error": "agent_id must be numeric"}), 400

    extraction_schema_id = data.get("extraction_schema_id")
    clear_extraction_schema = False
    if extraction_schema_id in ("", None) and "extraction_schema_id" in data:
        clear_extraction_schema = True
        extraction_schema_id = None
    elif extraction_schema_id is not None:
        try:
            extraction_schema_id = int(extraction_schema_id)
        except (ValueError, TypeError):
            return jsonify({"error": "extraction_schema_id must be numeric"}), 400

    extraction_workflow_id = data.get("extraction_workflow_id")
    clear_extraction_workflow = False
    if extraction_workflow_id in ("", None) and "extraction_workflow_id" in data:
        clear_extraction_workflow = True
        extraction_workflow_id = None
    elif extraction_workflow_id is not None:
        try:
            extraction_workflow_id = int(extraction_workflow_id)
        except (ValueError, TypeError):
            return jsonify({"error": "extraction_workflow_id must be numeric"}), 400

    ok = _engine.update_document_set(
        set_id=set_id,
        description=description,
        agent_id=agent_id,
        clear_agent=clear_agent,
        extraction_schema_id=extraction_schema_id,
        clear_extraction_schema=clear_extraction_schema,
        extraction_workflow_id=extraction_workflow_id,
        clear_extraction_workflow=clear_extraction_workflow,
    )
    if not ok:
        return jsonify({"error": "Document set not found or no changes"}), 404
    return jsonify({"success": True})


# ============================================================================
# Versions
# ============================================================================

@compliance_bp.route("/api/compliance/sets/<int:set_id>/versions", methods=["GET"])
@api_key_or_session_required(min_role=2)
def list_versions(set_id):
    versions = _engine.get_versions(set_id)
    for v in versions:
        for k, val in v.items():
            if isinstance(val, datetime):
                v[k] = val.isoformat()
    return jsonify({"versions": versions})


@compliance_bp.route("/api/compliance/sets/<int:set_id>/upload", methods=["POST"])
@api_key_or_session_required(min_role=2)
def upload_document(set_id):
    """Upload a compliance document for processing.

    Processing happens in a background daemon thread so large documents
    (100+ pages, multi-hour Claude extraction) don't tie up the browser
    AJAX request. The endpoint returns immediately with a job_id and
    'queued' status. The new version will appear in the versions list
    once processing completes — users can refresh manually.

    Accepts either multipart form upload or JSON with filePath.
    Set ?sync=1 to keep the legacy synchronous behaviour (small docs only).
    """
    import threading
    import uuid as _uuid

    file_path = None

    if request.content_type and "multipart" in request.content_type:
        uploaded = request.files.get("file")
        if not uploaded:
            return jsonify({"error": "No file provided"}), 400
        upload_dir = os.path.join(
            os.getenv("APP_ROOT", "."), "data", "compliance_uploads"
        )
        os.makedirs(upload_dir, exist_ok=True)
        # Unique filename to prevent collisions when multiple imports run concurrently
        unique_prefix = _uuid.uuid4().hex[:8]
        # Sanitise the user-supplied filename before joining it to the
        # upload directory. Without this, a malicious filename like
        # "../../../etc/passwd" would escape upload_dir via os.path.join.
        # secure_filename() strips path components, replaces unsafe chars
        # with underscores, and returns an empty string for filenames
        # made entirely of unsafe chars — fall back to "upload.bin" in
        # that case so the uuid-prefix still gives us a unique, writable
        # filename. Legitimate filenames like "Q1_Compliance.pdf" pass
        # through unchanged.
        safe_name = secure_filename(uploaded.filename or "") or "upload.bin"
        file_path = os.path.join(upload_dir, f"{unique_prefix}_{safe_name}")
        uploaded.save(file_path)
        excel_template = request.form.get("excel_template_path")
    else:
        data = request.get_json() or {}
        file_path = data.get("filePath") or data.get("file_path")
        excel_template = data.get("excel_template_path")

    if not file_path or not os.path.isfile(file_path):
        return jsonify({"error": "Valid file path is required"}), 400

    # Validate set up-front (fast, sync) so obvious errors don't get queued
    set_info = _engine.get_document_set(set_id)
    if not set_info:
        return jsonify({"error": "Document set not found"}), 404
    retailer_id = set_info["retailer_id"]
    agent_id = set_info.get("agent_id")
    user_id = _get_user_id()

    # Sync mode opt-in for small docs / API callers that want the result inline
    if request.args.get("sync") in ("1", "true"):
        result = _engine.process_compliance_document(
            file_path=file_path,
            retailer_id=retailer_id,
            set_id=set_id,
            uploaded_by=user_id,
            agent_id=agent_id,
            excel_template_path=excel_template,
        )

        if result.error and not result.is_duplicate:
            return jsonify({"error": result.error}), 500

        response = {
            "version_id": result.version_id,
            "version_number": result.version_number,
            "document_id": result.document_id,
            "requirements_count": len(result.requirements),
            "is_duplicate": result.is_duplicate,
        }
        if result.excel_path:
            response["excel_path"] = result.excel_path
        if result.change_summary:
            response["change_summary"] = result.change_summary
        if result.error:
            response["warning"] = result.error

        status = 200 if not result.is_duplicate else 409
        return jsonify(response), status

    # Default: async — spawn background worker, return immediately
    import compliance_jobs

    filename = os.path.basename(file_path)
    job_id = compliance_jobs.start_job(
        set_id=set_id, filename=filename, retailer_id=retailer_id
    )

    def worker():
        try:
            logger.info(
                f"Compliance import job {job_id} starting "
                f"(set={set_id}, file={filename}, retailer={retailer_id})"
            )
            compliance_jobs.mark_running(job_id, "Extracting requirements")
            result = _engine.process_compliance_document(
                file_path=file_path,
                retailer_id=retailer_id,
                set_id=set_id,
                uploaded_by=user_id,
                agent_id=agent_id,
                excel_template_path=excel_template,
            )
            if result.is_duplicate:
                logger.info(
                    f"Compliance import job {job_id} finished: duplicate "
                    f"(version_id={result.version_id})"
                )
                compliance_jobs.finish_job(
                    job_id, "duplicate",
                    version_id=result.version_id,
                    version_number=result.version_number,
                )
            elif result.error:
                logger.error(
                    f"Compliance import job {job_id} finished with error: {result.error}"
                )
                compliance_jobs.finish_job(
                    job_id, "error", error=result.error
                )
            else:
                logger.info(
                    f"Compliance import job {job_id} finished: success "
                    f"(version_id={result.version_id}, "
                    f"version={result.version_number}, "
                    f"requirements={len(result.requirements)})"
                )
                # If we got 0 requirements, surface the diagnostic info so the
                # user sees something more actionable than a silent zero.
                base_msg = (
                    f"Created version {result.version_number} "
                    f"({len(result.requirements)} requirements)"
                )
                if not result.requirements and result.extraction_diagnostics:
                    d = result.extraction_diagnostics
                    if d.get("fields_requested", 0) == 0:
                        base_msg += " — schema requested 0 fields (check the linked schema)"
                    elif d.get("fields_with_value", 0) == 0:
                        base_msg += (
                            f" — LLM returned null for all "
                            f"{d.get('fields_returned', 0)} fields "
                            f"(source: {d.get('source', '?')})"
                        )
                    elif d.get("error"):
                        base_msg += f" — extraction error: {d['error']}"
                compliance_jobs.finish_job(
                    job_id, "done",
                    version_id=result.version_id,
                    version_number=result.version_number,
                    message=base_msg,
                )
        except Exception as e:
            logger.error(
                f"Compliance import job {job_id} crashed: {e}", exc_info=True
            )
            compliance_jobs.finish_job(job_id, "error", error=str(e))

    threading.Thread(
        target=worker, daemon=True, name=f"compliance-{job_id[:8]}"
    ).start()

    return jsonify({
        "status": "queued",
        "job_id": job_id,
        "filename": filename,
        "set_id": set_id,
        "message": (
            "Document accepted. Processing continues in the background — "
            "safe to close this window. The new version will appear in the "
            "versions list when ready (typically a few minutes for small docs; "
            "up to two hours for very large or complex documents)."
        ),
    }), 202


# ============================================================================
# Admin: backfill helper for pre-fix compliance docs
# ============================================================================

@compliance_bp.route(
    "/api/compliance/admin/share-existing-knowledge", methods=["POST"]
)
@api_key_or_session_required(min_role=2)
def share_existing_compliance_knowledge():
    """One-shot helper: convert all existing user-tagged compliance docs to
    SHARED knowledge (visible to every user with access to the agent).

    Use this once after deploying the visibility fix. Idempotent.
    """
    result = _engine.share_existing_compliance_knowledge()
    return jsonify({"success": True, **result})


@compliance_bp.route(
    "/api/compliance/admin/cleanup-orphaned-knowledge", methods=["POST"]
)
@api_key_or_session_required(min_role=2)
def cleanup_orphaned_compliance_knowledge():
    """One-shot helper: remove orphaned compliance knowledge entries left
    behind by the pre-fix double-ingestion bug.

    Finds AgentKnowledge entries whose document is not linked to any current
    or historical RetailerDocumentVersion and soft-deletes them. Idempotent.
    """
    result = _engine.cleanup_orphaned_compliance_knowledge()
    return jsonify({"success": True, **result})


# ============================================================================
# Import Jobs (in-flight upload tracking)
# ============================================================================

@compliance_bp.route("/api/compliance/sets/<int:set_id>/jobs", methods=["GET"])
@api_key_or_session_required(min_role=2)
def list_jobs_for_set(set_id):
    """Return active and recently-finished import jobs for a document set.
    UI uses this to display in-flight uploads above the version list.
    """
    import compliance_jobs
    return jsonify({"jobs": compliance_jobs.get_active_jobs(set_id=set_id)})


# ============================================================================
# Requirements
# ============================================================================

@compliance_bp.route(
    "/api/compliance/versions/<int:version_id>/requirements", methods=["GET"]
)
@api_key_or_session_required(min_role=2)
def get_requirements(version_id):
    requirements = _engine.get_requirements(version_id)
    return jsonify({"requirements": requirements})


@compliance_bp.route("/api/compliance/versions/<int:version_id>", methods=["DELETE"])
@api_key_or_session_required(min_role=2)
def delete_version(version_id):
    result = _engine.delete_version(version_id)
    if not result.get("deleted"):
        return jsonify({"error": result.get("error", "Version not found")}), 404
    return jsonify({"success": True, **result})


@compliance_bp.route("/api/compliance/sets/<int:set_id>", methods=["DELETE"])
@api_key_or_session_required(min_role=2)
def delete_document_set(set_id):
    ok = _engine.delete_set(set_id)
    if not ok:
        return jsonify({"error": "Document set not found"}), 404
    return jsonify({"success": True})


# ============================================================================
# Comparison
# ============================================================================

@compliance_bp.route("/api/compliance/compare/versions", methods=["POST"])
@api_key_or_session_required(min_role=2)
def compare_versions():
    data = request.get_json() or {}
    version_a = data.get("version_a_id")
    version_b = data.get("version_b_id")
    if not version_a or not version_b:
        return jsonify({"error": "version_a_id and version_b_id are required"}), 400

    result = _comparison.compare_versions(int(version_a), int(version_b))
    if not result:
        return jsonify({"error": "No requirements found for comparison"}), 404

    comparison_id = _comparison.store_comparison(
        comparison_type="version",
        source_a_id=int(version_a),
        source_b_id=int(version_b),
        result=result,
        created_by=_get_user_id(),
    )

    return jsonify(
        {"comparison_id": comparison_id, "result": result}
    )


@compliance_bp.route("/api/compliance/compare/retailers", methods=["POST"])
@api_key_or_session_required(min_role=2)
def compare_retailers():
    data = request.get_json() or {}
    retailer_a = data.get("retailer_a_id")
    retailer_b = data.get("retailer_b_id")
    if not retailer_a or not retailer_b:
        return jsonify(
            {"error": "retailer_a_id and retailer_b_id are required"}
        ), 400

    result = _comparison.compare_retailers(
        int(retailer_a),
        int(retailer_b),
        category_filter=data.get("category"),
    )
    if not result:
        return jsonify({"error": "No requirements found for comparison"}), 404

    comparison_id = _comparison.store_comparison(
        comparison_type="cross_retailer",
        source_a_id=int(retailer_a),
        source_b_id=int(retailer_b),
        result=result,
        created_by=_get_user_id(),
    )

    return jsonify(
        {"comparison_id": comparison_id, "result": result}
    )


@compliance_bp.route(
    "/api/compliance/comparisons/<int:comparison_id>", methods=["GET"]
)
@api_key_or_session_required(min_role=2)
def get_comparison(comparison_id):
    result = _comparison.get_comparison(comparison_id)
    if not result:
        return jsonify({"error": "Comparison not found"}), 404
    if isinstance(result.get("created_at"), datetime):
        result["created_at"] = result["created_at"].isoformat()
    return jsonify(result)


# ============================================================================
# Export
# ============================================================================

@compliance_bp.route(
    "/api/compliance/versions/<int:version_id>/export/excel", methods=["GET"]
)
@api_key_or_session_required(min_role=2)
def export_requirements_excel(version_id):
    requirements = _engine.get_requirements(version_id)
    if not requirements:
        return jsonify({"error": "No requirements found"}), 404

    try:
        import openpyxl
        from io import BytesIO
        from flask import send_file

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requirements"

        headers = [
            "Category",
            "Subcategory",
            "Requirement",
            "Value",
            "Severity",
            "Source Page",
            "Confidence",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for req in requirements:
            ws.append(
                [
                    req.get("category", ""),
                    req.get("subcategory", ""),
                    req.get("requirement_text", ""),
                    req.get("specific_value", ""),
                    req.get("severity", ""),
                    req.get("source_page", ""),
                    req.get("confidence", ""),
                ]
            )

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"requirements_v{version_id}.xlsx",
        )
    except Exception as e:
        logger.error("Excel export failed: %s", e)
        return jsonify({"error": str(e)}), 500


@compliance_bp.route(
    "/api/compliance/comparisons/<int:comparison_id>/export/excel", methods=["GET"]
)
@api_key_or_session_required(min_role=2)
def export_comparison_excel(comparison_id):
    comp = _comparison.get_comparison(comparison_id)
    if not comp or not comp.get("result"):
        return jsonify({"error": "Comparison not found"}), 404

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from io import BytesIO
        from flask import send_file

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparison"

        headers = [
            "Category",
            "Subcategory",
            "Source A",
            "Source B",
            "Change Type",
            "Meaningful",
            "Reason",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", fill_type="solid")

        for detail in comp["result"].get("details", []):
            row_num = ws.max_row + 1
            ws.append(
                [
                    detail.get("category", ""),
                    detail.get("subcategory", ""),
                    detail.get("value_a", "") or "",
                    detail.get("value_b", "") or "",
                    detail.get("change_type", ""),
                    "Yes" if detail.get("is_meaningful") else "No",
                    detail.get("reason", ""),
                ]
            )
            change = detail.get("change_type", "")
            fill = (
                green_fill
                if change == "added"
                else red_fill
                if change == "removed"
                else yellow_fill
                if detail.get("is_meaningful")
                else None
            )
            if fill:
                for cell in ws[row_num]:
                    cell.fill = fill

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"comparison_{comparison_id}.xlsx",
        )
    except Exception as e:
        logger.error("Comparison Excel export failed: %s", e)
        return jsonify({"error": str(e)}), 500


# ============================================================================
# Taxonomy info
# ============================================================================

@compliance_bp.route("/api/compliance/accessible-agents", methods=["GET"])
@api_key_or_session_required(min_role=2)
def list_accessible_agents():
    """Return agents the current user has access to (group-based permissions)."""
    try:
        from DataUtils import select_user_agents_and_tools

        user_id = _get_user_id()
        user_role = (
            current_user.role if hasattr(current_user, "role") else 3
        )
        rows = select_user_agents_and_tools(user_id, user_role) or []

        # rows may be a list of dicts OR a DataFrame depending on caller
        if hasattr(rows, "to_dict"):
            rows = rows.to_dict("records")

        agents = []
        seen = set()
        for row in rows:
            aid = row.get("agent_id")
            if aid is None or aid in seen:
                continue
            seen.add(aid)
            # Schema uses agent_description as the display name (legacy column)
            name = (
                row.get("agent_name")
                or row.get("agent_description")
                or f"Agent #{aid}"
            )
            agents.append({"agent_id": int(aid), "agent_name": name})

        agents.sort(key=lambda a: (a["agent_name"] or "").lower())
        return jsonify({"agents": agents})
    except Exception as e:
        logger.exception("Failed to list accessible agents")
        return jsonify({"agents": [], "error": str(e)}), 500


@compliance_bp.route("/api/compliance/taxonomy", methods=["GET"])
@api_key_or_session_required(min_role=2)
def get_taxonomy():
    from compliance_engine import load_compliance_taxonomy

    taxonomy = load_compliance_taxonomy()
    categories = {}
    for cat_key, cat_val in taxonomy.get("categories", {}).items():
        subcats = list(cat_val.get("subcategories", {}).keys())
        categories[cat_key] = {
            "description": cat_val.get("description", ""),
            "subcategories": subcats,
        }
    return jsonify({"categories": categories})


# ============================================================================
# Schemas — reusable extraction schemas attachable to document sets
# ============================================================================

@compliance_bp.route("/api/compliance/schemas", methods=["GET"])
@api_key_or_session_required(min_role=2)
def list_schemas():
    # Auto-seed default schema from YAML if no schemas exist for this tenant
    from compliance_engine import load_compliance_taxonomy
    try:
        ComplianceEngine.seed_default_schema_if_empty(load_compliance_taxonomy())
    except Exception as e:
        logger.warning("Default schema seed failed: %s", e)

    schemas = ComplianceEngine.list_schemas()
    for s in schemas:
        for k, v in s.items():
            if isinstance(v, datetime):
                s[k] = v.isoformat()
    return jsonify({"schemas": schemas})


@compliance_bp.route("/api/compliance/schemas/<int:schema_id>", methods=["GET"])
@api_key_or_session_required(min_role=2)
def get_schema(schema_id):
    s = ComplianceEngine.get_schema(schema_id)
    if not s:
        return jsonify({"error": "Schema not found"}), 404
    for k, v in list(s.items()):
        if isinstance(v, datetime):
            s[k] = v.isoformat()
    return jsonify(s)


@compliance_bp.route("/api/compliance/schemas", methods=["POST"])
@api_key_or_session_required(min_role=2)
def create_schema():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400
    fields = data.get("fields") or []
    if not isinstance(fields, list):
        return jsonify({"error": "fields must be a JSON array"}), 400
    description = data.get("description") or ""

    created_by = None
    try:
        if current_user and getattr(current_user, "id", None) is not None:
            created_by = int(current_user.id)
    except Exception:
        pass

    schema_id = ComplianceEngine.create_schema(
        name=name,
        fields=fields,
        description=description,
        created_by=created_by,
    )
    return jsonify({"schema_id": schema_id, "name": name}), 201


@compliance_bp.route("/api/compliance/schemas/<int:schema_id>", methods=["PUT"])
@api_key_or_session_required(min_role=2)
def update_schema(schema_id):
    data = request.get_json() or {}
    name = data.get("name")
    if name is not None:
        name = name.strip()
        if not name:
            return jsonify({"error": "name cannot be empty"}), 400
    description = data.get("description")
    fields = data.get("fields")
    if fields is not None and not isinstance(fields, list):
        return jsonify({"error": "fields must be a JSON array"}), 400

    ok = ComplianceEngine.update_schema(
        schema_id=schema_id,
        name=name,
        description=description,
        fields=fields,
    )
    if not ok:
        return jsonify({"error": "Schema not found or no changes"}), 404
    return jsonify({"success": True})


@compliance_bp.route("/api/compliance/schemas/<int:schema_id>", methods=["DELETE"])
@api_key_or_session_required(min_role=2)
def delete_schema(schema_id):
    ok = ComplianceEngine.delete_schema(schema_id)
    if not ok:
        return jsonify({"error": "Schema not found"}), 404
    return jsonify({"success": True})
