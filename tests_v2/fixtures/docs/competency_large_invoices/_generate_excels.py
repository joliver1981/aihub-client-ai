"""Generate 3 complex multi-sheet Excel financial reports for the
competency suite that probes large-spreadsheet numerical Q&A.

   01_financial_report_global_logistics_fy2025.xlsx
       Logistics company P&L + departmental + customer concentration
   02_financial_report_megaretail_fy2025.xlsx
       Retail chain: regional revenue + SKU profitability + store ranking
   03_financial_report_pacific_mfg_fy2025.xlsx
       Manufacturer: COGS detail, inventory, production efficiency

Each report has:
   - Summary sheet with KPIs + anchor fingerprint values
   - 5-7 detail sheets, 50-500 rows each
   - Formulas (SUM, AVG) and cross-sheet references where appropriate
   - Multi-row merged headers on at least one sheet
   - An explicit "Anchors" sheet with ground-truth values

Numbers are deterministic (seeded random) so the test's expected
answers stay stable across runs.

Run:
    "$PY" _generate_excels.py
"""
from __future__ import annotations

import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent

# Styles
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2E4F8A")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
MONEY = '"$"#,##0.00'
INT = "#,##0"
PCT = "0.0%"


def _header_row(ws, row, headers, fill=HDR_FILL):
    for c, val in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = HDR_FONT
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt_money(ws, col, row_from, row_to):
    for r in range(row_from, row_to + 1):
        ws.cell(row=r, column=col).number_format = MONEY


# =========================================================================
# Report 1 — Global Logistics Corp FY2025
# =========================================================================

def build_global_logistics():
    wb = openpyxl.Workbook()
    rng = random.Random(20251115)

    # -- Summary sheet --
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "GLOBAL LOGISTICS CORP — Annual Financial Report FY2025"
    ws["A1"].font = Font(size=18, bold=True, color="2E4F8A")
    ws.merge_cells("A1:E1")
    ws["A3"] = "Headline KPIs"
    ws["A3"].font = Font(size=14, bold=True)

    kpis = [
        ("Total revenue (FY2025)",                 184_350_000.00,  MONEY),
        ("Cost of revenue",                        108_240_000.00,  MONEY),
        ("Gross profit",                            76_110_000.00,  MONEY),
        ("Gross margin",                                 0.4128,    PCT),
        ("Operating expenses",                      54_802_000.00,  MONEY),
        ("Operating income",                        21_308_000.00,  MONEY),
        ("Operating margin",                              0.1156,   PCT),
        ("Net income",                              15_847_500.00,  MONEY),
        ("Net margin",                                    0.0860,   PCT),
        ("Total employees (year-end)",                       1_842, INT),
        ("Total shipments handled (millions)",                42.3, "#,##0.0"),
        ("Largest customer (Vellichor Industries) % of rev", 0.118, PCT),
    ]
    _header_row(ws, 5, ["Metric", "Value"])
    for i, (k, v, fmt) in enumerate(kpis, 6):
        ws.cell(row=i, column=1, value=k).border = BORDER
        cell = ws.cell(row=i, column=2, value=v)
        cell.number_format = fmt
        cell.border = BORDER
    _set_widths(ws, [44, 18])

    # -- Quarterly revenue + COGS + opex --
    qws = wb.create_sheet("Quarterly P&L")
    qws["A1"] = "Quarterly P&L — FY2025 (USD)"
    qws["A1"].font = Font(size=14, bold=True)
    qws.merge_cells("A1:F1")

    _header_row(qws, 3, ["Line item", "Q1 2025", "Q2 2025",
                          "Q3 2025", "Q4 2025", "FY2025"])
    lines = [
        ("Express revenue",         11_240_000, 11_995_000, 12_710_000, 13_420_000),
        ("Ground revenue",          18_510_000, 19_140_000, 20_280_000, 21_440_000),
        ("Freight revenue",          9_625_000, 10_120_000, 10_770_000, 11_350_000),
        ("Customs brokerage",        2_710_000,  2_870_000,  3_010_000,  3_160_000),
    ]
    row = 4
    for name, q1, q2, q3, q4 in lines:
        qws.cell(row=row, column=1, value=name).border = BORDER
        for c, v in enumerate([q1, q2, q3, q4], 2):
            cell = qws.cell(row=row, column=c, value=v)
            cell.number_format = MONEY
            cell.border = BORDER
        # FY total formula
        cell = qws.cell(
            row=row, column=6,
            value=f"=SUM(B{row}:E{row})",
        )
        cell.number_format = MONEY
        cell.border = BORDER
        row += 1

    # Total revenue row
    qws.cell(row=row, column=1, value="TOTAL REVENUE").font = Font(bold=True)
    qws.cell(row=row, column=1).fill = SUB_FILL
    for c in range(2, 7):
        cell = qws.cell(
            row=row, column=c,
            value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}{row-1})",
        )
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
        cell.number_format = MONEY
        cell.border = BORDER
    revenue_total_row = row
    row += 2

    cost_lines = [
        ("Driver / pilot wages",    -14_200_000, -14_350_000, -14_780_000, -15_010_000),
        ("Fuel",                    -10_120_000, -10_510_000, -10_980_000, -11_240_000),
        ("Equipment maintenance",    -3_640_000,  -3_780_000,  -3_910_000,  -4_050_000),
        ("Sub-contracted carriers",  -1_840_000,  -1_910_000,  -2_010_000,  -2_120_000),
    ]
    for name, q1, q2, q3, q4 in cost_lines:
        qws.cell(row=row, column=1, value=name).border = BORDER
        for c, v in enumerate([q1, q2, q3, q4], 2):
            cell = qws.cell(row=row, column=c, value=v)
            cell.number_format = MONEY
            cell.border = BORDER
        cell = qws.cell(
            row=row, column=6,
            value=f"=SUM(B{row}:E{row})",
        )
        cell.number_format = MONEY
        cell.border = BORDER
        row += 1

    # OPEX rows
    opex_lines = [
        ("Sales & marketing",        -4_120_000,  -4_280_000,  -4_410_000,  -4_550_000),
        ("R&D / technology",         -2_120_000,  -2_245_000,  -2_390_000,  -2_510_000),
        ("G&A (HQ + finance + HR)",  -3_820_000,  -3_960_000,  -4_080_000,  -4_210_000),
        ("Insurance & claims",       -1_540_000,  -1_590_000,  -1_660_000,  -1_710_000),
        ("Depreciation",             -1_960_000,  -1_960_000,  -1_960_000,  -1_960_000),
    ]
    for name, q1, q2, q3, q4 in opex_lines:
        qws.cell(row=row, column=1, value=name).border = BORDER
        for c, v in enumerate([q1, q2, q3, q4], 2):
            cell = qws.cell(row=row, column=c, value=v)
            cell.number_format = MONEY
            cell.border = BORDER
        cell = qws.cell(
            row=row, column=6,
            value=f"=SUM(B{row}:E{row})",
        )
        cell.number_format = MONEY
        cell.border = BORDER
        row += 1

    _set_widths(qws, [28, 14, 14, 14, 14, 16])

    # -- Customer concentration --
    cws = wb.create_sheet("Customer Concentration")
    cws["A1"] = "Top 25 customers by FY2025 revenue contribution"
    cws["A1"].font = Font(size=14, bold=True)
    cws.merge_cells("A1:E1")
    _header_row(cws, 3, ["Rank", "Customer", "FY2025 revenue", "% of total",
                          "Tenure (yrs)"])
    cust_names = [
        "Vellichor Industries",       "Bramble & Forest Co.",
        "PolarKraft Packaging GmbH",  "Sundial Foods Cooperative",
        "Tessuto Holdings",           "Greenline Distributors",
        "Cypress Container Corp",     "Mira Pharmaceuticals AG",
        "Northstar Components",       "Westbrook & Vale LLP",
        "Eastgate Supplies",          "Halcyon Designs",
        "Bromley Apparel",            "Sterling Components",
        "Sycamore Foods",             "Maplewood Trading",
        "Quartz Imports",             "Pelham Engineering",
        "Cogswell Industries",        "Ravenscroft Holdings",
        "Atlas Mercantile",           "Sentinel-X Systems",
        "Aurora Bioplastics, Inc.",   "Hyperion Defense Group",
        "Tidewater Marine Logistics",
    ]
    # Revenue distribution — top one is 11.8%, decays
    pcts = [0.118, 0.082, 0.063, 0.054, 0.048, 0.041, 0.038, 0.034, 0.030,
            0.028, 0.025, 0.023, 0.021, 0.020, 0.019, 0.018, 0.017, 0.016,
            0.015, 0.014, 0.013, 0.012, 0.011, 0.010, 0.009]
    total_rev = 184_350_000.0
    for i, (name, pct) in enumerate(zip(cust_names, pcts), 4):
        cws.cell(row=i, column=1, value=i - 3).border = BORDER
        cws.cell(row=i, column=2, value=name).border = BORDER
        cell = cws.cell(row=i, column=3, value=round(total_rev * pct, 2))
        cell.number_format = MONEY
        cell.border = BORDER
        cell = cws.cell(row=i, column=4, value=pct)
        cell.number_format = PCT
        cell.border = BORDER
        cws.cell(row=i, column=5, value=rng.randint(2, 14)).border = BORDER
    _set_widths(cws, [6, 32, 18, 12, 14])

    # -- Departmental headcount + spend --
    dws = wb.create_sheet("Departmental Spend")
    dws["A1"] = "Departmental headcount and spend — FY2025"
    dws["A1"].font = Font(size=14, bold=True)
    dws.merge_cells("A1:E1")
    _header_row(dws, 3, ["Department", "Headcount", "Total comp ($)",
                          "Other opex ($)", "Total spend ($)"])
    depts = [
        ("Operations / Network",   1124, 84_300_000,  6_120_000),
        ("Customer Service",         283, 18_410_000,  1_840_000),
        ("Sales & Account Mgmt",     142, 17_750_000,  2_310_000),
        ("Engineering & Platform",    87, 14_200_000,  1_490_000),
        ("Finance & Accounting",      62,  8_540_000,    920_000),
        ("Legal & Compliance",        18,  4_120_000,    285_000),
        ("Marketing & Brand",         28,  4_950_000,  3_840_000),
        ("HR & People Operations",    41,  4_780_000,    540_000),
        ("Executive / G&A",           57, 15_840_000,  1_180_000),
    ]
    row = 4
    for dept, hc, comp, opex in depts:
        dws.cell(row=row, column=1, value=dept).border = BORDER
        dws.cell(row=row, column=2, value=hc).border = BORDER
        cell = dws.cell(row=row, column=3, value=comp)
        cell.number_format = MONEY
        cell.border = BORDER
        cell = dws.cell(row=row, column=4, value=opex)
        cell.number_format = MONEY
        cell.border = BORDER
        cell = dws.cell(row=row, column=5, value=f"=C{row}+D{row}")
        cell.number_format = MONEY
        cell.border = BORDER
        row += 1
    dws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, 6):
        cell = dws.cell(row=row, column=c,
                         value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}{row-1})")
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
        if c >= 3:
            cell.number_format = MONEY
        else:
            cell.number_format = INT
        cell.border = BORDER
    _set_widths(dws, [30, 12, 16, 16, 16])

    # -- Anchors sheet (ground truth) --
    aws = wb.create_sheet("Anchors")
    aws.append(["Fact", "Value"])
    for r in aws.iter_rows(min_row=1, max_row=1):
        for c in r: c.font = HDR_FONT; c.fill = HDR_FILL
    anchors = [
        ("Total revenue FY2025",                "$184,350,000.00"),
        ("Net income FY2025",                   "$15,847,500.00"),
        ("Operating margin FY2025",             "11.56%"),
        ("Total employees year-end",            "1,842"),
        ("Largest customer (rank 1)",           "Vellichor Industries — 11.8% of revenue"),
        ("Total customer concentration top-5",  "36.5%"),
        ("Largest department by headcount",     "Operations / Network — 1,124 employees"),
        ("Highest-spend department",            "Operations / Network — $90.42M total spend"),
        ("Quarter with highest revenue",        "Q4 2025"),
        ("Quarter with lowest revenue",         "Q1 2025"),
        ("Annual fuel cost",                    "$42,850,000 (sum of quarterly)"),
    ]
    for fact, val in anchors:
        aws.append([fact, val])
    _set_widths(aws, [42, 50])

    out = OUT_DIR / "01_financial_report_global_logistics_fy2025.xlsx"
    wb.save(out)
    print(f"Wrote {out}  ({out.stat().st_size:,} bytes)")


# =========================================================================
# Report 2 — Mega Retail Inc FY2025
# =========================================================================

def build_megaretail():
    wb = openpyxl.Workbook()
    rng = random.Random(20251116)

    # -- Summary --
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "MEGA RETAIL INC — Annual Financial Report FY2025"
    ws["A1"].font = Font(size=18, bold=True, color="2E4F8A")
    ws.merge_cells("A1:E1")

    kpis = [
        ("Total revenue (FY2025)",           412_900_000.00,  MONEY),
        ("E-commerce revenue",               228_440_000.00,  MONEY),
        ("In-store revenue",                 184_460_000.00,  MONEY),
        ("Cost of goods sold",              -247_740_000.00,  MONEY),
        ("Gross profit",                     165_160_000.00,  MONEY),
        ("Gross margin",                            0.4000,    PCT),
        ("Total store count (year-end)",                284,   INT),
        ("New stores opened in FY2025",                  22,   INT),
        ("Stores closed in FY2025",                       7,   INT),
        ("Active SKUs",                              12_485,  INT),
        ("Loyalty program members (millions)",          18.7, "#,##0.0"),
        ("Average basket size",                          74.82, MONEY),
        ("Top region (Southwest US) % of revenue",     0.224,   PCT),
    ]
    _header_row(ws, 3, ["Metric", "Value"])
    for i, (k, v, fmt) in enumerate(kpis, 4):
        ws.cell(row=i, column=1, value=k).border = BORDER
        cell = ws.cell(row=i, column=2, value=v)
        cell.number_format = fmt
        cell.border = BORDER
    _set_widths(ws, [46, 18])

    # -- Regional revenue (multi-row merged headers) --
    rws = wb.create_sheet("Regional Revenue")
    rws["A1"] = "Regional revenue by channel — FY2025 ($ thousands)"
    rws["A1"].font = Font(size=14, bold=True)
    rws.merge_cells("A1:G1")

    # Two-row header
    rws["A3"] = "Region"; rws.merge_cells("A3:A4")
    rws["B3"] = "E-commerce"; rws.merge_cells("B3:D3")
    rws["B4"] = "Direct"
    rws["C4"] = "Marketplace"
    rws["D4"] = "Subtotal"
    rws["E3"] = "In-store"; rws.merge_cells("E3:F3")
    rws["E4"] = "Retail"
    rws["F4"] = "Outlet"
    rws["G3"] = "TOTAL"; rws.merge_cells("G3:G4")

    for row in (3, 4):
        for col in range(1, 8):
            c = rws.cell(row=row, column=col)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

    region_rows = [
        ("Northeast US",     38_410, 18_220,   None, 24_510,  6_120, None),
        ("Mid-Atlantic US",  31_220, 15_410,   None, 22_180,  5_840, None),
        ("Southeast US",     35_700, 17_620,   None, 27_320,  7_120, None),
        ("Southwest US",     48_530, 24_280,   None, 23_840,  5_840, None),
        ("Midwest US",       28_140, 14_120,   None, 18_710,  4_980, None),
        ("West Coast US",    42_180, 21_500,   None, 16_410,  3_960, None),
        ("Canada (toy)",      6_240,  3_120,   None,  4_820,  1_240, None),
    ]
    for i, (region, direct, marketplace, _, retail, outlet, _t) in enumerate(region_rows, 5):
        rws.cell(row=i, column=1, value=region).border = BORDER
        rws.cell(row=i, column=2, value=direct).number_format = INT
        rws.cell(row=i, column=2).border = BORDER
        rws.cell(row=i, column=3, value=marketplace).number_format = INT
        rws.cell(row=i, column=3).border = BORDER
        # Subtotal = direct + marketplace
        cell = rws.cell(row=i, column=4, value=f"=B{i}+C{i}")
        cell.number_format = INT; cell.border = BORDER
        rws.cell(row=i, column=5, value=retail).number_format = INT
        rws.cell(row=i, column=5).border = BORDER
        rws.cell(row=i, column=6, value=outlet).number_format = INT
        rws.cell(row=i, column=6).border = BORDER
        # TOTAL row formula
        cell = rws.cell(row=i, column=7, value=f"=D{i}+E{i}+F{i}")
        cell.number_format = INT; cell.border = BORDER
    _set_widths(rws, [18, 12, 14, 12, 12, 12, 12])

    # -- Store ranking (large, 284 rows) --
    sws = wb.create_sheet("Store Ranking")
    sws["A1"] = "Stores by FY2025 revenue contribution"
    sws["A1"].font = Font(size=14, bold=True)
    sws.merge_cells("A1:F1")
    _header_row(sws, 3, ["Rank", "Store ID", "Location", "Region",
                          "FY2025 revenue ($)", "% of in-store rev"])
    # 284 stores; revenue dist-then-scaled
    regions = ["Northeast US", "Mid-Atlantic US", "Southeast US",
               "Southwest US", "Midwest US", "West Coast US",
               "Canada (toy)"]
    cities = ["New York", "Boston", "Philadelphia", "Atlanta", "Miami",
              "Houston", "Dallas", "Phoenix", "Los Angeles",
              "San Francisco", "Seattle", "Chicago", "Denver",
              "Minneapolis", "Toronto", "Vancouver", "Orlando", "Tampa"]
    in_store_total = 184_460_000.0
    weights = []
    for i in range(284):
        # Power-law-ish distribution
        weights.append(1.0 / (i + 8))
    s_total = sum(weights)
    weights = [w / s_total for w in weights]
    rng.shuffle(weights)
    weights.sort(reverse=True)
    for rank in range(284):
        store_id = f"S-{rank+1001:04d}"
        city = rng.choice(cities)
        region = rng.choice(regions)
        rev = round(in_store_total * weights[rank], 2)
        pct = weights[rank]
        sws.cell(row=rank + 4, column=1, value=rank + 1).border = BORDER
        sws.cell(row=rank + 4, column=2, value=store_id).border = BORDER
        sws.cell(row=rank + 4, column=3, value=city).border = BORDER
        sws.cell(row=rank + 4, column=4, value=region).border = BORDER
        cell = sws.cell(row=rank + 4, column=5, value=rev)
        cell.number_format = MONEY; cell.border = BORDER
        cell = sws.cell(row=rank + 4, column=6, value=pct)
        cell.number_format = PCT; cell.border = BORDER
    _set_widths(sws, [6, 12, 16, 18, 18, 16])

    # -- SKU profitability sample --
    sk = wb.create_sheet("Top SKUs")
    sk["A1"] = "Top 50 SKUs by FY2025 contribution margin"
    sk["A1"].font = Font(size=14, bold=True)
    _header_row(sk, 3, ["Rank", "SKU", "Category", "Units sold",
                         "Revenue ($)", "COGS ($)", "Margin ($)",
                         "Margin %"])
    skus = []
    cats = ["Apparel", "Electronics", "Home", "Beauty", "Toys",
            "Outdoor", "Pantry", "Wellness", "Auto", "Pet"]
    for i in range(50):
        sku_id = f"SKU-{30000 + i*7:05d}"
        cat = rng.choice(cats)
        units = rng.randint(20_000, 480_000)
        unit_price = round(rng.uniform(8.50, 240.00), 2)
        rev = round(units * unit_price, 2)
        cogs = round(rev * rng.uniform(0.45, 0.78), 2)
        margin = round(rev - cogs, 2)
        skus.append((sku_id, cat, units, rev, cogs, margin))
    # Sort by margin
    skus.sort(key=lambda x: -x[5])
    for rank, (sku_id, cat, units, rev, cogs, margin) in enumerate(skus, 1):
        sk.cell(row=rank + 3, column=1, value=rank).border = BORDER
        sk.cell(row=rank + 3, column=2, value=sku_id).border = BORDER
        sk.cell(row=rank + 3, column=3, value=cat).border = BORDER
        sk.cell(row=rank + 3, column=4, value=units).border = BORDER
        sk.cell(row=rank + 3, column=4).number_format = INT
        cell = sk.cell(row=rank + 3, column=5, value=rev)
        cell.number_format = MONEY; cell.border = BORDER
        cell = sk.cell(row=rank + 3, column=6, value=cogs)
        cell.number_format = MONEY; cell.border = BORDER
        cell = sk.cell(row=rank + 3, column=7, value=f"=E{rank+3}-F{rank+3}")
        cell.number_format = MONEY; cell.border = BORDER
        cell = sk.cell(row=rank + 3, column=8, value=f"=G{rank+3}/E{rank+3}")
        cell.number_format = PCT; cell.border = BORDER
    _set_widths(sk, [6, 14, 14, 12, 16, 14, 14, 10])

    # -- Anchors --
    aws = wb.create_sheet("Anchors")
    aws.append(["Fact", "Value"])
    for r in aws.iter_rows(min_row=1, max_row=1):
        for c in r: c.font = HDR_FONT; c.fill = HDR_FILL
    anchors = [
        ("Total revenue FY2025",                "$412,900,000.00"),
        ("E-commerce revenue",                   "$228,440,000.00"),
        ("In-store revenue",                     "$184,460,000.00"),
        ("Gross margin",                         "40.0%"),
        ("Total stores year-end",                "284"),
        ("Active SKUs",                          "12,485"),
        ("Loyalty members (M)",                  "18.7"),
        ("Top region by revenue",                "Southwest US — 22.4% of revenue"),
        ("Number of regions tracked",            "7"),
        ("New stores opened FY2025",             "22"),
        ("Average basket size",                  "$74.82"),
    ]
    for f, v in anchors:
        aws.append([f, v])
    _set_widths(aws, [42, 50])

    out = OUT_DIR / "02_financial_report_megaretail_fy2025.xlsx"
    wb.save(out)
    print(f"Wrote {out}  ({out.stat().st_size:,} bytes)")


# =========================================================================
# Report 3 — Pacific Manufacturing FY2025
# =========================================================================

def build_pacific_mfg():
    wb = openpyxl.Workbook()
    rng = random.Random(20251117)

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "PACIFIC MANUFACTURING LLC — Annual Financial Report FY2025"
    ws["A1"].font = Font(size=18, bold=True, color="2E4F8A")
    ws.merge_cells("A1:E1")

    kpis = [
        ("Total revenue (FY2025)",           248_700_000.00,  MONEY),
        ("Cost of goods sold",              -181_590_000.00,  MONEY),
        ("Gross profit",                      67_110_000.00,  MONEY),
        ("Gross margin",                            0.2697,    PCT),
        ("SG&A",                             -31_840_000.00,  MONEY),
        ("R&D",                              -12_440_000.00,  MONEY),
        ("Operating income",                  22_830_000.00,  MONEY),
        ("EBITDA",                            34_180_000.00,  MONEY),
        ("Total units produced (millions)",          18.4,    "#,##0.0"),
        ("Total facilities",                              7,   INT),
        ("Total production employees",               1_482,    INT),
        ("Inventory turns (FY2025)",                   6.8,    "#,##0.0"),
        ("Largest plant (Tacoma) % of output",       0.282,    PCT),
    ]
    _header_row(ws, 3, ["Metric", "Value"])
    for i, (k, v, fmt) in enumerate(kpis, 4):
        ws.cell(row=i, column=1, value=k).border = BORDER
        cell = ws.cell(row=i, column=2, value=v)
        cell.number_format = fmt
        cell.border = BORDER
    _set_widths(ws, [46, 18])

    # -- COGS Detail --
    cws = wb.create_sheet("COGS Detail")
    cws["A1"] = "Cost of goods sold — components, FY2025"
    cws["A1"].font = Font(size=14, bold=True)
    cws.merge_cells("A1:F1")
    _header_row(cws, 3, ["Component", "Q1 2025", "Q2 2025",
                          "Q3 2025", "Q4 2025", "FY2025"])
    cogs_rows = [
        ("Raw materials (steel)",       21_400_000, 21_780_000, 22_120_000, 22_580_000),
        ("Raw materials (polymer)",      8_240_000,  8_410_000,  8_620_000,  8_840_000),
        ("Direct labor",                14_200_000, 14_350_000, 14_580_000, 14_910_000),
        ("Plant utilities",              2_840_000,  2_920_000,  3_010_000,  3_080_000),
        ("Subcontracted machining",      3_220_000,  3_330_000,  3_440_000,  3_560_000),
        ("Freight inbound",              1_120_000,  1_180_000,  1_220_000,  1_260_000),
        ("Quality / scrap",              1_840_000,  1_920_000,  2_010_000,  2_080_000),
        ("Indirect manufacturing",       4_280_000,  4_410_000,  4_540_000,  4_680_000),
    ]
    row = 4
    for name, q1, q2, q3, q4 in cogs_rows:
        cws.cell(row=row, column=1, value=name).border = BORDER
        for c, v in enumerate([q1, q2, q3, q4], 2):
            cell = cws.cell(row=row, column=c, value=v)
            cell.number_format = MONEY; cell.border = BORDER
        cell = cws.cell(row=row, column=6, value=f"=SUM(B{row}:E{row})")
        cell.number_format = MONEY; cell.border = BORDER
        row += 1
    cws.cell(row=row, column=1, value="TOTAL COGS").font = Font(bold=True)
    for c in range(2, 7):
        cell = cws.cell(row=row, column=c,
                         value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}{row-1})")
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
        cell.number_format = MONEY
        cell.border = BORDER
    _set_widths(cws, [28, 14, 14, 14, 14, 16])

    # -- Production by plant --
    pws = wb.create_sheet("Production by Plant")
    pws["A1"] = "Plant production output and efficiency FY2025"
    pws["A1"].font = Font(size=14, bold=True)
    pws.merge_cells("A1:G1")
    _header_row(pws, 3, ["Plant", "City", "Units produced (M)",
                          "Yield %", "Headcount", "Output / FTE",
                          "Energy intensity (kWh/unit)"])
    plants = [
        ("Tacoma WA",      "Tacoma",        5.19, 0.964, 412, None, 0.42),
        ("Fresno CA",      "Fresno",        3.42, 0.951, 318, None, 0.38),
        ("Phoenix AZ",     "Phoenix",       2.78, 0.948, 248, None, 0.36),
        ("Wichita KS",     "Wichita",       2.21, 0.957, 192, None, 0.45),
        ("Toledo OH",      "Toledo",        1.84, 0.962, 158, None, 0.41),
        ("Chattanooga TN", "Chattanooga",   1.62, 0.949, 102, None, 0.39),
        ("Greenville SC",  "Greenville",    1.34, 0.953,  52, None, 0.37),
    ]
    for i, (plant, city, units_m, yield_, hc, _, energy) in enumerate(plants, 4):
        pws.cell(row=i, column=1, value=plant).border = BORDER
        pws.cell(row=i, column=2, value=city).border = BORDER
        cell = pws.cell(row=i, column=3, value=units_m)
        cell.number_format = "#,##0.00"; cell.border = BORDER
        cell = pws.cell(row=i, column=4, value=yield_)
        cell.number_format = PCT; cell.border = BORDER
        pws.cell(row=i, column=5, value=hc).border = BORDER
        # Output per FTE
        cell = pws.cell(row=i, column=6, value=f"=(C{i}*1000000)/E{i}")
        cell.number_format = INT; cell.border = BORDER
        cell = pws.cell(row=i, column=7, value=energy)
        cell.number_format = "0.00"; cell.border = BORDER
    _set_widths(pws, [18, 14, 16, 8, 12, 14, 24])

    # -- Inventory snapshot --
    iws = wb.create_sheet("Inventory Snapshot")
    iws["A1"] = "Inventory levels by SKU (snapshot at Dec 31, 2025)"
    iws["A1"].font = Font(size=14, bold=True)
    iws.merge_cells("A1:F1")
    _header_row(iws, 3, ["SKU", "Category", "Unit cost ($)",
                          "Units on hand", "Inventory value ($)",
                          "Days of supply"])
    cats = ["Fasteners", "Bearings", "Castings", "Fabricated parts",
            "Subassemblies", "Tools", "Maintenance items"]
    for i in range(120):
        sku = f"PM-{1000 + i*3:05d}"
        cat = rng.choice(cats)
        unit_cost = round(rng.uniform(0.85, 845.00), 2)
        units = rng.randint(120, 18_500)
        days = rng.randint(8, 95)
        iws.cell(row=i + 4, column=1, value=sku).border = BORDER
        iws.cell(row=i + 4, column=2, value=cat).border = BORDER
        cell = iws.cell(row=i + 4, column=3, value=unit_cost)
        cell.number_format = MONEY; cell.border = BORDER
        iws.cell(row=i + 4, column=4, value=units).border = BORDER
        iws.cell(row=i + 4, column=4).number_format = INT
        cell = iws.cell(row=i + 4, column=5,
                        value=f"=C{i+4}*D{i+4}")
        cell.number_format = MONEY; cell.border = BORDER
        iws.cell(row=i + 4, column=6, value=days).border = BORDER
    _set_widths(iws, [14, 18, 14, 14, 18, 14])

    # -- Anchors --
    aws = wb.create_sheet("Anchors")
    aws.append(["Fact", "Value"])
    for r in aws.iter_rows(min_row=1, max_row=1):
        for c in r: c.font = HDR_FONT; c.fill = HDR_FILL
    anchors = [
        ("Total revenue FY2025",            "$248,700,000.00"),
        ("Gross margin",                    "26.97%"),
        ("Operating income",                "$22,830,000.00"),
        ("EBITDA",                          "$34,180,000.00"),
        ("Units produced (M)",              "18.4"),
        ("Total production facilities",     "7"),
        ("Largest plant by output",         "Tacoma WA — 5.19M units (28.2% of total)"),
        ("Most efficient plant by yield",   "Tacoma WA — 96.4% yield"),
        ("Lowest-output plant",             "Greenville SC — 1.34M units"),
        ("Total production employees",      "1,482"),
        ("Inventory turns",                 "6.8"),
        ("Total inventory SKUs tracked",    "120"),
        ("FY2025 raw materials (steel + polymer)", "approx $122M combined"),
    ]
    for f, v in anchors:
        aws.append([f, v])
    _set_widths(aws, [42, 60])

    out = OUT_DIR / "03_financial_report_pacific_mfg_fy2025.xlsx"
    wb.save(out)
    print(f"Wrote {out}  ({out.stat().st_size:,} bytes)")


if __name__ == "__main__":
    build_global_logistics()
    build_megaretail()
    build_pacific_mfg()
