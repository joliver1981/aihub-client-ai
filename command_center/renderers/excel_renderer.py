"""
Command Center — Excel Renderer
==================================
Generates Excel files from data.
"""

import io
import logging
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def render_table_to_excel(
    headers: List[str],
    rows: List[List[Any]],
    sheet_name: str = "Data",
    title: Optional[str] = None,
) -> Optional[bytes]:
    """
    Render tabular data to an Excel file.
    Returns .xlsx bytes or None if openpyxl is not available.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed, Excel generation unavailable")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    start_row = 1

    # Title row
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
        start_row = 3

    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in rows:
            if col_idx <= len(row):
                max_length = max(max_length, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_blocks_to_excel(blocks: List[Dict[str, Any]], filename: str = "export") -> Optional[bytes]:
    """Render multiple content blocks to a multi-sheet Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed")
        return None

    wb = Workbook()
    sheet_idx = 0

    for block in blocks:
        if block.get("type") != "table":
            continue

        if sheet_idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()

        ws.title = (block.get("title") or f"Sheet{sheet_idx + 1}")[:31]  # Excel limit

        headers = block.get("headers", [])
        rows = block.get("rows", [])

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        sheet_idx += 1

    if sheet_idx == 0:
        return None

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
